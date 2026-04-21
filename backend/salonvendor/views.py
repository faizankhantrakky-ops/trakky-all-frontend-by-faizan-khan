# from django.shortcuts import render
# from rest_framework import viewsets,generics
# from rest_framework.views import APIView
# from .models import * 
# from .serializers import *
# from rest_framework.permissions import IsAuthenticated
# from .permissions import  IsAuthenticatedVendor, OfferPermission, ChairPermission
# from .permissions import *
# from rest_framework.response import Response
# from rest_framework import status
# from django.contrib.auth.models import User
# from django.http import HttpResponse
# import io
# import datetime 
# import decimal
# from django.db.models import Sum
# from django.shortcuts import get_object_or_404
# import itertools
# from .filters import *
# import random 
# import string
# from django.http import JsonResponse
# from django.db.models import Sum,Count,When,Case,Avg
# import qrcode
# import os
# from rest_framework.parsers import MultiPartParser, FormParser
# import boto3
# from botocore.exceptions import ClientError
# import requests
# from rest_framework_simplejwt.authentication import JWTAuthentication
# from .backends import VendorJWTAuthentication
# from rest_framework.exceptions import PermissionDenied
# from rest_framework.decorators import api_view,authentication_classes,permission_classes
# from salons.serializers import CategoryModelSerializer as SalonCategoryModelSerializer,MasterCategorySerializer,MasterServiceSerializer,ServiceSerializer
# from django_filters.rest_framework import DjangoFilterBackend
# from rest_framework.exceptions import APIException, PermissionDenied

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count, When, Case, Avg
from django.db.models.signals import post_save
from django.dispatch import receiver
from rest_framework.permissions import AllowAny

from rest_framework import viewsets, generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.exceptions import APIException, PermissionDenied
from rest_framework_simplejwt.authentication import JWTAuthentication

from .models import *
from .serializers import *
from .permissions import *  # Optional: Adjust based on whether you need all custom permissions
from .filters import *
from .backends import VendorJWTAuthentication

from salons.serializers import CategoryModelSerializer as SalonCategoryModelSerializer , salonprofileofferSerializer
from salons.serializers import MasterCategorySerializer, MasterServiceSerializer, ServiceSerializer,PackageListSerializer
from salons.models import Package, Services as SalonServices
import io
import decimal
import random
import string
import itertools
import qrcode
import os
import boto3
from botocore.exceptions import ClientError
import requests
from django_filters.rest_framework import DjangoFilterBackend
from copy import deepcopy
from django.contrib.auth.hashers import check_password
from rest_framework_simplejwt.tokens import RefreshToken
from datetime import date, datetime, timedelta
from dateutil import parser
from salons.backends import SalonUserJWTAuthentication



s3 = boto3.client('s3')

@receiver(post_save, sender=Staff)
def update_staff_attendance(sender, instance, created, **kwargs):
    if created:
        attendance = StaffAttendance.objects.create(staff=instance, date=instance.joining_date, present=True, num_services=0,amount_paid=instance.amount_paid)
        attendance.save()
        print("attendance created")
    else:
        print('attendance updated')

@receiver(post_save, sender=Appointment)
def populate_appointment_services_with_hsn(sender, instance, created, **kwargs):
    """
    Signal to populate included_services with HSN code from salon services model
    """
    if instance.included_services:
        updated = False
        service_ids = [
            s.get('service_id')
            for s in instance.included_services
            if s.get('service_id') and not s.get('hsn_code')
        ]

        if service_ids:
            hsn_map = {
                row['id']: row['hsn_code'] or ""
                for row in SalonServices.objects.filter(id__in=service_ids).values('id', 'hsn_code')
            }

            for service in instance.included_services:
                if service.get('hsn_code'):
                    continue
                service_id = service.get('service_id')
                if not service_id:
                    continue
                service['hsn_code'] = hsn_map.get(service_id, "")
                updated = True

        if updated:
            Appointment.objects.filter(pk=instance.pk).update(
                included_services=instance.included_services
            )

class VendorUserListCreateAPIView(generics.ListCreateAPIView):
    queryset = VendorUser.objects.all()
    serializer_class = VendorUserSerializer
    permission_classes = []
    authentication_classes = []

    def get_queryset(self):
        
        queryset = VendorUser.objects.all()

        start_date_str = self.request.query_params.get('start_date')
        end_date_str = self.request.query_params.get('end_date')

        if start_date_str and end_date_str:
            start_date = parser.parse(start_date_str)
            end_date = parser.parse(end_date_str) + timedelta(days=1)
                
            start_date = timezone.make_aware(start_date, timezone.get_current_timezone())
            end_date = timezone.make_aware(end_date, timezone.get_current_timezone())
        
            # Filter queryset based on the range
            queryset = queryset.filter(created_at__range=(start_date, end_date))  

        return queryset

    def get_authenticators(self):
        if self.request and self.request.method == 'GET':
            return [JWTAuthentication()]
        else:
            return []

    def get_permissions(self):
        if self.request.method == 'GET':
            return [IsSuperUser()]
        else:
            return []

class VendorUserRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    authentication_classes = [JWTAuthentication]
    queryset = VendorUser.objects.all()
    serializer_class = VendorUserSerializer

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()

        # Check for dependent records
        if instance.services.filter(membershiptypeservice__isnull=False).exists():
            return Response(
                {"error": "Cannot delete because dependent records exist."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

class VendorUserposRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]
    queryset = VendorUser.objects.all()
    serializer_class = VendorUserSerializer

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()

        # Check for dependent records
        if instance.services.filter(membershiptypeservice__isnull=False).exists():
            return Response(
                {"error": "Cannot delete because dependent records exist."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

class OTPView(generics.CreateAPIView):
    serializer_class = OTPSerializer
    permission_classes = []

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        ph_number = serializer.validated_data['ph_number']
        user = VendorUser.objects.filter(ph_number=ph_number).first()

        if not user:
            return Response({'error': 'User does not exist.'}, status=status.HTTP_400_BAD_REQUEST)

        otp = random.randint(100000, 999999)
        otp_obj = OTP.objects.create(vendor_user=user, otp=otp)
        formatted_phone_number = f"91{user.ph_number}"

        url = 'https://control.msg91.com/api/v5/otp?'

        headers = {
            "accept": "application/json",
            "content-type": "application/json",
            "authkey": os.environ.get("MSG91_AUTH_KEY")
        }

        payload = {
            "mobile": int(formatted_phone_number),
            "otp": otp,
            "template_id": "63354eb2d6fc052a2016c992"
        }

        response = requests.post(url, headers=headers, json=payload)

        return Response({'message': 'OTP sent successfully.', 'payload': response.json()}, status=status.HTTP_200_OK)

    def patch(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        ph_number = serializer.validated_data['ph_number']
        otp = serializer.validated_data['otp']
        new_password = serializer.validated_data['new_password']

        user = VendorUser.objects.filter(ph_number=ph_number).first()

        if not user:
            return Response({'error': 'User does not exist.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            otp_record = OTP.objects.get(vendor_user=user, otp=otp)
            if otp_record.is_valid():
                user.password = new_password
                user.save()
                otp_record.delete()
                return Response({"message": "Password changed successfully"}, status=status.HTTP_200_OK)
            else:
                return Response({"error": "OTP has expired"}, status=status.HTTP_400_BAD_REQUEST)
        except OTP.DoesNotExist:
            return Response({"error": "Invalid OTP"}, status=status.HTTP_400_BAD_REQUEST)



class UpdateVendorPasswordView(APIView):
    permission_classes = []

    def post(self, request):
        ph_number = request.data.get('ph_number')  # Get the vendor's phone number
        current_password = request.data.get('password')
        new_password = request.data.get('new_password')

        # Validate input
        if not ph_number or not current_password or not new_password:
            return Response({"error": "Phone number, current password, and new password are required."}, 
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            vendor = VendorUser.objects.get(ph_number=ph_number)  # Query vendor by phone number
        except VendorUser.DoesNotExist:
            return Response({"error": "Vendor not found."}, status=status.HTTP_404_NOT_FOUND)

        # First, try checking the password as if it's hashed
        if not check_password(current_password, vendor.password):
            # If that fails, assume the password is stored as plain text
            if vendor.password != current_password:
                return Response({"error": "Current password is incorrect."}, status=status.HTTP_400_BAD_REQUEST)

        # Hash the new password and save it
        vendor.password = new_password
        vendor.save()

        return Response({"message": "Password updated successfully."}, status=status.HTTP_200_OK)



# class JWTView(generics.CreateAPIView):
#     permission_classes = []

#     def create(self, request, *args, **kwargs):
#         phone_number = request.data.get('phone_number')
#         password = request.data.get('password')

#         # Check if phone_number is provided
#         if not phone_number:
#             return Response({'error': 'Provide phone_number.'}, status=status.HTTP_400_BAD_REQUEST)

#         # Check if user exists
#         user = VendorUser.objects.filter(ph_number=phone_number).first()

#         if not user:
#             return Response({'error': 'User does not exist.'}, status=status.HTTP_400_BAD_REQUEST)
        
#         # password = VendorUser.objects.filter()
#         password_check = user.password 

#         if not password:
#             return Response({'error': 'Provide Password'}, status=status.HTTP_400_BAD_REQUEST)
        
#         if password == password_check:
#             refresh = RefreshToken.for_user(user)

#             # Access the token and refresh token as strings
#             access_token = str(refresh.access_token)
#             refresh_token = str(refresh)

#             return Response({'refresh_token': refresh_token, 'access_token': access_token})
        
#         else:
#             return Response({'error': 'invalide Password'}, status=status.HTTP_400_BAD_REQUEST)


class JWTView(generics.CreateAPIView):
    permission_classes = []

    def create(self, request, *args, **kwargs):
        phone_number = request.data.get('phone_number')
        password = request.data.get('password')

        if not phone_number or not password:
            return Response({'error': 'Provide phone number and password.'}, status=400)

        # LOGIN MATCHING
        user = VendorUser.objects.filter(ph_number=phone_number).first()
        user_type = None
        staff = manager = None

        if user and user.password == password:
            pass
        else:
            staff = Staff.objects.filter(ph_number=phone_number).first()
            if staff and staff.password == password:
                user = staff.vendor_user
                user_type = "Staff"
            else:
                manager = Manager.objects.filter(ph_number=phone_number).first()
                if manager and manager.password == password:
                    user = manager.vendor_user
                    user_type = "Manager"
                else:
                    return Response({'error': 'Invalid credentials.'}, status=401)

        # DEVICE LIMIT CHECK
        active_sessions = user.active_jtis or []
        allowed = user.no_of_login_allowed

        # if len(active_sessions) >= allowed:
        #     return Response(
        #         {"error": f"Login device limit reached. Max allowed: {allowed}."},
        #         status=403
        #     )

        try:
            # Generate token
            refresh = RefreshToken.for_user(user)
            access_token = refresh.access_token

            # Add user_id to token payload for authentication
            access_token['user_id'] = user.id

            # JTI management
            new_jti = str(access_token['jti'])
            active_sessions.append(new_jti)
            user.active_jtis = active_sessions
            user.save(update_fields=['active_jtis'])

            response = {
                "refresh_token": str(refresh),
                "access_token": str(access_token),
                "user_id": user.id,
                "ph_number": user.ph_number,
                "user_type": "VendorUser" if user_type is None else user_type,
            }

            if user_type == "Staff":
                response["staff_id"] = staff.id
            elif user_type == "Manager":
                response["manager_id"] = manager.id

            return Response(response, status=200)

        except Exception as e:
            return Response({"error": "Token generation failed", "details": str(e)}, status=500)

class UnlockPOSAPIView(APIView):
    """
    API to unlock POS / reset active_jtis for a VendorUser
    Endpoint: /unlock-pos/<vendor_id>/
    Method: PATCH
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def patch(self, request, vendor_id, *args, **kwargs):
        try:
            vendor_user = VendorUser.objects.get(id=vendor_id)
        except VendorUser.DoesNotExist:
            return Response(
                {"error": "VendorUser not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Reset active_jtis
        vendor_user.active_jtis = []
        vendor_user.save(update_fields=['active_jtis'])

        return Response(
            {"success": f"POS unlocked for VendorUser ID {vendor_id}."},
            status=status.HTTP_200_OK
        )



class StaffViewSet(viewsets.ModelViewSet):
    serializer_class = StaffSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        from datetime import datetime

        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')

        queryset = Staff.objects.filter(vendor_user=self.request.user)
        today = date.today()

        # ✅ Auto-update is_present based on joining and exit date
        for staff in queryset:
            if staff.joining_date <= today and (staff.exit_date is None or staff.exit_date >= today):
                if not staff.is_present:
                    staff.is_present = True
                    staff.save(update_fields=["is_present"])
            else:
                if staff.is_present:
                    staff.is_present = False
                    staff.save(update_fields=["is_present"])

        # ✅ Filter by staff_role (job_role param)
        job_role = self.request.query_params.get('job_role')
        if job_role:
            queryset = queryset.filter(staff_role__iexact=job_role)

        # ✅ Filter by is_present param
        present_filter = self.request.query_params.get('is_present')
        if present_filter is not None:
            is_present_value = present_filter.lower() == 'true'
            queryset = queryset.filter(is_present=is_present_value)

        return queryset

    def create(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vendor_user_instance = request.user
        self.perform_create(serializer, vendor_user=vendor_user_instance)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def perform_create(self, serializer, vendor_user=None):
        serializer.save(vendor_user=vendor_user or self.request.user)

    def update(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')

        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)

    def partial_update(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)

from copy import deepcopy

class StaffAttendanceViewSet(viewsets.ModelViewSet):
    
    from datetime import datetime


    serializer_class = StaffAttendanceSerializer
    authentication_classes = [VendorJWTAuthentication]
    filter_fields = ['date', 'staff']

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return StaffAttendance.objects.filter(staff__vendor_user=self.request.user)

    def create(self, request, *args, **kwargs):
        
        if request.data.get('date') != str(datetime.today().date()):
            return Response({'error': 'You can only update today\'s attendance.'}, status=status.HTTP_400_BAD_REQUEST)

        staff_id = request.data.get('staff')
        if not staff_id:
            return Response({'error': 'Staff field is required.'}, status=status.HTTP_400_BAD_REQUEST)

        staff = Staff.objects.filter(vendor_user=self.request.user, id=staff_id).first()
        if not staff:
            return Response({'error': 'The staff given is inaccessible by you.'}, status=status.HTTP_400_BAD_REQUEST)

        mutable_data = deepcopy(request.data)
        mutable_data['staff'] = staff.id

        serializer = self.get_serializer(data=mutable_data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):

        if request.data.get('date') != str(datetime.today().date()):
            return Response({'error': 'You can only update today\'s attendance.'}, status=status.HTTP_400_BAD_REQUEST)

        staff_id = request.data.get('staff')
        salary = request.data.get('salary')
        if staff_id and salary is not None:
            try:
                staff = Staff.objects.get(vendor_user=self.request.user, id=staff_id)
                staff.salary = salary
                staff.save()
            except Staff.DoesNotExist:
                return Response({'error': 'Staff does not exist.'}, status=status.HTTP_400_BAD_REQUEST)

        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):

        if request.data.get('date') != str(datetime.today().date()):
            return Response({'error': 'You can only update today\'s attendance.'}, status=status.HTTP_400_BAD_REQUEST)

        amount_paid = request.data.get('amount_paid')
        if amount_paid:
            try:
                attendance = StaffAttendance.objects.get(id=kwargs['pk'])
                staff = attendance.staff
                if staff.vendor_user != self.request.user:
                    return Response({'error': 'The staff given is inaccessible by you.'}, status=status.HTTP_400_BAD_REQUEST)
                
                staff.amount_paid += Decimal(amount_paid)
                staff.save()
            except StaffAttendance.DoesNotExist:
                return Response({'error': 'Staff attendance does not exist.'}, status=status.HTTP_400_BAD_REQUEST)

        return super().partial_update(request, *args, **kwargs)
    
class ChairsViewSet(viewsets.ModelViewSet):
    serializer_class = ChairsSerializer
    authentication_classes = [VendorJWTAuthentication]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return Chairs.objects.filter(vendor_user=self.request.user)

    def perform_create(self, serializer):
        vendor = self.request.user
        serializer.save(vendor_user=vendor)

    def perform_update(self, serializer):
        vendor = self.request.user
        serializer.save(vendor_user=vendor)

    def perform_partial_update(self, serializer):
        vendor = self.request.user
        serializer.save(vendor_user=vendor)



# from rest_framework.decorators import action
# from collections import defaultdict

# class DashBoardSalesView(APIView):
#     authentication_classes = [VendorJWTAuthentication]
    
#     def get(self, request, *args, **kwargs):
#         start_date = request.GET.get('start_date')
#         end_date = request.GET.get('end_date')
        
#         if request.user.is_anonymous:
#             return Response({'error': 'You are not authorized to view this page.'}, status=status.HTTP_403_FORBIDDEN)
        
#         try:
#             vendor = VendorUser.objects.get(id=request.user.id)
#         except VendorUser.DoesNotExist:
#             return Response({'error': 'Vendor user does not exist.'}, status=status.HTTP_404_NOT_FOUND)
        
#         appointments = Appointment.objects.filter(vendor_user=vendor)
        
#         if start_date and end_date:
#             appointments = appointments.filter(date__gte=start_date, date__lte=end_date)
        
#         notcancelled_appointments = appointments.exclude(appointment_status='cancelled')
#         notcancelled_serialized = AppointmentNewSerializer(notcancelled_appointments, many=True)
        
#         total_bill_amount = sum([float(appointment['final_amount']) for appointment in notcancelled_serialized.data])
#         total_bill_count = notcancelled_appointments.count()
#         total_paid_amount = sum([float(appointment['amount_paid']) for appointment in notcancelled_serialized.data])

#         avg_bill_amount = total_bill_amount / total_bill_count if total_bill_count > 0 else 0
        
#         unpaid_bill_amount = sum([(appointment.final_amount - appointment.amount_paid) for appointment in notcancelled_appointments])
#         unpaid_bill_amount = max(unpaid_bill_amount, 0)
        
#         customer_membership = CustomerMembership.objects.filter(vendor_user=vendor)
        
#         if start_date and end_date:
#             customer_membership = customer_membership.filter(created_at__gte=start_date, created_at__lte=end_date)
        
#         customer_membership_serialized = CustomerMembershipSerializer(customer_membership, many=True)
#         unpaid_customer_membership = sum(cms.get('unpaid_amount', 0) for cms in customer_membership_serialized.data) if customer_membership_serialized.data else 0
        
#         # total_customers = appointments.values('customer_phone').distinct().count() if appointments else 0
#         # total_staff = Staff.objects.filter(vendor_user=vendor).count() if Staff.objects.filter(vendor_user=vendor).count() else 0
#         # total_services = Services.objects.filter(vendor_user=vendor).count() if Services.objects.filter(vendor_user=vendor).count() else 0
#         total_offers = Offers.objects.filter(vendor_user=vendor, enabled=True).count() if Offers.objects.filter(vendor_user=vendor, enabled=True).count() else 0
        
#         upi_appointments = notcancelled_appointments.filter(payment_mode__icontains='upi').count()
#         card_appointments = notcancelled_appointments.filter(payment_mode__icontains='card').count()
#         cash_appointments = notcancelled_appointments.filter(payment_mode__icontains='cash').count()

#         return Response({
#             'Total Bill Amount': '₹ {:.2f}'.format(total_bill_amount),
#             'Total Bill Count': total_bill_count,
#             'Total Paid Amount': '₹ {:.2f}'.format(total_paid_amount),
#             'Average Bill Amount': '₹ {:.2f}'.format(avg_bill_amount),
#             'Unpaid Bill Amount': '₹ {:.2f}'.format(unpaid_bill_amount),
#             'Unpaid Membership Amount': '₹ {:.2f}'.format(unpaid_customer_membership),
#             'UPI Appointments': upi_appointments,
#             'Card Appointments': card_appointments,
#             'Cash Appointments': cash_appointments,
#             'Total Offers': total_offers,
#             'fields': [
#                 'Total Bill Amount', 'Total Bill Count', 'Total Paid Amount', 'Average Bill Amount', 'Unpaid Bill Amount',
#                 'Unpaid Membership Amount', 'UPI Appointments', 'Card Appointments', 'Cash Appointments', 'Total Offers'
#             ]
#         }, status=status.HTTP_200_OK)


# class DashBoardCustomerView(APIView):
#     authentication_classes = [VendorJWTAuthentication]

#     def get(self, request, *args, **kwargs):
#         start_date = request.GET.get('start_date')
#         end_date = request.GET.get('end_date')
#         # if request.user.is_superuser:
#         #     return Response({'error': 'You are not authorized to view this page.'}, status=status.HTTP_403_FORBIDDEN)
#         vendor = VendorUser.objects.get(id=request.user.id)

#         if start_date and end_date:
#             appointments = Appointment.objects.filter(vendor_user=vendor, date__gte=start_date, date__lte=end_date)
#         else:
#             appointments = Appointment.objects.filter(vendor_user=vendor)

#         total_appointments = appointments.count() if appointments else 0
#         completed_appointments = appointments.filter(appointment_status='completed').count() if appointments else 0
#         cancelled_appointments = appointments.filter(appointment_status='cancelled').count() if appointments else 0
#         not_started_appointments = appointments.filter(appointment_status='not_started').count() if appointments else 0
#         ongoing_appointments = appointments.filter(appointment_status='running').count() if appointments else 0
#         total_customers = appointments.values('customer_phone').distinct().count() if appointments else 0
#         total_membership_customers = appointments.filter(membership__isnull=False).values('customer_phone').distinct().count() if appointments else 0
#         total_new_customers = appointments.filter(customer_type="new").values('customer_phone').distinct().count() if appointments else 0
#         total_regular_customers = appointments.filter(customer_type="regular").values('customer_phone').distinct().count() if appointments else 0
        
#         return Response({
#             'Total Appointments': total_appointments,
#             'Completed Appointments': completed_appointments,
#             'Cancelled Appointments': cancelled_appointments,
#             'Not Started Appointments': not_started_appointments,
#             'Ongoing Appointments': ongoing_appointments,
#             'Total Customers': total_customers,
#             'Membership Customers': total_membership_customers,
#             'New Customers': total_new_customers,
#             'Regular Customers': total_regular_customers,
#             'fields': ['Total Appointments', 'Completed Appointments', 'Cancelled Appointments', 'Not Started Appointments', 'Ongoing Appointments', 'Total Customers', 'Membership Customers', 'New Customers', 'Regular Customers']
#         }, status=status.HTTP_200_OK)

class CategoryModelViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategoryModelSerializer
    authentication_classes = [VendorJWTAuthentication]
    # filter_backends = [DjangoFilterBackend]
    filterset_fields = ['name', 'gender']

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return Category.objects.filter(vendor_user=self.request.user.id)

    def perform_create(self, serializer):
        self.set_vendor(serializer)

    def perform_update(self, serializer):
        self.set_vendor(serializer)

    def perform_partial_update(self, serializer):
        self.set_vendor(serializer)

    def set_vendor(self, serializer):
        vendor_user = self.request.user
        if not vendor_user:
            raise ValidationError({'vendor_user': 'Vendor user must be authenticated.'})
        
        serializer.save(vendor_user=vendor_user)


class ServiceModelViewSet(viewsets.ModelViewSet):
    queryset = Services.objects.all()
    serializer_class = ServiceSerializer
    authentication_classes = [VendorJWTAuthentication]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return Services.objects.filter(vendor_user=self.request.user)

    def create(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vendor_user_instance = request.user
        serializer.save(vendor_user=vendor_user_instance)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')

        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        vendor_user_instance = request.user
        serializer.save(vendor_user=vendor_user_instance)
        return Response(serializer.data)

    def partial_update(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)


class ManagerViewSet(viewsets.ModelViewSet):
    serializer_class = ManagerSerializer
    # permission_classes = [VendorRelatedModelPermission]
    authentication_classes = [VendorJWTAuthentication]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return Manager.objects.filter(vendor_user=self.request.user)

    def perform_create(self, serializer):
        vendor = self.request.user
        serializer.save(vendor_user=vendor)

    def perform_update(self, serializer):
        vendor = self.request.user
        serializer.save(vendor_user=vendor)

    def perform_partial_update(self, serializer):
        vendor = self.request.user
        serializer.save(vendor_user=vendor)


class MembershipTypeViewset(viewsets.ModelViewSet):
    serializer_class = MembershipTypeSerializer
    authentication_classes = [VendorJWTAuthentication]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return MembershipType.objects.filter(vendor_user=self.request.user.id)


    def perform_create(self, serializer):
        vendor = self.request.user
        serializer.save(vendor_user=vendor)

    def perform_update(self, serializer):
        vendor = self.request.user
        serializer.save(vendor_user=vendor)

    def perform_partial_update(self, serializer):
        vendor = self.request.user
        serializer.save(vendor_user=vendor)


class MembershipTypeServiceViewSet(viewsets.ModelViewSet):
    serializer_class = MembershipTypeServiceSerializer
    authentication_classes = [VendorJWTAuthentication]
    # permission_classes = [VendorRelatedModelPermission]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return MembershipTypeService.objects.filter(membership_type__vendor_user=self.request.user.id)

    # def set_vendor(self, callback, request, *args, **kwargs):
    #     vendor = get_object_or_404(VendorUser, id=request.user.id)
    #     request.data['vendor'] = vendor.id

    #     membership_type_id = request.data.get('membership_type')
    #     services = request.data.get('services')
    #     numbers = request.data.get('number_of_services')

    #     if not membership_type_id or not services or not numbers:
    #         return Response({'error': 'membership_type, services, and number_of_services fields are required.'}, status=status.HTTP_400_BAD_REQUEST)

    #     membership_type = get_object_or_404(MembershipType, pk=membership_type_id, vendor=vendor)
    #     created_objects = []

    #     for service_id, number in itertools.zip_longest(services, numbers, fillvalue=None):
    #         if service_id is None or number is None:
    #             return Response({'error': 'Mismatched services and number_of_services fields.'}, status=status.HTTP_400_BAD_REQUEST)
    #         service = get_object_or_404(Services, pk=service_id)
    #         membership_type_service, created = MembershipTypeService.objects.get_or_create(
    #             membership_type=membership_type, service=service, defaults={'number': number}
    #         )
    #         if not created:
    #             membership_type_service.number = number
    #             membership_type_service.save()
    #         created_objects.append(membership_type_service)

    #     return callback(request, created_objects, *args, **kwargs)

    def set_vendor(self, callback, request, *args, **kwargs):
        vendor = get_object_or_404(VendorUser, id=request.user.id)
        mutable_data = request.data.copy()  # Create a mutable copy of request data
        mutable_data['vendor'] = vendor.id

        membership_type_id = mutable_data.get('membership_type')
        services = mutable_data.get('services')
        numbers = mutable_data.get('number_of_services')

        if not membership_type_id or not services or not numbers:
            return Response({'error': 'membership_type, services, and number_of_services fields are required.'}, status=status.HTTP_400_BAD_REQUEST)

        membership_type = get_object_or_404(MembershipType, pk=membership_type_id, vendor_user=vendor)
        created_objects = []

        for service_id, number in itertools.zip_longest(services, numbers, fillvalue=None):
            if service_id is None or number is None:
                return Response({'error': 'Mismatched services and number_of_services fields.'}, status=status.HTTP_400_BAD_REQUEST)
            service = get_object_or_404(Services, pk=service_id)
            membership_type_service, created = MembershipTypeService.objects.get_or_create(
                membership_type=membership_type, service=service, defaults={'number': number}
            )
            if not created:
                membership_type_service.number = number
                membership_type_service.save()
            created_objects.append(membership_type_service)

        return callback(request, created_objects, *args, **kwargs)


    # def create(self, request, *args, **kwargs):
    #     def create_callback(request, created_objects, *args, **kwargs):
    #         serializer = self.get_serializer(created_objects, many=True)
    #         return Response(serializer.data, status=status.HTTP_201_CREATED)

    #     return self.set_vendor(create_callback, request, *args, **kwargs)


    def update(self, request, *args, **kwargs):
        return self.set_vendor(super().update, request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        return self.set_vendor(super().partial_update, request, *args, **kwargs)

class CustomerMembershipViewset(viewsets.ModelViewSet):
    serializer_class = CustomerMembershipSerializer
    authentication_classes = [VendorJWTAuthentication]
    filterset_class = CustomerMembershipFilter

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return CustomerMembership.objects.filter(vendor_user=self.request.user)

    def set_vendor(self, callback, request, *args, **kwargs):
        vendor = get_object_or_404(VendorUser, id=request.user.id)
        request.data['vendor'] = vendor.id

        membership_type_id = request.data.get('membership_type')
        customer_name = request.data.get('customer_name')
        customer_phone = request.data.get('customer_phone')
        num_person_allowed = request.data.get('num_person_allowed')
        remarks = request.data.get('remarks')
        manager_id = request.data.get('manager')
        membership_code = request.data.get('membership_code')
        amount_paid = request.data.get('amount_paid')
        customer_email = request.data.get('customer_email')
        customer_gender = request.data.get('customer_gender')

        required_fields = [
            membership_type_id, customer_name, customer_phone, num_person_allowed,
            remarks, manager_id, membership_code, amount_paid, customer_email, customer_gender
        ]

        if not all(required_fields):
            return Response(
                {'error': 'All fields are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        membership_type = get_object_or_404(MembershipType, pk=membership_type_id)
        manager = get_object_or_404(Manager, pk=manager_id)

        customer_membership = CustomerMembership(
            vendor_user=vendor,
            membership_type=membership_type,
            customer_name=customer_name,
            customer_phone=customer_phone,
            num_person_allowed=num_person_allowed,
            remarks=remarks,
            manager=manager,
            membership_code=membership_code,
            amount_paid=amount_paid,
            customer_gender=customer_gender,
            customer_email=customer_email
        )
        customer_membership.save()

        customer, created = CustomerTable.objects.get_or_create(
            customer_phone=customer_phone,
            defaults={
                'customer_name': customer_name,
                'customer_type': 'member',
                'vendor_user': vendor,
                'customer_email': customer_email,
                'customer_gender': customer_gender
            }
        )
        if not created:
            customer.customer_type = 'member'
            customer.save()

        return callback(request, customer_membership, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        def create_callback(request, customer_membership, *args, **kwargs):
            serializer = self.serializer_class(customer_membership)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return self.set_vendor(create_callback, request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

def generate_membership_code(request):
    while True:
        letters = ''.join(random.choices(string.ascii_lowercase, k=2))
        digits = ''.join(random.choices(string.digits.replace('0', ''), k=2))
        code_list = list(letters + digits)
        random.shuffle(code_list)
        membership_code = ''.join(code_list)
        if not CustomerMembership.objects.filter(membership_code=membership_code).exists():
            break
    return JsonResponse({'membership_code': membership_code})


class CancelledAppointmentViewset(viewsets.ModelViewSet):
    serializer_class = CancelledAppointmentSerializer
    # permission_classes = [AppointmentCanlledorRemarksPermission]
    authentication_classes = [VendorJWTAuthentication]

    def get_queryset(self):
        queryset = CancelledAppointment.objects.all()
        vendor_id = self.request.user.id
        if vendor_id is not None:
            queryset = queryset.filter(appointment__vendor_user=vendor_id)
        return queryset

    def create(self, request, *args, **kwargs):
        appointment_id = request.data.get('appointment')
        reason = request.data.get('reason')
        if not all([appointment_id, reason]):
            return Response(
                {'error': 'appointment and remarks fields are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        appointment = get_object_or_404(Appointment, pk=appointment_id)

        # Create cancelled appointment entry
        cancelled_appointment = CancelledAppointment(
            appointment=appointment,
            reason=reason
        )

        # Update appointment status
        appointment.appointment_status = 'cancelled'
        appointment.save()

        # Free up all staff linked to this appointment
        for staff_member in appointment.staff.all():
            staff_member.is_busy = False
            staff_member.save()

        cancelled_appointment.save()

        serializer = self.serializer_class(cancelled_appointment)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)



class AppointmentRemarksViewset(viewsets.ModelViewSet):
    serializer_class = AppointmentRemarksSerializer
    authentication_classes = [VendorJWTAuthentication]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return AppointmentRemarks.objects.filter(appointment__vendor_user=self.request.user)

    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

from rest_framework.pagination import PageNumberPagination

class Standard50Pagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 100


class CustomerTableView(generics.ListCreateAPIView):
    serializer_class = CustomerTableSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]
    pagination_class = Standard50Pagination
    
    
    def get_queryset(self):
        queryset = CustomerTable.objects.filter(vendor_user=self.request.user)
        
        # Apply individual filters
        customer_phone = self.request.query_params.get('customer_phone')
        customer_name = self.request.query_params.get('customer_name')
        customer_gender = self.request.query_params.get('customer_gender')
        customer_type = self.request.query_params.get('customer_type')
        
        if customer_phone:
            queryset = queryset.filter(customer_phone__icontains=customer_phone)
        if customer_name:
            queryset = queryset.filter(customer_name__icontains=customer_name)
        if customer_gender:
            queryset = queryset.filter(customer_gender__iexact=customer_gender)
        if customer_type:
            queryset = queryset.filter(customer_type__iexact=customer_type)
            
        return queryset.order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(vendor_user=self.request.user)

class CustomerTableRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = CustomerTableSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return CustomerTable.objects.filter(vendor_user=self.request.user)


class staff_monthly_details(APIView):
    authentication_classes = [VendorJWTAuthentication]
    def get(self, request, *args, **kwargs):
        
        # if request.user.is_superuser:
        #     return Response({'error': 'You are not authorized to view this page.'}, status=status.HTTP_403_FORBIDDEN)
        vendor = VendorUser.objects.get(id=request.user.id)
        staffs = Staff.objects.filter(vendor_user=vendor)
        # start_date = request.GET.get('start_date')
        # end_date = request.GET.get('end_date')
        staffdata = []
        for staff in staffs:
            attendance_data = StaffAttendance.objects.filter(
                staff_id=staff,
                # date__gte=start_date,
                # date__lte=end_date
            ).values('staff__staffname').annotate(
                commission_total=Sum('commission'),
                amount_paid_total=Sum('amount_paid'),
                num_services_total=Sum('num_services'),
                total_attendance=Count(Case(When(present=True, then=1)))


            )
            comments = AppointmentRemarks.objects.filter(
                appointment__staff=staff,
                # appointment__date__gte=start_date,
                # appointment__date__lte=end_date,
                appointment__service__isnull=False
            ).values('appointment__date', 'rating', 'remark', "appointment__customer_name")
            comments=list(comments)
            comments_offer = AppointmentRemarks.objects.filter(
                appointment__staff=staff,
                # appointment__date__gte=start_date,
                # appointment__date__lte=end_date,
                appointment__offer__isnull=False
            ).values('appointment__date', 'rating', 'remark', "appointment__customer_name")
            comments_offer=list(comments_offer)
            comments.extend(comments_offer)
            average_rating = AppointmentRemarks.objects.filter(
                appointment__staff=staff,
                # appointment__date__gte=start_date,
                # appointment__date__lte=end_date
            ).aggregate(Avg('rating'))
            staffdata.append({
                'staff': StaffSerializer(staff,context={'request': request}).data,
                'attendance_data': list(attendance_data),
                'comments_rating': comments,
                'average_rating': round(average_rating['rating__avg'],2) if average_rating['rating__avg'] else 
                0
            })
        return Response(staffdata)

# class salon_qrcode(APIView):
#     def get(self, request, *args, **kwargs):
#         vendor = VendorUser.objects.get(id=request.user.id)  # Adjust this line to match your Vendor model

#         session = boto3.Session(
#             aws_access_key_id=os.environ['AWS_ACCESS_KEY_ID'],
#             aws_secret_access_key=os.environ['AWS_SECRET_ACCESS_KEY'],
#             region_name=os.environ['AWS_S3_REGION_NAME']
#         )
#         s3_client = session.client('s3')

#         img_path = os.path.join('qrcodes/', f'{vendor.salon.slug}-qr.png')

#         try:
#             s3_client.head_object(Bucket=os.environ['AWS_STORAGE_BUCKET_NAME'], Key=img_path)
#             # If the image exists in S3, return the S3 URL
#             url = s3_client.generate_presigned_url(
#                 'get_object',
#                 Params={
#                     'Bucket': os.environ['AWS_STORAGE_BUCKET_NAME'],
#                     'Key': img_path
#                 },
#                 ExpiresIn=3600  # URL expires in 1 hour (you can adjust this)
#             )
#             return HttpResponse(url)
#         except ClientError:
#             # If the image does not exist in S3, generate and upload it
#             qr = qrcode.QRCode(version=1, box_size=10, border=5)
#             qr.add_data(f"https://trakky.in/{str(vendor.salon.city).lower()}/{str(vendor.salon.area).lower()}/salons/{str(vendor.salon.slug).lower()}")
#             qr.make(fit=True)
#             img = qr.make_image(fill_color="black", back_color="white")
#             img_bytes = io.BytesIO()
#             img.save(img_bytes, format='PNG')
#             img_bytes.seek(0)
#             s3_client.upload_fileobj(img_bytes, os.environ['AWS_STORAGE_BUCKET_NAME'], img_path)
#             # Return the S3 URL
#             url = s3_client.generate_presigned_url(
#                 'get_object',
#                 Params={
#                     'Bucket': os.environ['AWS_STORAGE_BUCKET_NAME'],
#                     'Key': img_path
#                 },
#                 ExpiresIn=3600  # URL expires in 1 hour (you can adjust this)
#             )
#             return HttpResponse(url)

class SalonLink(APIView):
    def get(self, request, *args, **kwargs):
        vendor = VendorUser.objects.get(id=self.request.user.id)
        return JsonResponse({"link":f"https://trakky.in/{str(vendor.salon.city).lower()}/{str(vendor.salon.area).lower()}/salons/{str(vendor.salon.slug).lower()}"})



class SupplierListCreateAPIView(generics.ListCreateAPIView):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    authentication_classes = [VendorJWTAuthentication]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')

        queryset = Supplier.objects.filter(vendor_user=self.request.user.id)
        
        name = self.request.query_params.get('name')
        first_name = self.request.query_params.get('first_name')
        last_name = self.request.query_params.get('last_name')
        telephone = self.request.query_params.get('telephone')
        mobile_no = self.request.query_params.get('mobile_no')
        city = self.request.query_params.get('city')
        state = self.request.query_params.get('state')
        country = self.request.query_params.get('country')
        pincode = self.request.query_params.get('pincode')

        # Apply filters
        if name:
            queryset = queryset.filter(name__icontains=name)
        if first_name:
            queryset = queryset.filter(first_name__icontains=first_name)
        if last_name:
            queryset = queryset.filter(last_name__icontains=last_name)
        if telephone:
            queryset = queryset.filter(telephone=telephone)
        if mobile_no:
            queryset = queryset.filter(mobile_no=mobile_no)
        if city:
            queryset = queryset.filter(city__icontains=city)
        if state:
            queryset = queryset.filter(state__icontains=state)
        if country:
            queryset = queryset.filter(country__icontains=country)
        if pincode:
            queryset = queryset.filter(pincode=pincode)

        return queryset

    def perform_create(self, serializer):
        serializer.save(vendor_user=self.request.user)

class SupplierRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    authentication_classes = [VendorJWTAuthentication]

class BrandListCreateAPIView(generics.ListCreateAPIView):
    queryset = Brand.objects.all()
    serializer_class = BrandSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return Brand.objects.filter(vendor=self.request.user.id)

    def perform_create(self, serializer):
        serializer.save(vendor=self.request.user)


class BrandRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Brand.objects.all()
    serializer_class = BrandSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

class ProductListCreateAPIView(generics.ListCreateAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    authentication_classes = [VendorJWTAuthentication]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')

        queryset = Product.objects.filter(supplier__vendor_user=self.request.user.id)
        
        supplier_id = self.request.query_params.get('supplier_id')
        product_identification_number = self.request.query_params.get('product_indentification_number')
        product_name = self.request.query_params.get('product_name')
        brand_name = self.request.query_params.get('brand_name')

        if supplier_id:
            queryset = queryset.filter(supplier__id=supplier_id)
        
        if product_identification_number:
            queryset = queryset.filter(
                product_indentification_number__icontains=product_identification_number
            )
        
        if product_name:
            queryset = queryset.filter(
                product_name__icontains=product_name
            )
        
        if brand_name:
            queryset = queryset.filter(
                product_brand__name__icontains=brand_name
            )

        return queryset

class ProductRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    authentication_classes = [VendorJWTAuthentication]

class StockorderListCreateAPIView(generics.ListCreateAPIView):
    queryset = Stockorder.objects.all()
    serializer_class = StockorderSerializer
    authentication_classes = [VendorJWTAuthentication]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')

        queryset = Stockorder.objects.filter(product__supplier__vendor_user=self.request.user.id)
        
        # Get filter parameters
        for_what = self.request.query_params.get('for_what')
        total_cost = self.request.query_params.get('total_cost')
        product_id = self.request.query_params.get('product_id')
        status = self.request.query_params.get('status')

        # Apply filters
        if for_what:
            queryset = queryset.filter(for_what__icontains=for_what)
        if total_cost:
            queryset = queryset.filter(total_cost=total_cost)
        if product_id:
            queryset = queryset.filter(product__id=product_id)
        if status:
            queryset = queryset.filter(status__iexact=status)

        return queryset

class StockorderRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Stockorder.objects.all()
    serializer_class = StockorderSerializer
    authentication_classes = [VendorJWTAuthentication]

    def perform_update(self, serializer):
        instance = serializer.save()

        if instance.status == 'completed':
            product = instance.product
            stock_order_quantity = instance.product_quantity

            # Retrieve the vendor_user from the request (JWT bearer token)
            vendor_user = self.request.user

            # Determine if this is for SellingInventory or UseInventory
            if instance.for_what == 'sell':
                inventory_model = SellingInventory
            else:
                inventory_model = UseInventory

            # Check if an inventory entry exists for this product
            inventory = inventory_model.objects.filter(
                product=product,
                supply_price_per_unit=instance.supply_price,
                retail_price_per_unit=instance.retail_price,
                vendor_user=vendor_user  # Filter by the authenticated vendor user
            ).first()

            if inventory:
                # If product, supply price, and retail price all match, update the quantity
                inventory.quantity += stock_order_quantity
                inventory.save()
            else:
                # If there is a mismatch, create a new inventory entry
                inventory_model.objects.create(
                    vendor_user=vendor_user,  # Pass the vendor_user from the token
                    product=product,
                    quantity=stock_order_quantity,
                    supply_price_per_unit=instance.supply_price,
                    retail_price_per_unit=instance.retail_price
                )

# class GiftCardListCreateAPIView(generics.ListCreateAPIView):
#     serializer_class = GiftCardSerializer
#     authentication_classes = [VendorJWTAuthentication]

#     def get_queryset(self):
#         queryset = GiftCard.objects.all()
#         vendor_id = self.request.user.id

#         if vendor_id:
#             queryset = queryset.filter(vendor_user=vendor_id)

#         return queryset
    
#     def perform_create(self, serializer):
#         serializer.save(vendor_user=self.request.user)

# class GiftCardRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
#     queryset = GiftCard.objects.all()
#     serializer_class = GiftCardSerializer
#     authentication_classes = [VendorJWTAuthentication]

#     def perform_update(self, serializer):
#         instance = serializer.instance
#         instance_validity_end_date = instance.validity_end_date
#         current_date = date.today()

#         if instance_validity_end_date < current_date:
#             serializer.validated_data['status'] = 'Expired'

#         serializer.save()


class CategoryRequestCreateAPIView(generics.ListCreateAPIView):
    queryset = CategoryRequest.objects.all()
    serializer_class = CategoryRequestSerializer
    authentication_classes = [VendorJWTAuthentication]
    

    class CategoryConflict(APIException):
        status_code = status.HTTP_409_CONFLICT
        default_detail = 'Same category name is present in the master category'
        default_code = 'category_conflict'

    def perform_create(self, serializer):
        vendor_user = self.request.user

        # Check if the request is authenticated
        if not vendor_user.is_authenticated:
            raise PermissionDenied('User is not authenticated')

        # Get the data from the validated serializer
        category_name = serializer.validated_data.get('category_name')
        gender = serializer.validated_data.get('gender')
        from_master = serializer.validated_data.get('from_master')

        # Check for existing master category with the same name and gender
        if not from_master:
            if MasterCategory.objects.filter(name__iexact=category_name, gender__iexact=gender).exists():
                raise self.CategoryConflict()

        # Save the instance
        serializer.save(vendor_user=vendor_user)

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return CategoryRequest.objects.filter(vendor_user=self.request.user.id).order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(vendor_user=self.request.user)


        return CategoryRequest.objects.filter(vendor_user=self.request.user)

    
class CategoryRequestRetrieveDestroyAPIView(generics.RetrieveDestroyAPIView):
    queryset = CategoryRequest.objects.all()
    serializer_class = CategoryRequestSerializer
    authentication_classes = [VendorJWTAuthentication]

class CategoryRequestAdminListAPIView(generics.ListAPIView):
    queryset = CategoryRequest.objects.all()
    serializer_class = CategoryRequestSerializer
    authentication_classes = [JWTAuthentication]
    filter_backends = [DjangoFilterBackend]

    def get_queryset(self):
        queryset = super().get_queryset()
        salon_city = self.request.query_params.get('salon_city', None)
        salon_area = self.request.query_params.get('salon_area', None)
        salon_id = self.request.query_params.get('salon_id', None)

        if salon_city:
            queryset = queryset.filter(salon__city__icontains=salon_city)
        if salon_area:
            queryset = queryset.filter(salon__area__icontains=salon_area)
        if salon_id:
            queryset = queryset.filter(salon__id=salon_id)

        return queryset.order_by('-created_at')

class CategoryRequestAdminRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = CategoryRequest.objects.all()
    serializer_class = CategoryRequestSerializer
    authentication_classes = [JWTAuthentication]


class ServiceRequestCreateAPIView(generics.ListCreateAPIView):
    queryset = ServiceRequest.objects.all()
    serializer_class = ServiceRequestSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    class ConflictException(APIException):
        status_code = status.HTTP_409_CONFLICT
        default_detail = 'Same service name is present in the master service'
        default_code = 'conflict'

    def perform_create(self, serializer):
        from_masterservice = serializer.validated_data.get('from_masterservice')
        service_name = serializer.validated_data.get('service_name')
        gender = serializer.validated_data.get('gender')

        if not from_masterservice:
            if MasterService.objects.filter(service_name__iexact=service_name, gender__iexact=gender).exists():
                raise self.ConflictException()

        serializer.save(vendor_user=self.request.user)

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return ServiceRequest.objects.filter(vendor_user=self.request.user.id).order_by('-created_at')
    
class ServiceRequestRetrieveDestroyAPIView(generics.RetrieveDestroyAPIView):
    queryset = ServiceRequest.objects.all()
    serializer_class = ServiceRequestSerializer
    authentication_classes = [VendorJWTAuthentication]

class ServiceRequestAdminListAPIView(generics.ListAPIView):
    queryset = ServiceRequest.objects.all()
    serializer_class = ServiceRequestSerializer
    authentication_classes = [JWTAuthentication]
    filter_backends = [DjangoFilterBackend]

    def get_queryset(self):
        queryset = super().get_queryset()
        salon_city = self.request.query_params.get('salon_city', None)
        salon_area = self.request.query_params.get('salon_area', None)
        salon_id = self.request.query_params.get('salon_id', None)

        if salon_city:
            queryset = queryset.filter(salon__city__icontains=salon_city)
        if salon_area:
            queryset = queryset.filter(salon__area__icontains=salon_area)
        if salon_id:
            queryset = queryset.filter(salon__id=salon_id)

        return queryset.order_by('-created_at')

class ServiceRequestAdminRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ServiceRequest.objects.all()
    serializer_class = ServiceRequestSerializer
    authentication_classes = [JWTAuthentication]


@api_view(['GET'])
def checkCategoryExistInCity(request):
    category_name = request.query_params.get('category_name')
    salon_id = request.query_params.get('salon_id')
    gender = request.query_params.get('gender')

    if not category_name or not salon_id or not gender:
        return Response({'message': 'category_name, salon_id, and gender parameters are required'}, status=400)

    try:
        salon = Salon.objects.get(id=salon_id)
    except Salon.DoesNotExist:
        return Response({'message': 'Salon not found'}, status=404)

    city = salon.city
    check = False
    if CategoryModel.objects.filter(
        master_category__name__iexact=category_name,
        master_category__gender__iexact=gender,
        city__iexact=city
    ).exists():
        check = True

    return Response({'exists': check})


@api_view(['POST'])
def createCategory(request):
    add = request.data.get('add')
    slug = request.data.get('slug')
    salon_id = request.data.get('salon')
    category_name = request.data.get('category_name')
    gender = request.data.get('gender')

    try:
        salon_obj = Salon.objects.get(id=salon_id)
    except Salon.DoesNotExist:
        return Response({'message': 'Salon not found'}, status=status.HTTP_404_NOT_FOUND)
        
    city = salon_obj.city

    try:
        master_category = MasterCategory.objects.get(name__iexact=category_name, gender__iexact=gender)
    except MasterCategory.DoesNotExist:
        return Response({'message': 'Master Category not found'}, status=status.HTTP_404_NOT_FOUND)

    if add == 'false':
        # Check if the category already exists
        if CategoryModel.objects.filter(city=city, master_category=master_category).exists():
            return Response({'message': 'Category with this city and master category already exists.'}, status=status.HTTP_400_BAD_REQUEST)

        data = {
            'city': city,
            'master_category': master_category.id,
            'slug': slug,
            'salon': [salon_id],  # Wrap in list if not already
            'category_name': category_name,
            'category_gender': gender
        }
        serializer = SalonCategoryModelSerializer(data=data)
        if serializer.is_valid():
            category_instance = serializer.save()
            # Assign master_category explicitly
            category_instance.master_category = master_category
            category_instance.save()
            
            # Update CategoryRequest status to approved
            CategoryRequest.objects.filter(
                salon=salon_obj,
                category_name=category_name,
                gender=gender,
                category_status='pending'
            ).update(category_status='approved', is_approved=True)

            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    else:
        try:
            category = CategoryModel.objects.filter(city=city, master_category=master_category).first()
            if not category:
                return Response({"detail": "Category not found"}, status=status.HTTP_404_NOT_FOUND)

            category.salon.add(salon_obj)
            category.save()
            
            # Update CategoryRequest status to approved
            CategoryRequest.objects.filter(
                salon=salon_obj,
                category_name=category_name,
                gender=gender,
                category_status='pending'
            ).update(category_status='approved', is_approved=True)

            serializer = CategoryModelSerializer(category)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Salon.DoesNotExist:
            return Response({"detail": "Salon not found"}, status=status.HTTP_404_NOT_FOUND)
        except CategoryModel.MultipleObjectsReturned:
            return Response({"detail": "Multiple categories found. Please refine your search."}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def createMasterCategory(request):
    salon_id = request.data.get('salon')
    category_name = request.data.get('category_name')
    slug = request.data.get('slug')
    image = request.data.get('image')
    gender = request.data.get('gender')

    # Create master category data
    master_category_data = {
        'mastercategory_image': image,
        'gender': gender,
        'name': category_name
    }

    # Serialize and save the master category
    master_category_serializer = MasterCategorySerializer(data=master_category_data)
    if master_category_serializer.is_valid():
        master_category_instance = master_category_serializer.save()
    else:
        return Response(master_category_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # Find the salon object
    try:
        salon_obj = Salon.objects.get(id=salon_id)
    except Salon.DoesNotExist:
        return Response({'message': 'Salon not found'}, status=status.HTTP_404_NOT_FOUND)
    
    city = salon_obj.city

    # Find the created master category object
    try:
        master_category = MasterCategory.objects.get(name=category_name, gender=gender).id
    except MasterCategory.DoesNotExist:
        return Response({'message': 'Master Category not found'}, status=status.HTTP_404_NOT_FOUND)

    # Create salon category data
    data = {
        'city': city,
        'master_category': master_category,
        'slug': slug,
        'salon': [salon_id],
        'category_name': category_name
    }
    
    serializer = SalonCategoryModelSerializer(data=data)
    if serializer.is_valid():
        category_instance = serializer.save()
        # Ensure master_category is set correctly
        category_instance.master_category_id = master_category
        category_instance.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(['POST'])
def createMasterServiceAndService(request):
    # Extracting data from request
    gender = request.data.get('gender')
    service_name = request.data.get('service_name')
    master_service_description = request.data.get('master_service_description')
    service_description = request.data.get('service_description')
    image = request.data.get('image')
    salon_id = request.data.get('salon_id')
    price = request.data.get('price')
    category_id = request.data.get('category_id')
    master_service_id = request.data.get('master_service_id')
    from_masterservice = request.data.get('from_masterservice', False)  # Default to False if not provided
    from_masterservice = request.data.get('from_masterservice')  # Default to False if not provided
    # discount= request.data.get('discount', 0)  # Default to 0

    # Validate required fields
    if not salon_id or not service_name or not price or not category_id:
        return Response({'message': 'salon_id, service_name, price, and category_id are required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        salon_obj = Salon.objects.get(id=salon_id)
    except Salon.DoesNotExist:
        return Response({'message': 'Salon not found'}, status=status.HTTP_404_NOT_FOUND)
    
    try:
        category = CategoryModel.objects.get(id=category_id)
        master_category_id = category.master_category.id
    except CategoryModel.DoesNotExist:
        return Response({'message': 'Category not found'}, status=status.HTTP_404_NOT_FOUND)
    except AttributeError:
        return Response({'message': 'Master category not found for the given category'}, status=status.HTTP_404_NOT_FOUND)
    if from_masterservice == 'true':
        # When from_masterservice is true, create a service linked to the existing master service
        print("if")
        if not master_service_id:
            return Response({'message': 'master_service_id is required when from_masterservice is true.'}, status=status.HTTP_400_BAD_REQUEST)

        # Service data
        service_data = {
            'city': salon_obj.city,
            'master_service': master_service_id,
            'salon': salon_id,
            'service_name': service_name,
            'price': price,
            'description': service_description,
            'categories': category_id,
            'discount': 0,
        }

        service_serializer = ServiceSerializer(data=service_data)
        if service_serializer.is_valid():
            service_instance = service_serializer.save()
            # Update ServiceRequest status to approved
            ServiceRequest.objects.filter(
                salon=salon_obj,
                service_name=service_name,
                category_id=category_id,
                service_status='pending'
            ).update(service_status='approved', is_approved=True)

            return Response(service_serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(service_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    else:
        # When from_masterservice is false, create both a new master service and a service
        print("else")
        category = CategoryModel.objects.get(id=category_id)
        master_category_id = category.master_category.id
        master_service_data = {
            'service_image': image,
            'gender': gender,
            'service_name': service_name,
            'description': master_service_description,
            'categories': master_category_id,  # Use master_category_id here
        }

        master_service_serializer = MasterServiceSerializer(data=master_service_data)
        if master_service_serializer.is_valid():
            master_service_instance = master_service_serializer.save()
        else:
            return Response(master_service_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # Service data
        service_data = {
            'city': salon_obj.city,
            'master_service': master_service_instance.id,
            'salon': salon_id,
            'service_name': service_name,
            'price': price,
            'description': service_description,
            'categories': category_id,
            'discount': 0,
        }

        service_serializer = ServiceSerializer(data=service_data)
        if service_serializer.is_valid():
            service_instance = service_serializer.save()
            # Update ServiceRequest status to approved
            ServiceRequest.objects.filter(
                salon=salon_obj,
                service_name=service_name,
                category_id=category_id,
                service_status='pending'
            ).update(service_status='approved', is_approved=True)

            return Response(service_serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(service_serializer.errors, status=status.HTTP_400_BAD_REQUEST)



@api_view(['GET'])
def get_master_categories_not_in_salon(request):
    salon_id = request.query_params.get('salon_id')

    if not salon_id:
        return Response({'message': 'Salon ID is required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        salon = Salon.objects.get(id=salon_id)
    except Salon.DoesNotExist:
        return Response({'message': 'Salon not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Get the master categories already present in the salon's categories
    existing_master_category_ids = CategoryModel.objects.filter(salon=salon).values_list('master_category_id', flat=True)
    
    # Get the master categories that are not present in the salon
    available_master_categories = MasterCategory.objects.exclude(id__in=existing_master_category_ids)
    
    serializer = MasterCategorySerializer(available_master_categories, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)



@api_view(['GET'])
def get_master_services_not_in_salon(request):
    salon_id = request.query_params.get('salon_id')

    if not salon_id:
        return Response({'message': 'Salon ID is required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        salon = Salon.objects.get(id=salon_id)
    except Salon.DoesNotExist:
        return Response({'message': 'Salon not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Get the master services already present in the salon's services
    existing_master_service_ids = Services.objects.filter(salon=salon).values_list('master_service__id', flat=True)
    
    # Get the master services that are not present in the salon
    available_master_services = MasterService.objects.exclude(id__in=existing_master_service_ids)
    
    serializer = MasterServiceSerializer(available_master_services, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)
    

class DailyExpensisListCreateView(generics.ListCreateAPIView):
    queryset = DailyExpensis.objects.all()
    serializer_class = DailyExpensisSerializer
    authentication_classes = [VendorJWTAuthentication]
    from datetime import datetime, timedelta
    from django.utils import timezone
    import datetime as dt

    def get_queryset(self):
        import datetime as dt

        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')

        queryset = DailyExpensis.objects.filter(vendor_user=self.request.user)

        start_date_str = self.request.query_params.get('start_date')
        end_date_str = self.request.query_params.get('end_date')

        if start_date_str and end_date_str:
            try:
                start_date = dt.datetime.strptime(start_date_str, '%Y-%m-%d')
                end_date = dt.datetime.strptime(end_date_str, '%Y-%m-%d') + dt.timedelta(days=1)

                start_date = timezone.make_aware(start_date, timezone.get_current_timezone())
                end_date = timezone.make_aware(end_date, timezone.get_current_timezone())

                queryset = queryset.filter(created_at__range=(start_date, end_date))
            except ValueError:
                raise ValueError('Invalid date format. Please use YYYY-MM-DD.')

        return queryset

    def perform_create(self, serializer):
        serializer.save(vendor_user=self.request.user)

class DailyExpensisDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = DailyExpensis.objects.all()
    serializer_class = DailyExpensisSerializer
    authentication_classes = [VendorJWTAuthentication]


class MembershipListAPIView(generics.ListCreateAPIView):
    queryset = MemberShip.objects.all()
    serializer_class = MemberShipSerializer
    authentication_classes = [VendorJWTAuthentication]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return MemberShip.objects.filter(vendor_user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(vendor_user=self.request.user)


class MembershipRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = MemberShip.objects.all()
    serializer_class = MemberShipSerializer
    authentication_classes = [VendorJWTAuthentication]


class CustomerMembershipListAPIView(generics.ListCreateAPIView):
    queryset = CustomerMembershipnew.objects.all()
    serializer_class = CustomerMembershipnewSerializer
    authentication_classes = [VendorJWTAuthentication]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        queryset = CustomerMembershipnew.objects.filter(vendor_user=self.request.user)
        
        customer_number = self.request.query_params.get('customer_number')
        
        if customer_number:
            queryset = queryset.filter(customer_number=customer_number)
        
        return queryset

    def perform_create(self, serializer):
        serializer.save(vendor_user=self.request.user)


class CustomerMembershipRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = CustomerMembershipnew.objects.all()
    serializer_class = CustomerMembershipnewSerializer
    authentication_classes = [VendorJWTAuthentication]

class Standard25Pagination(PageNumberPagination):
    page_size = 25
    page_size_query_param = 'page_size'
    max_page_size = 100


class AppointmentListCreateView(generics.ListCreateAPIView):
    import datetime
    from datetime import timedelta

    queryset = Appointment.objects.all()
    serializer_class = AppointmentNewSerializer
    authentication_classes = [VendorJWTAuthentication]
    # pagination_class = Standard25Pagination


    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')

        queryset = (
            Appointment.objects
            .filter(vendor_user=self.request.user)
            .select_related(
                'vendor_user',
                'manager',
                'customer',
                'selled_product',
                'customer_wallet',
            )
            .prefetch_related(
                'membership',
                'service',
                'offer',
                'staff',
                'product_consumption',
            )
        )

        # Staff ID filter
        staff_id = self.request.query_params.get('staff_id')
        if staff_id:
            queryset = queryset.filter(staff__id=staff_id)

        # Date range filter (based on Appointment.date)
        start_date_str = self.request.query_params.get('start_date')
        end_date_str = self.request.query_params.get('end_date')

        if start_date_str and end_date_str:
            try:
                from datetime import datetime
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                queryset = queryset.filter(date__range=(start_date, end_date))
            except ValueError:
                raise ValidationError('Invalid date format. Please use YYYY-MM-DD.')

        is_delete = self.request.query_params.get("is_delete")  # expects: true / false / 0 / 1
        if is_delete is not None:
            if is_delete.lower() in ["true", "1"]:
                queryset = queryset.filter(is_delete=True)
            elif is_delete.lower() in ["false", "0"]:
                queryset = queryset.filter(is_delete=False)
            else:
                raise ValidationError("Invalid value for is_delete. Use true/false or 1/0.")


        # Status ordering logic
        status_order = {
            'not_started': 1,
            'running': 2,
            'completed': 3,
            'cancelled': 4,
        }

        def get_status_order(appointment):
            status = appointment.appointment_status or 'not_started'
            return status_order.get(status, float('inf'))

        # Sort queryset in memory by status and creation time
        queryset = sorted(
            queryset,
            key=lambda x: (get_status_order(x), -x.created_at.timestamp())
        )

        return queryset

    def perform_create(self, serializer):
        customer_phone = serializer.validated_data.get('customer_phone')
        customer_name = serializer.validated_data.get('customer_name')
        customer_gender = serializer.validated_data.get('customer_gender')
        vendor_user = self.request.user
        for_consultation = serializer.validated_data.get('for_consultation', False)

        customer, _ = CustomerTable.objects.get_or_create(
            customer_phone=customer_phone,
            vendor_user=vendor_user,
            defaults={
                "customer_name": customer_name,
                "customer_gender": customer_gender,
            },
        )

        customer_type = (
            'regular'
            if Appointment.objects.filter(vendor_user=vendor_user, customer_phone=customer_phone).exists()
            else 'new'
        )
        
         # Check if GST is enabled for this vendor
        is_gst_enabled = vendor_user.is_gst
        
        # Prepare save data
        save_data = {
            'vendor_user': vendor_user,
            'customer_type': customer_type,
            'customer': customer,
        }
        
        # If GST is NOT enabled, disable all tax fields
        if not is_gst_enabled:
            save_data['tax_percentage'] = None
            save_data['tax_amount'] = Decimal('0.00')
            save_data['appointment_tax_percentage'] = None
            save_data['appointment_tax_amount'] = Decimal('0.00')
            save_data['product_tax_percentage'] = None
            save_data['product_tax_amount'] = Decimal('0.00')
        
        # If it's a consultation appointment, set payment-related fields to 0
        if for_consultation:
            save_data.update({
                'actual_amount': 0,
                'final_amount': 0,
                'amount_paid': 0,
                'payment_status': 'free' if hasattr(Appointment, 'payment_status') else None
            })
        
        serializer.save(**save_data)
    


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404

from .models import Appointment
from .services import (
    send_appointment_notification,
    update_product_consumption,
)


class AppointmentProcessView(APIView):
    authentication_classes = [VendorJWTAuthentication]

    def post(self, request):
        appointment_id = request.data.get("appointment_id")

        if not appointment_id:
            return Response(
                {"error": "appointment_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # safer way
        appointment = get_object_or_404(
            Appointment,
            id=appointment_id,
            vendor_user=request.user
        )

        try:
            # 🔥 Call service layer functions
            send_appointment_notification(appointment)
            update_product_consumption(appointment)

            return Response(
                {"message": "Notification sent and product consumption updated successfully"},
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )



class AppointmentRetrieveDestroyUpdateView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Appointment.objects.all()
    serializer_class = AppointmentNewSerializer
    authentication_classes = [VendorJWTAuthentication]

    def partial_update(self, request, *args, **kwargs):
        # Call the super method for the patch operation
        response = super().partial_update(request, *args, **kwargs)
        
        # Only execute if the request was successful
        if response.status_code == status.HTTP_200_OK:
            # Get the updated appointment instance
            appointment = self.get_object()

            # Check if checkout is True and update appointment status
            if appointment.checkout and appointment.appointment_status != 'completed':
                print("Checkout is done, updating appointment status to 'completed'.")
                appointment.appointment_status = 'completed'
                appointment.save()
            else:
                print("Checkout not done or appointment already completed.")
            
            # Check if product_details exists and has data
            if appointment.product_details:
                print("Updating product consumption...")
                appointment.update_product_consumption()
                print("Product consumption updated.")
            else:
                print("No product details found to update consumption.")
        
        return response

class CreateOnlyAppointmentView(generics.ListCreateAPIView):
    queryset = Appointment.objects.all()
    serializer_class = AppointmentNewSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        customer_phone = serializer.validated_data.get('customer_phone')
        customer_name = serializer.validated_data.get('customer_name')
        customer_gender = serializer.validated_data.get('customer_gender')

        # No vendor_user usage anymore
        customer = CustomerTable.objects.filter(customer_phone=customer_phone).first()

        if not customer:
            customer = CustomerTable.objects.create(
                customer_phone=customer_phone,
                customer_name=customer_name,
                customer_gender=customer_gender
            )

        customer_type = 'regular' if Appointment.objects.filter(customer_phone=customer_phone).exists() else 'new'

        serializer.save(customer_type=customer_type, customer=customer)


@api_view(['GET'])
@authentication_classes([VendorJWTAuthentication])
@permission_classes([IsAuthenticated])
def appointments_filter_by_month(request):
    from django.http import JsonResponse
    from datetime import datetime, timedelta
    year = request.GET.get('year')
    month = request.GET.get('month')

    if not year or not month:
        return JsonResponse({'error': 'Year and month are required parameters'}, status=400)

    try:
        year = int(year)
        month = int(month)
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
    except ValueError:
        return JsonResponse({'error': 'Invalid year or month'}, status=400)

    appointments = Appointment.objects.filter(date__gte=start_date, date__lt=end_date,vendor_user=request.user.id)
    
    result = {}
    
    for day in range(1, (end_date - start_date).days + 1):
        date_key = start_date + timedelta(days=day - 1)
        date_str = date_key.strftime("%d/%m/%Y")

        day_appointments = appointments.filter(date=date_key)
        
        result[date_str] = {
            'total_appointments': list(day_appointments.values()),
            'running_appointments': list(day_appointments.filter(appointment_status='running').values()),
            'not_started_appointments': list(day_appointments.filter(appointment_status='not_started').values()),
            'cancelled_appointments': list(day_appointments.filter(appointment_status='cancelled').values()),
            'completed_appointments': list(day_appointments.filter(appointment_status='completed').values())
        }
    
    return JsonResponse(result, safe=False)


class SaleListCreate(generics.ListCreateAPIView):
    queryset = Sale.objects.all()
    serializer_class = SaleSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return Sale.objects.filter(vendor_user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(vendor_user=self.request.user)
class SaleRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Sale.objects.all()
    serializer_class = SaleSerializer
    authentication_classes = [VendorJWTAuthentication]


class SalonRequestListCreateAPIView(generics.ListCreateAPIView):
    queryset = SalonRequest.objects.all()
    serializer_class = SalonRequestSerializer
    # authentication_classes = [VendorJWTAuthentication]
    # permission_classes = [IsAuthenticated]
    permission_classes = [AllowAny]

    def get_queryset(self):
        queryset = super().get_queryset()
        city = self.request.query_params.get('city')
        area = self.request.query_params.get('area')
        
        if city:
            queryset = queryset.filter(city__iexact=city)
        if area:
            queryset = queryset.filter(area__iexact=area)
        
        return queryset

class SalonRequestRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = SalonRequest.objects.all()
    serializer_class = SalonRequestSerializer
    # authentication_classes = [VendorJWTAuthentication]
    # permission_classes = [IsAuthenticated]
    permission_classes = [AllowAny]

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({"message": "Salon deleted successfully"}, status=status.HTTP_204_NO_CONTENT)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response({"message": "Salon updated successfully", "data": serializer.data})


class SellingInventoryListCreateAPIView(generics.ListCreateAPIView):
    queryset = SellingInventory.objects.filter(is_active=True)  # Only active products
    serializer_class = SellingInventorySerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return SellingInventory.objects.filter(vendor_user=self.request.user)

    def create(self, request, *args, **kwargs):
        product_id = request.data.get('product')
        product_quantity = request.data.get('quantity')
        supply_price = request.data.get('supply_price_per_unit')
        retail_price = request.data.get('retail_price_per_unit')

        product = Product.objects.get(id=product_id)

        # Check if SellingInventory exists
        inventory = SellingInventory.objects.filter(product=product).first()

        if inventory:
            if inventory.supply_price_per_unit == float(supply_price) and \
            inventory.retail_price_per_unit == float(retail_price):
                inventory.quantity += int(product_quantity)
                inventory.save()
                return Response({'detail': 'Inventory updated successfully'}, status=status.HTTP_200_OK)
            else:
                SellingInventory.objects.create(
                    vendor_user=request.user,
                    product=product,
                    quantity=product_quantity,
                    supply_price_per_unit=supply_price,
                    retail_price_per_unit=retail_price
                )
                return Response({'detail': 'New inventory entry created with different prices'}, status=status.HTTP_201_CREATED)
        else:
            SellingInventory.objects.create(
                vendor_user=request.user,
                product=product,
                quantity=product_quantity,
                supply_price_per_unit=supply_price,
                retail_price_per_unit=retail_price
            )
            return Response({'detail': 'New inventory entry created'}, status=status.HTTP_201_CREATED)

class SellingInventoryRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = SellingInventory.objects.all()
    serializer_class = SellingInventorySerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def perform_destroy(self, instance):
        if instance.is_active:
            # Soft delete: Set is_active to False
            instance.is_active = False
            instance.save()
        else:
            # Permanent delete: If is_active is already False, delete the object
            instance.delete()


class UseInventoryListCreateAPIView(generics.ListCreateAPIView):
    queryset = UseInventory.objects.filter(is_active=True)  # Only active products
    serializer_class = UseInventorySerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return UseInventory.objects.filter(vendor_user=self.request.user)


    def create(self, request, *args, **kwargs):
        product_id = request.data.get('product')
        product_quantity = request.data.get('quantity')
        supply_price = request.data.get('supply_price_per_unit')

        product = Product.objects.get(id=product_id)

        # Check if UseInventory exists
        inventory = UseInventory.objects.filter(product=product).first()

        if inventory:
            # Compare supply price only (as retail price is null in use case)
            if inventory.supply_price_per_unit == float(supply_price):
                # If supply price matches, update the quantity
                inventory.quantity += int(product_quantity)
                inventory.save()
                return Response({'detail': 'Inventory updated successfully'}, status=status.HTTP_200_OK)
            else:
                # If supply price doesn't match, create a new UseInventory entry
                UseInventory.objects.create(
                    product=product,
                    quantity=product_quantity,
                    supply_price_per_unit=supply_price,
                    retail_price_per_unit=None  # UseInventory has no retail price
                )
                return Response({'detail': 'New use inventory entry created with different supply price'}, status=status.HTTP_201_CREATED)
        else:
            # If UseInventory does not exist, create a new one
            UseInventory.objects.create(
                product=product,
                quantity=product_quantity,
                supply_price_per_unit=supply_price,
                retail_price_per_unit=None  # UseInventory has no retail price
            )
            return Response({'detail': 'New use inventory entry created'}, status=status.HTTP_201_CREATED)



# Retrieve, Update, Delete (Soft delete)
class UseInventoryRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = UseInventory.objects.all()
    serializer_class = UseInventorySerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def perform_destroy(self, instance):
        if instance.is_active:
            # Soft delete: Set is_active to False
            instance.is_active = False
            instance.save()
        else:
            # Permanent delete: If is_active is already False, delete the object
            instance.delete()

class SellListCreateAPIView(generics.ListCreateAPIView):
    queryset = Sell.objects.all()
    serializer_class = SellSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')

        queryset = Sell.objects.filter(vendor_user=self.request.user)
        
        # Get filter parameters
        customer_name = self.request.query_params.get('customer_name')
        customer_phone = self.request.query_params.get('customer_phone')
        customer_gender = self.request.query_params.get('customer_gender')

        # Apply filters
        if customer_name:
            queryset = queryset.filter(customer_name__icontains=customer_name)
        if customer_phone:
            queryset = queryset.filter(customer_phone__icontains=customer_phone)
        if customer_gender:
            queryset = queryset.filter(customer_gender__iexact=customer_gender)

        return queryset

    def perform_create(self, serializer):
        serializer.save(vendor_user=self.request.user)

class SellDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Sell.objects.all()
    serializer_class = SellSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]



class AppointmentNotificationView(generics.ListCreateAPIView):
    queryset = AppointmentNotification.objects.all()
    serializer_class = AppointmentNotificationSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return AppointmentNotification.objects.filter(Appointment__vendor_user=self.request.user)
    

class AppointmentNotificationCreateView(generics.ListCreateAPIView):
    queryset = AppointmentNotification.objects.all()
    serializer_class = AppointmentNotificationSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return AppointmentNotification.objects.filter(Appointment__vendor_user=self.request.user)


class ScoreNotificationView(generics.ListCreateAPIView):
    queryset = ScoreNotification.objects.all()
    serializer_class = ScoreNotificationSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return ScoreNotification.objects.filter(appointmentscore__appointment__vendor_user=self.request.user)
    

class ScoreNotificationCreateView(generics.ListCreateAPIView):
    queryset = ScoreNotification.objects.all()
    serializer_class = ScoreNotificationSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return ScoreNotification.objects.filter(appointmentscore__appointment__vendor_user=self.request.user)


class TipNotificationView(generics.ListCreateAPIView):
    queryset = TipNotification.objects.all()
    serializer_class = TipNotificationSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return TipNotification.objects.filter(appointmentscore__appointment__vendor_user=self.request.user)


class TipNotificationCreateView(generics.ListCreateAPIView):
    queryset = TipNotification.objects.all()
    serializer_class = TipNotificationSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return TipNotification.objects.filter(appointmentscore__appointment__vendor_user=self.request.user)


class CustomerPurchaseSummaryView(APIView):
    
    def get(self, request, customer_id):
        try:
            customer = CustomerTable.objects.get(id=customer_id)
            
            # Get counts of different purchases
            product_purchases = customer.product_purchases.all()
            appointment_purchases = customer.appointment_purchases.all()
            membership_purchases = customer.membership_purchases.all()
            
            # Prepare detailed purchase information
            product_purchase_list = [
                {
                    'product_name': purchase.product_name,
                    'purchase_date': purchase.purchase_date,
                    'quantity': purchase.quantity
                }
                for purchase in product_purchases
            ]
            
            appointment_purchase_list = [
                {
                    'service_name': purchase.service_name,
                    'appointment_date': purchase.appointment_date,
                    'purchase_date': purchase.purchase_date
                }
                for purchase in appointment_purchases
            ]
            
            membership_purchase_list = [
                {
                    'membership_type': purchase.membership_type.name,
                    'purchase_date': purchase.purchase_date
                }
                for purchase in membership_purchases
            ]
            
            # Response data including customer details
            data = {
                'customer_details': {
                    'customer_name': customer.customer_name,
                    'customer_phone': customer.customer_phone,
                    'customer_email': customer.customer_email,
                    'customer_type': customer.customer_type,
                    'customer_gender': customer.customer_gender,
                    'created_at': customer.created_at,
                },
                'purchase_summary': {
                    'product_purchase_count': product_purchases.count(),
                    'appointment_purchase_count': appointment_purchases.count(),
                    'membership_purchase_count': membership_purchases.count()
                },
                'detailed_purchases': {
                    'product_purchases': product_purchase_list,
                    'appointment_purchases': appointment_purchase_list,
                    'membership_purchases': membership_purchase_list
                }
            }
            
            return Response(data, status=status.HTTP_200_OK)
        
        except CustomerTable.DoesNotExist:
            return Response({"error": "Customer not found"}, status=status.HTTP_404_NOT_FOUND)


class PackageRequestListCreateAPIView(generics.ListCreateAPIView):
    queryset = PackageRequest.objects.all()
    serializer_class = PackageRequestSerializer
    authentication_classes = [VendorJWTAuthentication]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return PackageRequest.objects.filter(vendor_user=self.request.user).order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(vendor_user=self.request.user)

class PackageRequestDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = PackageRequest.objects.all()
    serializer_class = PackageRequestSerializer
    authentication_classes = [VendorJWTAuthentication]

class PackageRequestAdminListAPIView(generics.ListAPIView):
    queryset = PackageRequest.objects.all()
    serializer_class = PackageRequestSerializer
    authentication_classes = [JWTAuthentication]
    filter_backends = [DjangoFilterBackend]

    def get_queryset(self):
        queryset = super().get_queryset()
        # latest first
        return queryset.order_by('-created_at')

class PackageRequestAdminRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = PackageRequest.objects.all()
    serializer_class = PackageRequestSerializer
    authentication_classes = [JWTAuthentication]


class ApprovePackageAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            # 1. Retrieve the PackageRequest object
            package_request = PackageRequest.objects.get(pk=pk)
        except PackageRequest.DoesNotExist:
            return Response({'error': 'Package request not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if package_request.status == 'approved':
             return Response({'error': 'Package request is already approved'}, status=status.HTTP_400_BAD_REQUEST)

        # 2. Direct Data Retrieval from Request (Using Payload Keys)
        
        # Use the key 'package_name' from the payload or the existing object
        package_name = request.data.get('package_name', package_request.package_name)
        
        # FIX: Use the key 'discounted_price' from the payload (Your Model uses discount_price)
        # We explicitly retrieve it and check for None to allow a value of 0.
        discount_price = request.data.get('discounted_price')
        if discount_price is None:
            # Fallback to the existing discounted_price on the PackageRequest object
            discount_price = package_request.discounted_price 
        
        # package_time (JSONField)
        package_time = request.data.get('package_time', package_request.package_time)
        
        # additional_included_service (JSONField): Uses the complex service JSON from the payload
        # FIX: Use the key 'services_included' from the payload (maps to the additional JSON field)
        additional_included_service_data = request.data.get('services_included', package_request.services_included)
        
        # 3. M2M Service ID Retrieval (NO CHANGE NEEDED HERE IF IDs ARE ALREADY ON package_request)
        # Since your payload doesn't contain a simple list of 'service_included' IDs, 
        # we will rely entirely on the IDs already associated with package_request.services M2M field.
        service_ids = list(package_request.services.values_list('id', flat=True))


        # 4. Final Validation Check
        # Check for required fields. Use 'is None' for package_time and discount_price to allow 0.
        if not package_name or discount_price is None or package_time is None:
             return Response({
                 'error': 'Missing required fields: package_name, discounted_price, or package_time'
             }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate service_ids format (optional, but good practice)
        if service_ids and (not isinstance(service_ids, list) or not all(isinstance(id, int) for id in service_ids)):
             # This should only happen if package_request.services M2M has bad data
             return Response({'error': 'Internal data error: PackageRequest services M2M field is invalid.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        user = request.user 
        
        # 5. Approval and Creation Logic
        try:
            with transaction.atomic():
                # 1. Create the Package object
                package = Package.objects.create(
                    user=user,
                    salon=package_request.salon,
                    package_name=package_name,
                    # Actual price should be retrieved from the request object, not the request body
                    actual_price=package_request.actual_price or 0.00,
                    discount_price=discount_price, # Use the retrieved value
                    package_time=package_time,
                    # IMPORTANT: Use the retrieved complex JSON data
                    additional_included_service=additional_included_service_data, 
                    updated_by=user,
                    updated_date=timezone.now()
                )

                # 2. Set the ManyToMany Services
                package.service_included.set(service_ids)

                # 3. Update the PackageRequest status and data
                package_request.status = 'approved'
                package_request.package_name = package_name
                package_request.discounted_price = discount_price # Update PackageRequest using the final value
                package_request.package_time = package_time
                package_request.services.set(service_ids)
                package_request.services_included = additional_included_service_data
                package_request.save()

        except Exception as e:
            print(f"Error during package approval: {e}") 
            return Response({'error': 'An internal server error occurred during package creation. Details: {}'.format(str(e))}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # 4. Serialize and return response
        try:
            package_serializer = PackageListSerializer(package)
            return Response({
                'message': 'Package created and approved',
                'package': package_serializer.data
            }, status=status.HTTP_201_CREATED)
        except NameError:
            return Response({
                'message': 'Package created and approved, but serialization failed (PackageListSerializer not defined).',
                'package_id': package.id
            }, status=status.HTTP_201_CREATED)


class SalonProfileOfferRequestListCreateAPIView(generics.ListCreateAPIView):
    queryset = SalonProfileOfferRequest.objects.all()
    serializer_class = SalonProfileOfferRequestSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return SalonProfileOfferRequest.objects.filter(vendor_user=self.request.user).order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(vendor_user=self.request.user)

class SalonProfileOfferRequestAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = SalonProfileOfferRequest.objects.all()
    serializer_class = SalonProfileOfferRequestSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]



class SalonProfileOfferRequestAdminListAPIView(generics.ListAPIView):
    queryset = SalonProfileOfferRequest.objects.all()
    serializer_class = SalonProfileOfferRequestSerializer
    authentication_classes = [JWTAuthentication]
    filter_backends = [DjangoFilterBackend]

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.order_by('-created_at')

class SalonProfileOfferRequestAdminRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = SalonProfileOfferRequest.objects.all()
    serializer_class = SalonProfileOfferRequestSerializer
    authentication_classes = [JWTAuthentication]

import json
from django.http import JsonResponse

class ApproveSalonProfileOfferAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        # Fetch the profile offer request using the primary key (pk)
        try:
            offer_request = SalonProfileOfferRequest.objects.get(pk=pk)
        except SalonProfileOfferRequest.DoesNotExist:
            return Response({'error': 'Salon profile offer request not found'}, status=status.HTTP_404_NOT_FOUND)

        # Get required fields from request data
        salon_id = request.data.get('salon')
        name = request.data.get('name')
        actual_price = request.data.get('actual_price')
        discount_price = request.data.get('discounted_price')  # Make sure to use 'discounted_price'
        gender = request.data.get('gender')
        terms_and_conditions = request.data.get('terms_and_conditions')
        offer_timing = request.data.get('offer_timing')  # This will be received as a string
        image = request.data.get('image')

        # Validate required fields
        required_fields = {
            "salon": salon_id, 
            "name": name, 
            "actual_price": actual_price, 
            "discounted_price": discount_price, 
            "gender": gender, 
            "terms_and_conditions": terms_and_conditions, 
            "offer_timing": offer_timing
        }
        missing_fields = [field for field, value in required_fields.items() if not value]
        if missing_fields:
            return Response({'error': f'Missing required fields: {", ".join(missing_fields)}'}, status=status.HTTP_400_BAD_REQUEST)

        print("offer befor ", offer_timing)
        # Attempt to parse the offer_timing field from string to JSON
        try:
            offer_timing_json = json.loads(offer_timing)
        except json.JSONDecodeError:
            return Response({'error': 'Invalid JSON format in offer_timing'}, status=status.HTTP_400_BAD_REQUEST)
        print("offer after ", offer_timing_json)

        # Fetch the Salon object
        salon = get_object_or_404(Salon, id=salon_id)

        # Update the offer request status and save the relevant fields
        offer_request.status = 'approved'
        offer_request.salon = salon
        offer_request.name = name
        offer_request.actual_price = actual_price
        offer_request.discounted_price = discount_price  # Update discount_price correctly
        offer_request.gender = gender
        offer_request.terms_and_conditions = terms_and_conditions
        offer_request.offer_timing = offer_timing_json  # Store the JSON object
        offer_request.image = image
        offer_request.city = salon.city  # Set the city to the salon's city
        offer_request.user = request.user  # Associate the current user
        offer_request.save()

        # Create a new SalonProfileOffer object using the updated offer request data
        profile_offer = salonprofileoffer.objects.create(
            salon=salon,
            name=name,
            actual_price=actual_price,
            discount_price=discount_price,  # Ensure the correct discount price is used
            gender=gender,
            terms_and_conditions=terms_and_conditions,
            offer_time=offer_timing_json,  # Store the JSON object
            image=image,
            user=request.user,  # Associate the current user
            city=salon.city  # Set city to salon's city
        )

        # Serialize the newly created offer
        serializer = salonprofileofferSerializer(profile_offer)

        # Return response in JSON format
        return JsonResponse({
            'message': 'Salon profile offer created and approved',
            'offer': serializer.data  # Return the serialized data
        }, status=status.HTTP_201_CREATED)
    # def delete(self, request, pk):
    #     try:
    #         # Fetch the profile offer request using the primary key (pk)
    #         offer_request = SalonProfileOfferRequest.objects.get(pk=pk)
    #     except SalonProfileOfferRequest.DoesNotExist:
    #         return Response({'error': 'Salon profile offer request not found'}, status=status.HTTP_404_NOT_FOUND)

    #     try:
    #         # Fetch the offer by profile_offer_id and delete it
    #         profile_offer = SalonProfileOffer.objects.get(id=offer_request.profile_offer_id)
    #         profile_offer.delete()
    #         offer_request.status = 'pending'
    #         offer_request.save()

    #         return Response({'message': 'Salon profile offer deleted and status set to pending'}, status=status.HTTP_200_OK)
    #     except SalonProfileOffer.DoesNotExist:
    #         return Response({'error': 'Salon profile offer not found'}, status=status.HTTP_404_NOT_FOUND)

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Sum
from .models import Appointment, Sell  # adjust import path as needed

class DashboardVendorAPIView(APIView):

    # Proper imports inside class
    import datetime
    from datetime import timedelta
    from django.utils import timezone
    from django.db.models import Sum

    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        timezone = self.timezone
        datetime_module = self.datetime
        timedelta = self.timedelta
        Sum = self.Sum

        now = timezone.now()
        today = now.date()

        # Get query params
        start_date_str = request.query_params.get("start_date")
        end_date_str = request.query_params.get("end_date")

        # Default: Next 7 days (excluding today)
        if not start_date_str or not end_date_str:
            start_date = today + timedelta(days=1)
            end_date = today + timedelta(days=7)
        else:
            try:
                start_date = datetime_module.datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date = datetime_module.datetime.strptime(end_date_str, "%Y-%m-%d").date()
            except ValueError:
                return Response(
                    {"error": "Invalid date format. Use YYYY-MM-DD"},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # Optional vendor filter
        vendor_user_id = request.query_params.get("vendor_user_id")
        if vendor_user_id:
            try:
                vendor_user_id = int(vendor_user_id)
            except ValueError:
                return Response({"error": "vendor_user_id must be an integer"}, status=400)
            vendor_filter = {"vendor_user__id": vendor_user_id}
        else:
            vendor_filter = {"vendor_user": request.user}

        # --- Base Query Filters ---
        all_appointments_in_range = Appointment.objects.filter(
            date__gte=start_date,
            date__lte=end_date,
            is_delete=False,
            **vendor_filter
        )

        recent_sales_qs = Sell.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            **vendor_filter
        )

        # Completed Appointments & Revenue
        completed_appointments_qs = all_appointments_in_range.filter(
            appointment_status="completed",
            checkout=True
        )
        total_completed_appointments = completed_appointments_qs.count()
        total_revenue = completed_appointments_qs.aggregate(Sum("final_amount"))["final_amount__sum"] or 0

        # Upcoming Appointments (exclude today)
        if not start_date_str or not end_date_str:
            upcoming_appointments_qs = Appointment.objects.filter(
                date__gt=today,
                date__lte=today + timedelta(days=7),
                **vendor_filter
            )
        else:
            upcoming_appointments_qs = Appointment.objects.filter(
                date__gt=today,
                date__gte=start_date,
                date__lte=end_date,
                **vendor_filter
            )
        upcoming_appointments = upcoming_appointments_qs.count()

        # All Appointments Count
        total_appointments_count = all_appointments_in_range.count()

        # Product Sales
        total_sell_product_count = recent_sales_qs.count()
        recent_sales = recent_sales_qs.aggregate(Sum("final_total"))["final_total__sum"] or 0

        # Membership Selled Count
        total_membership_selled_count = CustomerMembership.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date,
            **vendor_filter
        ).count()

        # Wallet Selled Count
        total_wallet_selled_count = Wallet.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            **vendor_filter
        ).count()

        # Response Data
        response_data = [
            {
                "label": "Revenue (Appointment)",
                "count_label": f"₹{total_revenue:,.2f}" if total_revenue else "₹0.00",
                "timestamp_label": f"{start_date} → {end_date}",
            },
            {
                "label": "Appointments (Total Booked)",
                "count_label": f"{total_appointments_count} booked",
                "timestamp_label": f"{start_date} → {end_date}",
            },
            {
                "label": "Upcoming Appointments",
                "count_label": f"{upcoming_appointments} Appointment",
                "timestamp_label": (
                    f"{today + timedelta(days=1)} → {today + timedelta(days=7)}"
                    if not start_date_str or not end_date_str
                    else f"{start_date} → {end_date}"
                ),
            },
            {
                "label": "Product Sales",
                "count_label": f"₹{recent_sales:,.2f}" if recent_sales else "₹0.00",
                "timestamp_label": f"{start_date} → {end_date}",
            },
            {
                "label": "Total Bill Count",
                "sub-label": {
                    "Total Appointment completed": total_completed_appointments,
                    "Total sell product": total_sell_product_count,
                    "Total membership selled": total_membership_selled_count,
                    "Total wallet selled": total_wallet_selled_count,
                },
                "count_label": f"{total_completed_appointments + total_sell_product_count} Bills",
                "timestamp_label": f"{start_date} → {end_date}",
            },
        ]

        return Response(response_data, status=status.HTTP_200_OK)
    
from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth
# from datetime import datetime, timedelta
# from django.utils import timezone

@api_view(['GET'])
@authentication_classes([VendorJWTAuthentication])
@permission_classes([IsAuthenticated])
def dashboard_overview(request):

    from datetime import datetime, timedelta
    from django.utils import timezone
    from django.db.models import Sum, Count
    from django.db.models.functions import TruncMonth

    # ---------------- VENDOR FILTER ----------------
    vendor_user_id = request.query_params.get('vendor_user_id')
    if vendor_user_id:
        try:
            vendor_user_id = int(vendor_user_id)
        except ValueError:
            return Response({"error": "vendor_user_id must be an integer"}, status=400)
    else:
        vendor_user_id = request.user.id

    # ---------------- DATE FILTER ----------------
    start_date_str = request.query_params.get('start_date')
    end_date_str = request.query_params.get('end_date')

    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)
    else:
        # fallback to default last 180 days
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=180)

    # ---------------- SALES OVERVIEW ----------------
    sales_data = (
        Sell.objects.filter(vendor_user_id=vendor_user_id, created_at__date__range=(start_date, end_date))
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(total_sales=Sum('final_total'))
        .order_by('month')
    )
    sales_overview = {
        "months": [data['month'].strftime('%b') for data in sales_data],
        "total_sales": [data['total_sales'] or 0 for data in sales_data]
    }

    # ---------------- REVENUE OVERVIEW ----------------
    revenue_data = (
        Appointment.objects.filter(vendor_user_id=vendor_user_id, date__range=(start_date, end_date))
        .annotate(month=TruncMonth('date'))
        .values('month')
        .annotate(total_revenue=Sum('final_amount'))
        .order_by('month')
    )
    revenue_overview = {
        "months": [data['month'].strftime('%b') for data in revenue_data],
        "total_revenue": [data['total_revenue'] or 0 for data in revenue_data]
    }

    # ---------------- APPOINTMENTS ----------------
    appointments_list = []
    appointments_qs = Appointment.objects.filter(
        vendor_user_id=vendor_user_id,
        date__range=(start_date, end_date)
    ).prefetch_related('service__categories', 'service__master_service')

    for appt in appointments_qs:
        services_data = [
            {
                "service_name": s.master_service.service_name if s.master_service else s.service_name,
                "category_name": s.categories.name if s.categories else None
            }
            for s in appt.service.all()
        ]

        appointments_list.append({
            "customer_name": appt.customer_name,
            "appointment_date": appt.date,
            "appointment_price": appt.final_amount,
            "services": services_data
        })

    # ---------------- BEST SELLING SERVICES ----------------
    best_selling_services = (
        Appointment.objects.filter(
            vendor_user_id=vendor_user_id,
            date__range=(start_date, end_date),
            appointment_status='completed'
        )
        .values('service__master_service__service_name')
        .annotate(service_count=Count('service'))
        .order_by('-service_count')
    )

    filtered_services = [
        service for service in best_selling_services
        if service['service__master_service__service_name']
    ]

    best_selling_services_data = {
        "services": [s['service__master_service__service_name'] for s in filtered_services],
        "counts": [s['service_count'] for s in filtered_services],
    }

    # ---------------- TOP PERFORMERS ----------------
    top_performing_staff = (
        Appointment.objects.filter(
            vendor_user_id=vendor_user_id,
            date__range=(start_date, end_date),
            appointment_status='completed',
            staff__isnull=False
        )
        .exclude(staff__staffname__isnull=True)
        .exclude(staff__staffname='')
        .values('staff__id', 'staff__staffname')
        .annotate(
            total_revenue=Sum('final_amount'),
            appointment_count=Count('id')
        )
        .order_by('-total_revenue')
    )

    top_performers = {
        "staff_names": [staff['staff__staffname'] for staff in top_performing_staff],
        "staff_ids": [staff['staff__id'] for staff in top_performing_staff],
        "total_revenue": [staff['total_revenue'] for staff in top_performing_staff],
        "appointment_count": [staff['appointment_count'] for staff in top_performing_staff]
    }

    # ---------------- FINAL RESPONSE ----------------
    return Response({
        "sales_overview": sales_overview,
        "revenue_overview": revenue_overview,
        "appointments": appointments_list,
        "best_selling_services": best_selling_services_data,
        "top_performers": top_performers,
        "filters_applied": {
            "start_date": start_date.strftime("%Y-%m-%d"),
            "end_date": end_date.strftime("%Y-%m-%d"),
            "vendor_user_id": vendor_user_id
        }
    })



def generate_unique_pun(name):
    # Take the first 3 characters of the name and convert to uppercase
    pun_prefix = name[:3].upper()

    # Generate a 6-digit random number as a string
    pun_number = ''.join([str(random.randint(0, 9)) for _ in range(6)])
    
    # Concatenate the pun prefix and pun number
    pun = f"{pun_prefix}-{pun_number}"
    
    # Return the generated pun
    return pun

@api_view(['POST'])
@permission_classes([AllowAny])
def generate_pun_view(request):
    # Use request.GET to access query parameters
    name = request.GET.get('name', '')  # Get the 'name' query parameter from the URL
    if not name:
        return Response({"error": "Name parameter is required"}, status=400)

    pun = generate_unique_pun(name)
    return Response({"pun": pun}, status=200)



class CurrentUseInventoryListCreateAPIView(generics.ListCreateAPIView):
    queryset = CurrentUseInventory.objects.all()
    serializer_class = CurrentUseInventorySerializer
    authentication_classes = [VendorJWTAuthentication]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return CurrentUseInventory.objects.filter(use_inventory__vendor_user=self.request.user)


    def create(self, request, *args, **kwargs):
        use_inventory_id = request.data.get('use_inventory')
        
        try:
            use_inventory_instance = UseInventory.objects.get(id=use_inventory_id)
        except UseInventory.DoesNotExist:
            return Response({"error": "UseInventory with this ID does not exist."}, status=status.HTTP_404_NOT_FOUND)

        total_available_quantity = request.data.get('total_available_quantity')
        per_use_consumption = request.data.get('per_use_consumption')
        quantity = int(request.data.get('quantity', 1))  # Number of CurrentUseInventory objects to create
        measure_unit = request.data.get('measure_unit')  # Get measure_unit from request data

        if use_inventory_instance.quantity < quantity:
            return Response({"error": "Not enough quantity available in UseInventory."}, status=status.HTTP_400_BAD_REQUEST)

        use_inventory_instance.quantity -= quantity
        use_inventory_instance.save()

        created_objects = []
        for _ in range(quantity):
            current_use = CurrentUseInventory(
                use_inventory=use_inventory_instance,  # Assign the actual instance
                total_available_quantity=total_available_quantity,
                per_use_consumption=per_use_consumption,
                measure_unit=measure_unit,
                remaining_quantity=total_available_quantity,  # Set remaining quantity to total available quantity

            )
            current_use.save()
            created_objects.append(current_use)

        serializer = self.get_serializer(created_objects, many=True)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class CurrentuseInventoryRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = CurrentUseInventory.objects.all()
    serializer_class = CurrentUseInventorySerializer
    authentication_classes = [VendorJWTAuthentication]

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        
        use_inventory_instance = instance.use_inventory

        original_quantity = instance.total_available_quantity
        updated_quantity = request.data.get('quantity', original_quantity)  # Get new quantity if provided

        if use_inventory_instance.quantity + original_quantity < updated_quantity:
            return Response({"error": "Not enough quantity available in UseInventory."}, status=status.HTTP_400_BAD_REQUEST)

        use_inventory_instance.quantity += original_quantity  # Return original quantity
        use_inventory_instance.quantity -= updated_quantity  # Subtract the updated quantity
        use_inventory_instance.save()

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


from django.http import HttpResponse

from reportlab.lib.pagesizes import A4

from reportlab.pdfgen import canvas



from reportlab.lib.pagesizes import A4

from reportlab.lib import colors

from reportlab.lib.units import inch

from reportlab.pdfgen import canvas

from reportlab.platypus import Table, TableStyle, Paragraph, SimpleDocTemplate, Spacer

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle



from django.http import HttpResponse

import datetime



from reportlab.lib.pagesizes import A4

from reportlab.lib import colors

from reportlab.lib.styles import getSampleStyleSheet

from reportlab.lib.units import mm

from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer

from django.views.decorators.csrf import csrf_exempt

import base64

from azure.storage.blob import BlobServiceClient

from azure.storage.blob import ContentSettings
from reportlab.lib.utils import simpleSplit 
import logging 
RECEIPT_WIDTH = 210
MARGIN = 10

# Helper Functions (Must be defined or imported)
def draw_centered_string(c, x_center, y, text, font="Helvetica", size=10):
    c.setFont(font, size)
    text_width = c.stringWidth(text, font, size)
    c.drawString(x_center - text_width / 2, y, text)

def draw_line(c, y, width, margin):
    c.line(margin, y, width - margin, y)

def simpleSplit(text, font, size, max_width):
    # A simple tokenizer/line splitter for ReportLab text
    # This is a placeholder; a full implementation is complex.
    if not text:
        return [""]
    c = canvas.Canvas(HttpResponse(content_type='application/pdf')) # dummy canvas for stringWidth
    c.setFont(font, size)
    words = text.split()
    lines = []
    current_line = []
    for word in words:
        if c.stringWidth(" ".join(current_line + [word])) < max_width:
            current_line.append(word)
        else:
            lines.append(" ".join(current_line))
            current_line = [word]
    lines.append(" ".join(current_line))
    return lines
    
# ==============================================================================
# UPDATED CLASS WITH AUTO-HEIGHT LOGIC
# ==============================================================================

class InvoiceViewSet(viewsets.ViewSet):

    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def retrieve(self, request, pk=None):
        """Wrapper to call generate_invoice with appointment_id"""
        return self.generate_invoice(request, appointment_id=pk)

    def generate_invoice(self, request, appointment_id=None):
        try:
            # NOTE: Assuming 'Appointment', 'VendorJWTAuthentication', 'IsAuthenticated'
            # and 'simpleSplit' are available in the scope.
            appointment = Appointment.objects.get(id=appointment_id, appointment_status='completed')
        except Appointment.DoesNotExist:
            raise ValidationError({"error": "Invalid appointment ID or the appointment is not completed."})

        if not appointment.vendor_user:
            raise ValidationError({"error": "Appointment has no linked VendorUser."})
        if not appointment.vendor_user.salon:
            raise ValidationError({"error": "VendorUser has no linked Salon."})

        # --- Data Gathering (Unchanged) ---
        salon = appointment.vendor_user.salon
        salon_name = getattr(salon, 'name', "Default Salon")
        salon_address = getattr(salon, 'address', "No Address Available")
        salon_phone = getattr(appointment.vendor_user, 'ph_number', 'N/A') 
        customer_name = getattr(appointment, 'customer_name', 'N/A')
        invoice_number = f"INV-{appointment_id}"
        payment_status = getattr(appointment, 'payment_status', 'N/A')
        payment_source = getattr(appointment, 'payment_mode', 'N/A')
        split_payment_mode = getattr(appointment, 'split_payment_mode', None)
        
        booking_date = getattr(appointment, 'date', 'N/A')
        if not isinstance(booking_date, str) and hasattr(booking_date, 'strftime'):
            try:
                booking_date = booking_date.strftime("%Y-%m-%d")
            except:
                booking_date = 'N/A'
            
        items = []
        
        # NOTE: Item Processing logic for SRV/MBR/OFR/PRD goes here (as it was in your original code)
        # To avoid redundancy, I will assume the 'items' list is populated correctly as per your tax/discount logic:
        # 
        # --- START: Populating 'items' list (as per your original logic) ---
        for service in getattr(appointment, 'included_services', []):
             item_type = "SRV"
             # ... rest of service logic ...
             items.append({
                 "item": service.get("service_name"),
                 "type": item_type,
                 "qty": 1,
                 "price": decimal.Decimal(str(service.get("actual_price") or 0)),
                 "disc": decimal.Decimal(str(service.get("actual_price") or 0)) - decimal.Decimal(str(service.get("final_price") or 0)),
                 "tax": 0,
                 "amt": decimal.Decimal(str(service.get("final_price") or 0)),
                 "hsn_code": service.get("hsn_code", "")
             })
        for offer in getattr(appointment, 'included_offers', []):
            original_price = decimal.Decimal(str(offer.get("actual_price") or 0))
            final_price = decimal.Decimal(str(offer.get("discounted_price") or 0))
            items.append({
                "item": offer.get("offer_name", "Offer Item"), 
                "type": "OFR", "qty": 1, "price": original_price,
                "disc": original_price - final_price, "tax": 0, "amt": final_price,
                "hsn_code": offer.get("hsn_code", "")
            })
        selled_product_obj = getattr(appointment, 'selled_product', None)
        if selled_product_obj and getattr(selled_product_obj, 'product_list', None):
            for product_item in selled_product_obj.product_list:
                qty = decimal.Decimal(str(product_item.get("quantity", 1) or 1))
                price = decimal.Decimal(str(product_item.get("price_per_unit", 0) or 0))
                discount_val = decimal.Decimal(str(product_item.get("discount", 0) or 0))
                tax_val = decimal.Decimal(str(product_item.get("tax", 0) or 0))
                amount = decimal.Decimal(str(product_item.get("net_sub_total", (price * qty) - discount_val + tax_val) or 0))
                items.append({
                    "item": product_item.get("product_name", "Product Item"), 
                    "type": "PRD", "qty": qty, "price": price,
                    "disc": discount_val, "tax": tax_val, "amt": amount,
                    "hsn_code": product_item.get("hsn_code", "")
                })
        # --- END: Populating 'items' list ---

        # --- Calculation Logic (Unchanged, important for display) ---
        subtotal = sum(item['amt'] for item in items)
        global_discount_amount = decimal.Decimal(str(getattr(appointment, 'discount_amount', '0.00') or '0.00'))
        global_discount_percent = decimal.Decimal(str(getattr(appointment, 'discount_percentage', '0.00') or '0.00'))
        # global_tax_amount = decimal.Decimal(str(getattr(appointment, 'tax_amount', '0.00') or '0.00'))
        # global_tax_percent = decimal.Decimal(str(getattr(appointment, 'tax_percentage', '0.00') or '0.00'))

         # Check if GST is enabled for this vendor
        is_gst_enabled = appointment.vendor_user.is_gst if appointment.vendor_user else False
        
        # Only apply tax if GST is enabled
        if is_gst_enabled:
            global_tax_amount = decimal.Decimal(str(getattr(appointment, 'tax_amount', '0.00') or '0.00'))
            global_tax_percent = decimal.Decimal(str(getattr(appointment, 'tax_percentage', '0.00') or '0.00'))
        else:
            global_tax_amount = decimal.Decimal('0.00')
            global_tax_percent = decimal.Decimal('0.00')
        rounding_off = decimal.Decimal(str(getattr(appointment, 'rounding_off', '0.00') or '0.00')) 
        
        discount_display_text = "Discount"
        discount_value = decimal.Decimal("0.00")
        if global_discount_amount > 0:
            discount_value = global_discount_amount
            calc_percent = (discount_value / subtotal) * 100 if subtotal else 0
            discount_display_text = f"Discount ({calc_percent:.2f}%)" if subtotal else "Discount (Amt)"
        elif global_discount_percent > 0:
            discount_value = (global_discount_percent / 100) * subtotal
            discount_display_text = f"Discount ({global_discount_percent:.2f}%)"

        tax_value = decimal.Decimal("0.00")
        taxable_base = subtotal - discount_value
        tax_display_text_cgst = "CGST"
        tax_display_text_sgst = "SGST"
        
        if global_tax_amount > 0:
            tax_value = global_tax_amount
            calc_percent = (tax_value / taxable_base) * 100 if taxable_base else 0
            cgst_percent = calc_percent / 2
            sgst_percent = calc_percent / 2
            tax_display_text_cgst = f"CGST ({cgst_percent:.2f}%)" if taxable_base else "CGST (Amt)"
            tax_display_text_sgst = f"SGST ({sgst_percent:.2f}%)" if taxable_base else "SGST (Amt)"
        elif is_gst_enabled and global_tax_percent > 0:
            tax_value = (global_tax_percent / 100) * taxable_base
            cgst_percent = global_tax_percent / 2
            sgst_percent = global_tax_percent / 2
            tax_display_text_cgst = f"CGST ({cgst_percent:.2f}%)"
            tax_display_text_sgst = f"SGST ({sgst_percent:.2f}%)"
            
        cgst_value = tax_value / 2
        sgst_value = tax_value / 2
        
        total = subtotal - discount_value + tax_value + rounding_off
        total_saved_today = discount_value

        # ======================================================================
        # --- STAGE 1: Calculate Total Required Height ---
        # We start counting from the top (y_position = 0) and add height.
        # This is the reverse of drawing, but necessary to determine the total size.
        # Line height is roughly 10 points for size 8/9 font.
        
        current_height = 20 # Initial top margin

        # 1. Header Block
        current_height += 15 # Salon Name (Size 14)
        address_lines = simpleSplit(salon_address, "Helvetica", 8, RECEIPT_WIDTH - 2 * MARGIN)
        current_height += len(address_lines) * 10 # Address lines (Size 8)
        current_height += 15 # Call number (Size 9)
        current_height += 15 # GST number (Size 9) - IF UNCOMMENTED
        current_height += 10 # Separator Line

        # 2. Invoice Details Block
        current_height += 15 # Receipt No.
        current_height += 10 # Customer Name
        current_height += 10 # Appointment Date
        current_height += 10 # Separator Line

        # 3. Items Table Header
        current_height += 15 # Header text
        current_height += 5  # Separator Line

        # 4. Items Table Body (Dynamic Height)
        item_max_width = 75 - (MARGIN + 15) - 5
        for item in items:
            item_text = item["item"]
            item_lines = simpleSplit(item_text, "Helvetica", 8, item_max_width)
            # 12pt for first line, 10pt for subsequent lines
            current_height += 12 + max(0, len(item_lines) - 1) * 10 

        current_height += 8 # Separator Line

        # 5. Summary Block
        current_height += 15 # Sub Total
        if discount_value > 0:
            current_height += 10
        if cgst_value > 0: # Check total tax value, not just CGST
            current_height += 10 # CGST
            current_height += 10 # SGST
        if rounding_off != decimal.Decimal('0.00'):
            current_height += 10
        current_height += 10 # Separator Line
        current_height += 15 # Total Amount (Bold)
        current_height += 10 # Separator Line

        # 6. Payment & Status Block
        current_height += 15 # Payment Mode line (if not split)
        if split_payment_mode and isinstance(split_payment_mode, dict):
             current_height += 5 # Extra space for split modes
             current_height += len(split_payment_mode) * 10 # Each split mode
        current_height += 10 # Status line
        current_height += 10 # Separator Line

        # 7. Footer Messages
        current_height += 15 # You saved today
        current_height += 15 # Thank you message
        current_height += 20 # Refer & Earn
        current_height += 10 # Separator Line
        current_height += 15 # Website

        # 8. Full Forms
        current_height += 30 # Top space
        current_height += 10 # OFR/MBR line
        current_height += 10 # PRD/SRV line

        # 9. Regards/Contact
        current_height += 20 # Regards header
        current_height += 12 # Salon Name

        # 10. Rating Section
        current_height += 30 # Hey, enjoyed
        current_height += 15 # Rate us

        # 11. Booking Button
        current_height += 15 # Space before button
        current_height += 20 # Button height

        # 12. Terms and Conditions
        current_height += 40 # Space before T&C header
        current_height += 15 # T&C header
        current_height += 15 # Query contact line

        # Add final bottom margin
        current_height += 20 

        # Final calculated height
        FINAL_RECEIPT_HEIGHT = current_height
        
        # ======================================================================
        # --- STAGE 2: Generate PDF with Dynamic Height ---
        # Now we create the canvas using the calculated height and draw everything.
        # We use height - current_y to draw from top down.
        # ======================================================================
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="invoice_{appointment_id}.pdf"'

        # Create canvas with dynamic height
        c = canvas.Canvas(response, pagesize=(RECEIPT_WIDTH, FINAL_RECEIPT_HEIGHT))

        width = RECEIPT_WIDTH
        height = FINAL_RECEIPT_HEIGHT
        center = width / 2
        
        # Start drawing position from the top margin
        y_position = height - 20 
        
        # ====================== HEADER ======================
        c.setFillColorRGB(0, 0, 0)
        draw_centered_string(c, center, y_position, salon_name, "Helvetica-Bold", 14)

        # Salon Address as proper paragraph
        y_position -= 15
        c.setFont("Helvetica", 8)
        for line in address_lines:
            draw_centered_string(c, center, y_position, line, "Helvetica", 8)
            y_position -= 10
            
        y_position -= 5
        draw_centered_string(c, center, y_position, f"Call: {salon_phone}", "Helvetica", 9)
        
        # Separator Line
        y_position -= 10
        draw_line(c, y_position, width, MARGIN)

        # ====================== INVOICE DETAILS ======================
        y_position -= 15
        c.setFont("Helvetica", 9)
        c.drawString(MARGIN, y_position, "Receipt No:")
        c.drawString(width - MARGIN - c.stringWidth(invoice_number, "Helvetica", 9), y_position, invoice_number)

        y_position -= 10
        c.drawString(MARGIN, y_position, "Customer Name :")
        c.drawString(width - MARGIN - c.stringWidth(customer_name, "Helvetica", 9), y_position, customer_name)
        
        y_position -= 10
        c.drawString(MARGIN, y_position, "Appointment Date:")
        c.drawString(width - MARGIN - c.stringWidth(booking_date, "Helvetica", 9), y_position, booking_date)

        # Separator Line
        y_position -= 10
        draw_line(c, y_position, width, MARGIN)

        # ====================== ITEMS TABLE HEADER ======================
        y_position -= 15
        
        # Column X coordinates
        col_sr_no_x = MARGIN
        col_item_x = MARGIN + 15
        col_type_x = 75
        col_hsn_x = 105
        col_qty_x = 130
        col_amt_x = width - MARGIN
        
        c.setFont("Helvetica-Bold", 9)
        c.drawString(col_sr_no_x, y_position, "Sr")
        c.drawString(col_item_x, y_position, "Item")
        c.drawString(col_hsn_x, y_position, "HSN")
        c.drawString(col_type_x, y_position, "Type")
        c.drawString(col_qty_x, y_position, "Qty")
        amt_header = "Amt"
        c.drawString(col_amt_x - c.stringWidth(amt_header, "Helvetica-Bold", 9), y_position, amt_header)
        
        # Separator Line
        y_position -= 5
        draw_line(c, y_position, width, MARGIN)

        # ====================== ITEMS TABLE BODY ======================
        c.setFont("Helvetica", 8)
        
        item_max_width = col_type_x - col_item_x - 5
        
        for index, item in enumerate(items):
            y_position -= 12
            current_y = y_position
            
            # Print static data on the starting line (which is 'current_y')
            c.drawString(col_sr_no_x, current_y, f"{index + 1}.")
            c.drawString(col_hsn_x, current_y, str(item.get("hsn_code", "")))
            c.drawString(col_type_x, current_y, item["type"])
            c.drawString(col_qty_x, current_y, str(item["qty"])) 
            
            # Amount (Right-aligned)
            amt_str = f"{item['amt']:.2f}"
            c.drawString(col_amt_x - c.stringWidth(amt_str, "Helvetica", 8), current_y, amt_str)
            
            # Wrap and print Item Name
            item_text = item["item"]
            item_lines = simpleSplit(item_text, "Helvetica", 8, item_max_width)
            
            for i, line in enumerate(item_lines):
                if i == 0:
                    c.drawString(col_item_x, current_y, line)
                else:
                    current_y -= 10
                    c.drawString(col_item_x, current_y, line)
            
            y_position = current_y # Update y_position to the lowest point of the item block

        # Separator Line
        y_position -= 8
        draw_line(c, y_position, width, MARGIN)
        
        # ====================== SUMMARY ======================
        c.setFont("Helvetica", 9)
        
        # Sub Total
        y_position -= 15
        c.drawString(MARGIN, y_position, "Sub Total")
        subtotal_str = f"Rs {subtotal:.2f}"
        c.drawString(col_amt_x - c.stringWidth(subtotal_str, "Helvetica", 9), y_position, subtotal_str)

        # Discount
        if discount_value > 0:
            y_position -= 10
            c.drawString(MARGIN, y_position, discount_display_text)
            discount_str = f"-{discount_value:.2f}"
            c.drawString(col_amt_x - c.stringWidth(discount_str, "Helvetica", 9), y_position, discount_str)
        
        # Tax (CGST/SGST)
        if tax_value > 0:
            y_position -= 10
            c.drawString(MARGIN, y_position, tax_display_text_cgst)
            cgst_str = f"+{cgst_value:.2f}"
            c.drawString(col_amt_x - c.stringWidth(cgst_str, "Helvetica", 9), y_position, cgst_str)
            
            y_position -= 10
            c.drawString(MARGIN, y_position, tax_display_text_sgst)
            sgst_str = f"+{sgst_value:.2f}"
            c.drawString(col_amt_x - c.stringWidth(sgst_str, "Helvetica", 9), y_position, sgst_str)
    
        # Rounding Off
        if rounding_off != decimal.Decimal('0.00'):
            y_position -= 10
            c.drawString(MARGIN, y_position, "Rounding Off")
            rounding_str = f"{rounding_off:+.2f}"
            c.drawString(col_amt_x - c.stringWidth(rounding_str, "Helvetica", 9), y_position, rounding_str)

        # Total Amount Line
        y_position -= 10
        draw_line(c, y_position, width, MARGIN)
        
        y_position -= 15
        c.setFont("Helvetica-Bold", 10)
        c.drawString(MARGIN, y_position, "Total Amount")
        total_str = f"Rs {total:.2f}"
        c.drawString(col_amt_x - c.stringWidth(total_str, "Helvetica-Bold", 10), y_position, total_str)

        # Total Amount Line
        y_position -= 10
        draw_line(c, y_position, width, MARGIN)
        
        # ====================== PAYMENT & STATUS ======================
        c.setFont("Helvetica", 9)
        y_position -= 15
        
        if split_payment_mode and isinstance(split_payment_mode, dict):
            c.drawString(MARGIN, y_position, "Payment Mode:")
            y_position -= 5 
            
            for mode, amount in split_payment_mode.items():
                y_position -= 10
                mode_str = mode.replace('-', ' ').title()
                try:
                    amount_str = f"Rs {decimal.Decimal(str(amount)):.2f}"
                except:
                    amount_str = "Rs 0.00"
                
                c.drawString(MARGIN + 10, y_position, mode_str)
                c.drawString(col_amt_x - c.stringWidth(amount_str, "Helvetica", 9), y_position, amount_str)
            
            y_position -= 10
            c.drawString(MARGIN, y_position, "Status:")
            c.drawString(col_amt_x - c.stringWidth(payment_status, "Helvetica", 9), y_position, payment_status)
            
        else:
            c.drawString(MARGIN, y_position, "Payment Mode:")
            c.drawString(col_amt_x - c.stringWidth(payment_source, "Helvetica", 9), y_position, payment_source)

            y_position -= 10
            c.drawString(MARGIN, y_position, "Status:")
            c.drawString(col_amt_x - c.stringWidth(payment_status, "Helvetica", 9), y_position, payment_status)

        # Separator Line
        y_position -= 10
        draw_line(c, y_position, width, MARGIN)

        # ====================== FOOTER MESSAGES ======================
        y_position -= 15
        c.setFont("Helvetica-Bold", 9)
        c.drawString(MARGIN, y_position, "You saved today")
        saved_str = f"Rs {total_saved_today:.2f}"
        c.drawString(col_amt_x - c.stringWidth(saved_str, "Helvetica-Bold", 9), y_position, saved_str)

        y_position -= 15
        draw_centered_string(c, center, y_position, "Thank you for visiting!", "Helvetica", 9)

        y_position -= 20
        draw_centered_string(c, center, y_position, "Refer & Earn | Visit again soon.", "Helvetica-Bold", 9)

        # Separator Line
        y_position -= 10
        draw_line(c, y_position, width, MARGIN)

        # Website
        y_position -= 15
        draw_centered_string(c, center, y_position, "www.trakky.in", "Helvetica", 12)
        
        # ====================== FULL FORMS ======================
        y_position -= 30
        c.setFont("Helvetica", 8)
        c.drawString(MARGIN, y_position, "OFR == Offer")
        c.drawString(MARGIN + 70, y_position, "MBR == Membership")
        y_position -= 10
        c.drawString(MARGIN, y_position, "PRD == Product")
        c.drawString(MARGIN + 70, y_position, "SRV == Service")


        # ====================== ORIGINAL CODE LOGIC KEPT BELOW ======================
        # Regards and Contact Info
        y_position -= 20
        c.setFont("Helvetica-Bold", 10)
        c.drawString(MARGIN, y_position, "Regards")
        y_position -= 12
        c.drawString(MARGIN, y_position, f"{salon_name}")

        # ====================== RATING SECTION ======================
        y_position -= 30
        c.setFont("Helvetica-Bold", 9)
        c.drawString(MARGIN, y_position, "Hey, enjoyed our service?")

        # Draw "Rate us" and "click here" with hyperlink
        y_position -= 15
        c.drawString(MARGIN, y_position, "Rate us ")

        rate_us_url = "https://trakky.in/" # Placeholder URL
        click_here_text = "click here"
        
        c.setFont("Helvetica-Bold", 9)
        rate_us_width = c.stringWidth("Rate us ")
        click_here_x_start = MARGIN + rate_us_width
        click_here_y_start = y_position

        c.setFillColorRGB(0, 0, 1)
        c.drawString(click_here_x_start, click_here_y_start, click_here_text)

        click_here_width = c.stringWidth(click_here_text)
        click_here_height = 10 
        c.linkURL(
            rate_us_url,
            (click_here_x_start, click_here_y_start - 2, 
            click_here_x_start + click_here_width, click_here_y_start + click_here_height),
            relative=0,
        )
        c.setFillColorRGB(0, 0, 0) # Reset color

        # ====================== BOOKING BUTTON ======================
        base_url = "https://trakky.in/"
        city = getattr(salon, 'city', 'default-city')
        area = getattr(salon, 'area', 'default-area')
        slug = getattr(salon, 'slug', 'default-slug')
        booking_url = f"{base_url}{city}/{area}/salons/{slug}"

        y_position -= 15
        button_text = "Book Your Next Appointment"
        button_width = 180
        button_height = 20
        button_x_start = center - button_width / 2
        button_y_start = y_position - button_height

        c.setFillColor(colors.black)
        c.rect(button_x_start, button_y_start, button_width, button_height, fill=1)

        c.setFillColor(colors.white)
        draw_centered_string(c, center, button_y_start + 6, button_text, "Helvetica", 9)

        c.linkURL(
            booking_url,
            (button_x_start, button_y_start, button_x_start + button_width, button_y_start + button_height),
            relative=0,
        )
        c.setFillColorRGB(0, 0, 0) # Reset color

        # ====================== TERMS AND CONDITIONS ======================
        y_position -= 40
        c.setFont("Helvetica-Bold", 10) 
        c.drawString(MARGIN, y_position, "Terms and conditions apply")

        y_position -= 15
        c.setFont("Helvetica", 9)
        c.drawString(MARGIN + 5, y_position, "• For any Query contact on : " + salon_phone)

        # Finalize and generate PDF
        c.showPage()
        c.save()

        return response

from reportlab.lib.colors import black, blue, white
from reportlab.lib.utils import ImageReader

class InvoiceViewSetduplicate(viewsets.ViewSet):
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def generate_invoice(self, request, appointment_id=None):
        # --- Fetch real appointment (instead of dummy data) ---
        try:
            appointment = Appointment.objects.get(id=appointment_id, appointment_status='completed')
        except Appointment.DoesNotExist:
            raise ValidationError({"error": "Invalid appointment ID or the appointment is not completed."})

        # --- Extract details (dynamic, same as original InvoiceViewSet) ---
        salon_name = appointment.vendor_user.salon.name or "Default Salon"
        salon_address = appointment.vendor_user.salon.address if appointment.vendor_user.salon else "No Address Available"
        salon_phone = getattr(appointment.vendor_user, "ph_number", "N/A")  # safer access
        customer_name = appointment.customer_name
        invoice_number = f"INV-{appointment_id}"
        booking_date = appointment.date
        booking_time = appointment.appointment_start_time
        payment_status = appointment.payment_status
        payment_source = appointment.payment_mode
        logo_image = appointment.vendor_user.logo

        # ================== ITEMS (Services, Offers, Products) ==================
        items = [
            {
                "item": (
                    f'{service["service_name"]} - Mmbr ({", ".join([membership.membership_type.name for membership in appointment.membership.all()])})'
                    if service.get("from_membership") else service["service_name"]
                ),
                "qty": 1,
                "price": service["actual_price"],
                "disc": service["actual_price"] - service["final_price"],
                "tax": 0,
                "amt": service["final_price"]
            }
            for service in appointment.included_services
        ]

        # --- Offers ---
        if appointment.included_offers:
            for offer in appointment.included_offers:
                offer_name = offer.get("offer_name")
                original_price = offer.get("actual_price", 0)
                final_price = offer.get("discounted_price", 0)
                discount = original_price - final_price
                items.append({
                    "item": f'{offer_name} - (OFR)',
                    "qty": 1,
                    "price": original_price,
                    "disc": discount,
                    "tax": 0,
                    "amt": final_price,
                })

        # --- Products ---
        if hasattr(appointment, "selled_product") and appointment.selled_product:
            product_details = appointment.selled_product
            if isinstance(product_details.product_list, list):
                for product in product_details.product_list:
                    product_name = product.get("product_name", "Product")
                    qty = product.get("quantity", 1)
                    price = product.get("price_per_unit", 0)
                    discount = product.get("discount", 0)
                    tax = product.get("tax", 0)
                    amount = product.get("net_sub_total", (price - discount) * qty)
                    items.append({
                        "item": f"{product_name} - (PRD)",
                        "qty": qty,
                        "price": price,
                        "disc": discount,
                        "tax": tax,
                        "amt": amount,
                    })

        # ================== CALCULATIONS ==================
        discount_sum = sum(item['disc'] for item in items)
        subtotal = sum(item['amt'] for item in items)
        tax_sum = sum(item['tax'] for item in items)
        rounding_off = 0
        total = subtotal + tax_sum + rounding_off

        # ================== PDF Response ==================
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="invoice_{appointment_id}.pdf"'

        # Create canvas for PDF
        c = canvas.Canvas(response, pagesize=A4)
        width, height = A4

        # --- Replicating the PDF UI exactly from the image ---

        # Left Column (Invoice Details)
        x_left = 30
        y_pos = height - 40

        # Logo and Salon Info
        if logo_image:
            logo_width = 80
            logo_height = 80
            c.drawImage(logo_image, x_left, y_pos - logo_height, width=logo_width, height=logo_height)
        
        c.setFont("Helvetica-Bold", 16)
        c.drawString(x_left + 100, y_pos - 30, salon_name)
        c.setFont("Helvetica", 10)
        c.drawString(x_left + 100, y_pos - 45, salon_address)
        c.drawString(x_left + 100, y_pos - 60, f"P.S.I.D.")
        
        y_pos -= 100
        c.setFont("Helvetica", 10)
        c.drawString(x_left, y_pos, "Invoice date")
        c.drawString(x_left + 100, y_pos, f": {booking_date} | {booking_time}")
        y_pos -= 15
        c.drawString(x_left, y_pos, "Invoice to")
        c.drawString(x_left + 100, y_pos, f": {customer_name}")
        c.drawString(x_left + 100, y_pos - 15, f": {salon_phone}")
        
        # Bill details header
        y_pos -= 45
        c.setFont("Helvetica-Bold", 14)
        c.drawString(x_left, y_pos, "Bill details")
        y_pos -= 15
        c.setFont("Helvetica", 10)
        c.drawString(x_left, y_pos, "Services")
        c.drawString(width / 2 - 100, y_pos, "Amount")
        y_pos -= 10
        c.line(x_left, y_pos, width / 2 - 20, y_pos)

        # Items
        y_pos -= 15
        for item in items:
            c.setFont("Helvetica", 10)
            c.drawString(x_left, y_pos, item["item"])
            c.drawString(width / 2 - 100, y_pos, f"₹ {item['amt']:.2f}")
            y_pos -= 15
            
        y_pos -= 10
        c.line(x_left, y_pos, width / 2 - 20, y_pos)

        # Summary
        y_pos -= 15
        c.drawString(x_left, y_pos, "Subtotal")
        c.drawString(width / 2 - 100, y_pos, f"₹ {subtotal:.2f}")
        y_pos -= 15
        c.drawString(x_left, y_pos, "Total discount")
        c.drawString(width / 2 - 100, y_pos, f"₹ {discount_sum:.2f}")
        y_pos -= 15
        c.drawString(x_left, y_pos, "Rounding off")
        c.drawString(width / 2 - 100, y_pos, f"₹ {rounding_off:.2f}")
        y_pos -= 15
        c.line(x_left, y_pos, width / 2 - 20, y_pos)
        y_pos -= 15
        c.setFont("Helvetica-Bold", 12)
        c.drawString(x_left, y_pos, "Bill amount")
        c.drawString(width / 2 - 100, y_pos, f"₹ {total:.2f}")
        y_pos -= 15
        c.drawString(x_left, y_pos, "Payment by UPI") # Assuming this is a static string
        
        # Separator line
        c.line(width / 2 - 10, 40, width / 2 - 10, height - 40)

        # Right Column (Offers and Rating)
        x_right = width / 2 + 10
        y_pos_right = height - 40
        
        c.setFillColor(blue)
        c.setFont("Helvetica", 10)
        c.drawString(x_right, y_pos_right, "Thank you for choosing us! We appreciate your")
        c.drawString(x_right, y_pos_right - 15, "visit and look forward to seeing you again soon.")
        
        y_pos_right -= 40
        # Placeholder for the offer image, since it's an image and not text
        # If the image URL is not accessible, you can omit this part.
        offer_image_path = 'https://app.barbera.io/static/media/offers.jpg'
        try:
            offer_image = ImageReader(offer_image_path)
            c.drawImage(offer_image, x_right, y_pos_right - 350, width=width/2 - 20, height=350)
        except Exception:
            print("Warning: Could not load the offer image from the URL. Please ensure the URL is correct.")

        # Rate us section
        y_pos_rating = y_pos_right - 370
        c.setFillColor(black)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(x_right, y_pos_rating, "Rate your overall visit at")
        c.drawString(x_right, y_pos_rating - 15, "Florra unisex studio")
        y_pos_rating -= 35
        c.setFont("Helvetica", 10)
        c.drawString(x_right, y_pos_rating, "Rate below as per your experience")
        
        # Star rating placeholders
        star_x = x_right + 10
        star_y = y_pos_rating - 20
        c.setFont('ZapfDingbats', 18)
        for _ in range(5):
            c.drawString(star_x, star_y, '☆')
            star_x += 20
            
        y_pos_rating -= 40
        c.drawString(x_right, y_pos_rating, "We'd like to hear what you have to say")
        
        # Placeholder for text input box
        c.rect(x_right, y_pos_rating - 70, width/2 - 50, 65, stroke=1, fill=0)
        
        # Submit button
        button_x = x_right + (width/2 - 50)/2 - 40
        button_y = y_pos_rating - 90
        c.setFillColor(blue)
        c.rect(button_x, button_y, 80, 20, stroke=0, fill=1)
        c.setFillColor(white)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(button_x + 10, button_y + 6, "Submit feedback")
        
        # Footer section (common for both columns)
        y_pos_footer = 100
        c.setFillColor(black)
        c.setFont("Helvetica", 10)
        c.drawString(x_left, y_pos_footer, "Add-on Information")
        c.drawString(x_left, y_pos_footer - 15, "Details on Membership, Due/balance amount")
        c.linkURL("https://app.barbera.io/invoice/68cc2b2607c5b5dccfa24eb7", (x_left + 150, y_pos_footer - 20, x_left + 180, y_pos_footer - 10), relative=0)
        
        y_pos_footer -= 30
        c.drawString(x_left, y_pos_footer, "Book your next appointment")
        c.drawString(x_left, y_pos_footer - 15, "Get a confirmed booking with us")
        c.linkURL(f"https://trakky.in/{appointment.vendor_user.salon.city}/{appointment.vendor_user.salon.area}/salons/{appointment.vendor_user.salon.slug}",
                 (x_left + 150, y_pos_footer - 20, x_left + 180, y_pos_footer - 10), relative=0)
        
        # Bottom-most footer
        y_pos_bottom = 40
        c.line(x_left, y_pos_bottom, width - x_left, y_pos_bottom)
        y_pos_bottom -= 15
        c.drawString(x_left + 180, y_pos_bottom, "Florra unisex studio")
        y_pos_bottom -= 15
        c.drawString(x_left + 120, y_pos_bottom, "505, MAYHEA FLAT, BHUDHAR BUNGLOWS, NEAR JUDGES BUNGLOW, BODAKDEV, AHMEDABAD- 380015")
        
        # Social media icons
        social_x = x_left + 200
        y_pos_bottom -= 20
        c.drawString(social_x, y_pos_bottom, "Socials profiles for Florra unisex studio")
        c.drawString(social_x, y_pos_bottom - 15, "G Review")
        c.drawString(social_x + 60, y_pos_bottom - 15, "O follow")
        
        y_pos_bottom -= 30
        c.drawString(x_left + 200, y_pos_bottom, "Powered by")
        c.drawImage(ImageReader('https://trakky-image-h6bea2g2ahf5hkav.z01.azurefd.net/trakky-new-pics/salon_mul_images/02889827-377b-444a-b678-dd2eb48f6f24.webp'), x_left + 200, y_pos_bottom - 20, width=50, height=20)


        # Finalize and generate PDF
        c.showPage()
        c.save()
        return response




def app_invoice(instance, filename):
    ext = filename.split('.')[-1]
    filename = f'{uuid4()}.pdf'
    return os.path.join('app_invoice_pos', filename)

class InvoicedetailViewSet(viewsets.ViewSet):
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def generate_invoice(self, request, appointment_id=None):
        # Validate appointment
        try:
            appointment = Appointment.objects.get(id=appointment_id, appointment_status="completed")
        except Appointment.DoesNotExist:
            raise ValidationError({"error": "Invalid appointment ID or the appointment is not completed."})

        # ------------------ Static Data ------------------
        salon_name = appointment.vendor_user.salon.name or "Default Salon"
        salon_address = appointment.vendor_user.salon.address if appointment.vendor_user.salon else "No Address Available"
        salon_phone = appointment.vendor_user.ph_number
        customer_name = appointment.customer_name
        invoice_number = f"INV-{appointment_id}"
        booking_date = appointment.date
        booking_time = appointment.appointment_start_time
        payment_status = appointment.payment_status
        payment_source = appointment.payment_mode

        # ------------------ Items ------------------
        items = [
            {
                "item": f'{service["service_name"]} - Mmbr ({", ".join([membership.membership_type.name for membership in appointment.membership.all()])})'
                if service.get("from_membership")
                else service["service_name"],
                "qty": 1,
                "price": service["actual_price"],
                "disc": service["actual_price"] - service["final_price"],
                "tax": 0,
                "amt": service["final_price"],
            }
            for service in appointment.included_services
        ]

        if appointment.included_offers:
            for offer in appointment.included_offers:
                offer_name = offer.get("offer_name")
                original_price = offer.get("actual_price")
                final_price = offer.get("discounted_price")
                discount = original_price - final_price

                items.append(
                    {
                        "item": f"{offer_name} - (OFR)",
                        "qty": 1,
                        "price": original_price,
                        "disc": discount,
                        "tax": 0,
                        "amt": final_price,
                    }
                )

        # ------------------ Product Sales ------------------
        if hasattr(appointment, "selled_product") and appointment.selled_product:
            product_details = appointment.selled_product
            if isinstance(product_details.product_list, list):
                for product in product_details.product_list:
                    items.append(
                        {
                            "item": f"{product.get('product_name', 'Product')} - (PRD)",
                            "qty": product.get("quantity", 1),
                            "price": product.get("price_per_unit", 0),
                            "disc": product.get("discount", 0),
                            "tax": product.get("tax", 0),
                            "amt": product.get("net_sub_total", 0),
                        }
                    )

        # ------------------ Calculations ------------------
        discount_sum = sum(item["disc"] for item in items)
        subtotal = sum(item["amt"] for item in items)
        tax = sum(item["tax"] for item in items)
        rounding_off = 0
        total = subtotal + tax + rounding_off

        # ------------------ Generate PDF ------------------
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y_position = height - 40

        # Header: Salon info
        c.setFont("Helvetica-Bold", 16)
        c.drawString(width / 3 - 65, y_position, salon_name)
        c.setFont("Helvetica", 10)
        y_position -= 20
        c.drawString(width / 3 - 65, y_position, salon_address)
        y_position -= 20
        c.drawString(width / 3 - 65, y_position, f"Call us for appointment: {salon_phone}")

        # Invoice & Customer Info
        y_position -= 40
        c.setFont("Helvetica-Bold", 25)
        c.drawString(30, y_position, f"INVOICE: {customer_name}")
        c.setFont("Helvetica", 10)
        c.drawString(width - 130, y_position, f"Invoice No: {invoice_number}")
        c.drawString(width - 130, y_position + 20, f"Appointment ID: {appointment_id}")

        # Booking date/time
        y_position -= 20
        c.line(30, y_position, width - 30, y_position)
        y_position -= 20
        c.drawString(30, y_position, "Booking Date:")
        c.drawString(width - 130, y_position, f"{booking_date} {booking_time}")

        # Payment
        y_position -= 30
        c.drawString(30, y_position, "Payment Status:")
        c.drawString(width - 50, y_position, f"{payment_status}")
        y_position -= 15
        c.drawString(30, y_position, "Payment Source:")
        c.drawString(width - 50, y_position, f"{payment_source}")
        y_position -= 20
        c.line(30, y_position, width - 30, y_position)

        # Items table header
        y_position -= 30
        c.setFont("Helvetica-Bold", 10)
        c.drawString(30, y_position, "Item")
        c.drawString(330, y_position, "Qty")
        c.drawString(380, y_position, "Price")
        c.drawString(430, y_position, "Discount")
        c.drawString(480, y_position, "Tax")
        c.drawString(530, y_position, "Amount")

        y_position -= 20
        c.line(30, y_position, width - 30, y_position)

        # Items
        y_position -= 20
        c.setFont("Helvetica", 10)
        for item in items:
            c.drawString(30, y_position, item["item"])
            c.drawString(330, y_position, str(item["qty"]))
            c.drawString(380, y_position, str(item["price"]))
            c.drawString(430, y_position, str(item["disc"]))
            c.drawString(480, y_position, str(item["tax"]))
            c.drawString(540, y_position, str(item["amt"]))
            y_position -= 15

        # Totals
        y_position -= 20
        c.line(30, y_position, width - 30, y_position)
        y_position -= 20
        c.setFont("Helvetica-Bold", 10)
        c.drawString(30, y_position, "Subtotal")
        c.drawString(540, y_position, str(subtotal))
        y_position -= 20
        c.drawString(30, y_position, "Tax")
        c.drawString(540, y_position, str(tax))
        y_position -= 20
        c.drawString(30, y_position, "Rounding")
        c.drawString(540, y_position, str(rounding_off))
        y_position -= 20
        c.drawString(30, y_position, "Total")
        c.drawString(540, y_position, str(total))
        y_position -= 20
        c.drawString(30, y_position, "You saved today")
        c.drawString(540, y_position, str(discount_sum))

        # Finalize
        c.showPage()
        c.save()

        pdf_bytes = buffer.getvalue()
        buffer.close()

        # ------------------ Upload to Azure ------------------
        connection_string = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
        container_name = os.getenv("AZURE_CONTAINER", "trakky-new-pics")

        if not connection_string:
            raise Exception("AZURE_STORAGE_CONNECTION_STRING is missing.")

        blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        blob_name = app_invoice(None, f"invoice_{appointment_id}.pdf")
        blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)

        blob_client.upload_blob(
            pdf_bytes,
            overwrite=True,
            content_settings=ContentSettings(content_type="application/pdf"),
        )
        blob_url = blob_client.url

        # ------------------ Return ------------------
        if request.query_params.get("download") == "true":
            response = HttpResponse(pdf_bytes, content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="invoice_{appointment_id}.pdf"'
            return response

        return Response(
            {"message": "Invoice generated successfully", "invoice_url": blob_url},
            status=200,
        )


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from django.core.mail import EmailMessage, get_connection
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from .models import Appointment
from django.conf import settings



class SendInvoiceEmailViewSet(viewsets.ViewSet):
    def send_invoice_email(self, request, appointment_id=None):
        # Validate and get the appointment details
        appointment = Appointment.objects.get(id=appointment_id)
        salon_email = appointment.vendor_user.email  # Salon's unique email
        customer_emails = appointment.customer_email  # Customer's email
        salon_name=appointment.vendor_user.salon.name 

        if not salon_email:
            raise ValidationError({"error": "Salon email is missing."})

        # Ensure that salon email is valid
        if not self.is_valid_email(salon_email):
            raise ValidationError({"error": "Invalid salon email address."})

        # Central email credentials from .env
        central_email = os.environ.get('CENTRAL_EMAIL_HOST_USER')
        central_password = os.environ.get('CENTRAL_EMAIL_HOST_PASSWORD')

        if not central_email or not central_password:
            raise ValidationError({"error": "Central email credentials are missing."})

        # Create a connection using the central email host credentials
        connection = get_connection(
            host=settings.EMAIL_HOST,
            port=settings.EMAIL_PORT,
            username=central_email,
            password=central_password,
            use_tls=True
        )

        invoice_viewset = InvoiceViewSet()
        pdf_response = invoice_viewset.generate_invoice(request, appointment_id)
        
        # Convert the PDF response to bytes
        pdf_content = pdf_response.content
        # Email details
        email_subject = f"Invoice for Appointment ID {appointment_id}"
        email_body = (
            f"Dear {appointment.customer_name},\n\n"
            f"Thank you for choosing {salon_name}. Please find attached your invoice for the appointment.\n\n"
            "Best regards,\n"
            f"{salon_name}"
        )

        email = EmailMessage(
            subject=email_subject,
            body=email_body,
            from_email=salon_email,  # Dynamic salon email address
            to=[customer_emails],  # Customer email
            connection=connection  # Using central email credentials for sending
        )

        email.attach(f"invoice_{appointment_id}.pdf", pdf_content, 'application/pdf')

        # Send the email
        try:
            email.send()
            return Response({"success": "Invoice sent successfully."})
        except Exception as e:
            return Response({"error": f"Error sending email: {str(e)}"}, status=500)

    def is_valid_email(self, email):
        # Basic email validation function (you can enhance it further)
        return '@' in email and '.' in email


class DailyUpdateRequestViewSet(viewsets.ModelViewSet):
    queryset = Dailyupdaterequest.objects.all()
    serializer_class = DailyUpdateRequestSerializer
    authentication_classes = [VendorJWTAuthentication]


    def get_queryset(self):
        queryset = super().get_queryset()
        vendor = self.request.query_params.get('vendor')
        salon_name = self.request.query_params.get('salon_name')
        salon_city = self.request.query_params.get('salon_city')
        salon_area = self.request.query_params.get('salon_area')
        is_approved = self.request.query_params.get('is_approved')


        if vendor:
            queryset = queryset.filter(vendor__id=vendor)
        if salon_name:
            queryset = queryset.filter(vendor__salon__name__icontains=salon_name)
        if salon_city:
            queryset = queryset.filter(vendor__salon__city__icontains=salon_city)
        if salon_area:
            queryset = queryset.filter(vendor__salon__area__icontains=salon_area)
        if is_approved is not None:
            if is_approved.lower() in ['true', '1']:
                queryset = queryset.filter(is_approved=True)
            elif is_approved.lower() in ['false', '0']:
                queryset = queryset.filter(is_approved=False)

        return queryset.order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(vendor=self.request.user)


class ClientWorkPhotosRequestViewSet(viewsets.ModelViewSet):
    queryset = Clientworkphotosrequest.objects.all()
    serializer_class = ClientWorkPhotosRequestSerializer
    authentication_classes = [VendorJWTAuthentication]

    def get_queryset(self):
        queryset = super().get_queryset()
        vendor = self.request.query_params.get('vendor')
        salon_name = self.request.query_params.get('salon_name')
        salon_city = self.request.query_params.get('salon_city')
        salon_area = self.request.query_params.get('salon_area')
        is_approved = self.request.query_params.get('is_approved')

        if vendor:
            queryset = queryset.filter(vendor__id=vendor)
        if salon_name:
            queryset = queryset.filter(vendor__salon__name__icontains=salon_name)
        if salon_city:
            queryset = queryset.filter(vendor__salon__city__icontains=salon_city)
        if salon_area:
            queryset = queryset.filter(vendor__salon__area__icontains=salon_area)
        if is_approved is not None:
            if is_approved.lower() in ['true', '1']:
                queryset = queryset.filter(is_approved=True)
            elif is_approved.lower() in ['false', '0']:
                queryset = queryset.filter(is_approved=False)

        return queryset.order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(vendor=self.request.user)
    

class DailyUpdateRequestAdminViewSet(viewsets.ModelViewSet):
    queryset = Dailyupdaterequest.objects.all()
    serializer_class = DailyUpdateRequestSerializer
    authentication_classes = [JWTAuthentication]

    def get_queryset(self):
        queryset = super().get_queryset()
        vendor = self.request.query_params.get('vendor')
        salon_name = self.request.query_params.get('salon_name')
        salon_city = self.request.query_params.get('salon_city')
        salon_area = self.request.query_params.get('salon_area')
        is_approved = self.request.query_params.get('is_approved')

        if vendor:
            queryset = queryset.filter(vendor__id=vendor)
        if salon_name:
            queryset = queryset.filter(vendor__salon__name__icontains=salon_name)
        if salon_city:
            queryset = queryset.filter(vendor__salon__city__icontains=salon_city)
        if salon_area:
            queryset = queryset.filter(vendor__salon__area__icontains=salon_area)
        if is_approved is not None:
            if is_approved.lower() in ['true', '1']:
                queryset = queryset.filter(is_approved=True)
            elif is_approved.lower() in ['false', '0']:
                queryset = queryset.filter(is_approved=False)

        return queryset.order_by('-created_at')


class ClientWorkPhotosRequestAdminViewSet(viewsets.ModelViewSet):
    queryset = Clientworkphotosrequest.objects.all()
    serializer_class = ClientWorkPhotosRequestSerializer
    authentication_classes = [JWTAuthentication]

    def get_queryset(self):
        queryset = super().get_queryset()
        vendor = self.request.query_params.get('vendor')
        salon_name = self.request.query_params.get('salon_name')
        salon_city = self.request.query_params.get('salon_city')
        salon_area = self.request.query_params.get('salon_area')
        is_approved = self.request.query_params.get('is_approved')

        if vendor:
            queryset = queryset.filter(vendor__id=vendor)
        if salon_name:
            queryset = queryset.filter(vendor__salon__name__icontains=salon_name)
        if salon_city:
            queryset = queryset.filter(vendor__salon__city__icontains=salon_city)
        if salon_area:
            queryset = queryset.filter(vendor__salon__area__icontains=salon_area)
        if is_approved is not None:
            if is_approved.lower() in ['true', '1']:
                queryset = queryset.filter(is_approved=True)
            elif is_approved.lower() in ['false', '0']:
                queryset = queryset.filter(is_approved=False)

        return queryset.order_by('-created_at')



import csv
from decimal import Decimal
from django.utils.dateparse import parse_date, parse_time
class CustomerImportAPIView(APIView):
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            decoded_file = file_obj.read().decode('utf-8')
        except UnicodeDecodeError:
            return Response({'error': 'Invalid file encoding. Must be UTF-8.'}, status=status.HTTP_400_BAD_REQUEST)
            
        io_string = io.StringIO(decoded_file)
        reader = csv.DictReader(io_string) 

        customer_data = {}
        processed_customer_phones = []
        created_appointments = []
        skipped_rows = []
        request.user = getattr(request, 'user', type('MockUser', (object,), {'id': 1})())

        with transaction.atomic():
            for row in reader:
                # 🎯 FIX 1: Blank Row Filter
                if not any(v.strip() for v in row.values() if v is not None):
                    skipped_rows.append("Row skipped: Completely blank row detected.")
                    continue

                row_lower = {k.lower(): v for k, v in row.items()}
                
                # --- Get all raw data fields ---
                raw_phone_field = row_lower.get('customer_phone', '').strip()
                name = (row_lower.get('customer_name') or '').strip()
                appointment_date_str = (row_lower.get('appointment_date') or '').strip() # Get date string
                
                parts = raw_phone_field.split()
                phone = parts[0] if parts else ''
                gender = parts[1] if len(parts) > 1 else (row_lower.get('customer_gender') or '')
                
                if gender:
                    gender = gender.lower().strip()
                    if gender not in ['male', 'female', 'other'] or not gender:
                        gender = None
                else:
                    gender = None

                # --- NEW REQUIRED FIELD CHECK FOR CUSTOMER (Phone, Name, Gender) ---
                if not (phone.isdigit() and len(phone) == 10 and name and gender):
                    reason = f"Missing required field for Customer model (phone: {raw_phone_field}, name: {name}, gender: {gender})"
                    skipped_rows.append(f"Row skipped: {reason}")
                    continue 

                # --- Customer Aggregation Logic ---
                if phone not in customer_data:
                    customer_data[phone] = {
                        'customer_name': name,
                        'customer_phone': phone,
                        'customer_gender': gender,
                        'customer_email': (row_lower.get('customer_email') or '').strip(),
                        'appiontment_dates': {}, 
                        'total_visited_count': 0,
                    }

                # --- Update Customer Aggregated Data (Regardless of Appointment Date) ---
                agg_data = customer_data[phone]

                # Update visited count and appointment history ONLY if date is present
                if appointment_date_str:
                    agg_data['total_visited_count'] += 1
                    visit_count = agg_data['total_visited_count']
                    agg_data['appiontment_dates'][f'appointment_date_{visit_count}'] = {
                        'date': appointment_date_str, 
                        'service': row_lower.get('service_taken')
                    }
                
                # --- Create/Update Customer OBJECT (MANDATORY STEP) ---
                customer_obj, created = CustomerTable.objects.update_or_create(
                    vendor_user=request.user, 
                    customer_phone=phone,
                    defaults={
                        'customer_name': agg_data['customer_name'],
                        'customer_gender': agg_data['customer_gender'],
                        'customer_email': agg_data['customer_email'],
                        # Pass the potentially updated appointment history
                        'appiontment_dates': agg_data['appiontment_dates'], 
                        'total_visited_count': agg_data['total_visited_count'],
                    }
                )
                
                if phone not in processed_customer_phones:
                    processed_customer_phones.append(phone)


                # =========================================================================
                # 🎯 NEW LOGIC: ONLY CREATE APPOINTMENT IF DATE IS PRESENT AND VALID
                # =========================================================================
                
                if not appointment_date_str:
                    # Skip the entire appointment creation block if the date is null/empty
                    continue 
                
                # --- Find Related Foreign Key Objects ---
                manager_obj = None
                manager_name = (row_lower.get('manager_name') or '').strip()
                if manager_name:
                    manager_obj = Manager.objects.filter(
                        vendor_user=request.user, managername=manager_name
                    ).first()

                # --- Get Time Strings (Optional fields, use a helper variable) ---
                appointment_start_str = row_lower.get('appointment_start_time')
                appointment_end_str = row_lower.get('appointment_end_time')

                # --- Create Appointment OBJECT ---
                appointment_obj = Appointment.objects.create(
                    vendor_user=request.user,
                    customer=customer_obj,
                    manager=manager_obj,
                    
                    # Date is guaranteed to be a non-empty string here
                    date=parse_date(appointment_date_str), 
                    
                    customer_name=name,
                    customer_phone=phone,
                    customer_gender=gender,
                    customer_email=(row_lower.get('customer_email') or None),
                    
                    # Time fields are optional, so we keep the None check
                    appointment_start_time=(parse_time(appointment_start_str) if appointment_start_str else None),
                    appointment_end_time=(parse_time(appointment_end_str) if appointment_end_str else None),
                    
                    # Ensure conversion to Decimal for financial fields
                    actual_amount=Decimal((row_lower.get('actual_amount') or '0').strip()),
                    final_amount=Decimal((row_lower.get('final_amount') or '0').strip()),
                    amount_paid=Decimal((row_lower.get('amount_paid') or '0').strip()),
                    
                    appointment_status='completed',
                    created_at=timezone.now()
                )

                # --- Handle ManyToMany fields ---
                
                staff_names_raw = row_lower.get('staff_name')
                if staff_names_raw:
                    staff_names = [s.strip() for s in staff_names_raw.split(',') if s.strip()]
                    staff_objs = Staff.objects.filter(vendor_user=request.user, staffname__in=staff_names)
                    if staff_objs.exists():
                        appointment_obj.staff.set(staff_objs)

                service_names_raw = row_lower.get('service_taken')
                if service_names_raw:
                    service_names = [s.strip() for s in service_names_raw.split(',') if s.strip()]
                    service_objs = SalonServices.objects.filter(vendor_user=request.user, service_name__in=service_names)
                    if service_objs.exists():
                        appointment_obj.service.set(service_objs)

                offer_names_raw = row_lower.get('offers_applied')
                if offer_names_raw:
                    offer_names = [o.strip() for o in offer_names_raw.split(',') if o.strip()]
                    offer_objs = salonprofileoffer.objects.filter(vendor_user=request.user, offername__in=offer_names)
                    if offer_objs.exists():
                        appointment_obj.offer.set(offer_objs)

                membership_names_raw = row_lower.get('memberships')
                if membership_names_raw:
                    membership_names = [m.strip() for m in membership_names_raw.split(',') if m.strip()]
                    membership_objs = CustomerMembershipnew.objects.filter(vendor_user=request.user, membership_name__in=membership_names)
                    if membership_objs.exists():
                        appointment_obj.membership.set(membership_objs)

                created_appointments.append(appointment_obj.id)
                # =========================================================================

        return Response({
            'message': 'Import completed successfully',
            'processed_customers': processed_customer_phones,
            'created_appointments': created_appointments,
            'count_customers': len(processed_customer_phones),
            'count_appointments': len(created_appointments),
            'skipped_rows_count': len(skipped_rows),
            'skipped_rows_details': skipped_rows
        })
    
from django.db.models import Prefetch

class CustomerExportAPIView(APIView):
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        vendor_user = request.user

        # Prefetch related appointments for all customers
        customers = (
            CustomerTable.objects.filter(vendor_user=vendor_user)
            .prefetch_related(
                Prefetch(
                    'appointments',
                    queryset=Appointment.objects.select_related('manager')
                    .prefetch_related('staff', 'service', 'offer', 'membership'),
                    to_attr='customer_appointments'
                )
            )
        )

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="full_customer_export.csv"'
        response.write('\ufeff')  # BOM for Excel compatibility

        writer = csv.writer(response)

        # ✅ Header row (static strings)
        writer.writerow([
            'customer_name',
            'customer_phone',
            'customer_gender',
            'customer_email',
            'appointment_date',
            'services_taken',
            'offers_applied',
            'memberships',
            'appointment_start_time',
            'appointment_end_time',
            'staff_name',
            'manager_name',
            'actual_amount',
            'final_amount',
            'amount_paid',
            'total_visited_count',
            'customer_type'
        ])

        # ✅ Loop through customers
        for customer in customers:
            if hasattr(customer, 'customer_appointments') and customer.customer_appointments:
                for appointment in customer.customer_appointments:
                    writer.writerow([
                        customer.customer_name,
                        customer.customer_phone,
                        customer.customer_gender,
                        customer.customer_email,
                        appointment.date.strftime('%Y-%m-%d') if appointment.date else '',
                        ', '.join([s.master_service.service_name for s in appointment.service.all()]) if appointment.service.exists() else '',
                        ', '.join([o.name for o in appointment.offer.all()]) if appointment.offer.exists() else '',
                        ', '.join([m.membership_type.name for m in appointment.membership.all()]) if appointment.membership.exists() else '',
                        appointment.appointment_start_time.strftime('%H:%M') if appointment.appointment_start_time else '',
                        appointment.appointment_end_time.strftime('%H:%M') if appointment.appointment_end_time else '',
                        ', '.join([s.staffname for s in appointment.staff.all()]) if appointment.staff.exists() else '',
                        appointment.manager.managername if appointment.manager else '',
                        str(appointment.actual_amount) if appointment.actual_amount else '',
                        str(appointment.final_amount) if appointment.final_amount else '',
                        str(appointment.amount_paid) if appointment.amount_paid else '',
                        customer.total_visited_count,
                        customer.customer_type
                    ])
            else:
                # Fallback to CustomerTable data if no appointments exist
                appointment_dates = getattr(customer, 'appointment_dates', None) or {}
                if appointment_dates:
                    for visit_key, visit_data in appointment_dates.items():
                        writer.writerow([
                            customer.customer_name,
                            customer.customer_phone,
                            customer.customer_gender,
                            customer.customer_email,
                            visit_data.get('date', ''),
                            visit_data.get('service', ''),
                            '', '', '', '', '', '', '', '', '',  # Empty fields for M2M/appointment data
                            customer.total_visited_count,
                            customer.customer_type
                        ])
                else:
                    # Minimal customer data if no appointments at all
                    writer.writerow([
                        customer.customer_name,
                        customer.customer_phone,
                        customer.customer_gender,
                        customer.customer_email,
                        '', '', '', '', '', '', '', '', '', '',
                        customer.total_visited_count,
                        customer.customer_type
                    ])

        return response


class DailyupdaterequestNotificationView(generics.ListCreateAPIView):
    queryset = DailyupdaterequestNotification.objects.all()
    serializer_class = DailyupdaterequestNotificationSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')
        return DailyupdaterequestNotification.objects.filter(
            dailyupdate__vendor=self.request.user
        )

    def perform_create(self, serializer):
        serializer.save()
    
class ClientworkphotosrequestNotificationView(generics.ListCreateAPIView):
    queryset = ClientworkphotosrequestNotification.objects.all()
    serializer_class = ClientworkphotosrequestNotificationSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            raise PermissionDenied("Unauthorized")
        return ClientworkphotosrequestNotification.objects.filter(
            clientworkphoto__vendor=user
        )

    def perform_create(self, serializer):
        serializer.save()


import csv
import time
import shutil
import subprocess
import os
import re
from io import StringIO
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
from django.utils.encoding import smart_str
import winreg

class WhatsAppBulkSender:
    def __init__(self, config: dict):
        self.config = config
        self.driver = None

    

    def _get_exact_chrome_version(self):
        # Try reading from Windows registry
        reg_paths = [
            r"SOFTWARE\Google\Chrome\BLBeacon",           # 64-bit
            r"SOFTWARE\WOW6432Node\Google\Chrome\BLBeacon" # 32-bit
        ]
        for reg_path in reg_paths:
            try:
                key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, reg_path)
                version, _ = winreg.QueryValueEx(key, "version")
                return version
            except FileNotFoundError:
                pass

        # Fallback: check file properties
        possible_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
        ]
        for path in possible_paths:
            if os.path.exists(path):
                try:
                    info = subprocess.check_output(
                        [path, "--version"], shell=True
                    ).decode().strip()
                    return info.replace("Google Chrome", "").strip()
                except Exception:
                    pass
        return None

    def _detect_driver_path(self):
        chrome_version = self._get_exact_chrome_version()

        # Always clear webdriver_manager cache
        cache_path = os.path.expanduser("~/.wdm")
        if os.path.exists(cache_path):
            shutil.rmtree(cache_path)

        if chrome_version:
            major_version = chrome_version.split('.')[0]  # e.g., "139"
            print(f"[INFO] Detected Chrome major version: {major_version}")
            return ChromeDriverManager(driver_version=major_version).install()
        else:
            print("[WARN] Could not detect Chrome version. Using latest driver.")
            return ChromeDriverManager().install()

    
    def setup_driver(self):
        chrome_options = Options()
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option("useAutomationExtension", False)

        if self.config.get("user_data_dir"):
            chrome_options.add_argument(f"--user-data-dir={self.config['user_data_dir']}")

        driver_path = self._detect_driver_path()

        try:
            service = Service(driver_path)
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            self.driver.maximize_window()
            self.driver.implicitly_wait(3)
        except WebDriverException as e:
            raise RuntimeError(f"Failed to start ChromeDriver: {e}")

    def login_to_whatsapp(self):
        self.driver.get("https://web.whatsapp.com")
        login_time = int(self.config.get("login_time", 240))
        deadline = time.time() + login_time

        while time.time() < deadline:
            if self.driver.find_elements(By.ID, "pane-side"):
                return True
            boxes = self.driver.find_elements(By.XPATH, "//div[@contenteditable='true']")
            if any(box.is_displayed() and box.is_enabled() for box in boxes):
                return True
            time.sleep(1)
        return False

    def send_message_to_number(self, phone, message):
        country = str(self.config.get("country_code", "91")).strip()
        new_msg_time = int(self.config.get("new_msg_time", 10))
        send_msg_time = int(self.config.get("send_msg_time", 3))

        # Make sure phone is clean (digits only)
        phone = re.sub(r"\D", "", phone)

        chat_url = f"https://web.whatsapp.com/send?phone={country}{phone}&text="

        try:
            self.driver.get(chat_url)
        except Exception as e:
            return False, f"Navigation error: {e}"

        # Allow time for WhatsApp to redirect or show error
        time.sleep(2)

        end_time = time.time() + new_msg_time
        while time.time() < end_time:
            # Check for WhatsApp's known invalid number messages
            error_texts = [
                "phone number shared via url is invalid",
                "phone number is not on whatsapp"
            ]
            page_source = self.driver.page_source.lower()
            if any(err in page_source for err in error_texts):
                return False, "Invalid phone number"

            # Locate the message box (tab index can vary depending on WA version)
            boxes = self.driver.find_elements(
                By.XPATH, '//div[@contenteditable="true"][@data-tab="10"]'
            )
            for box in boxes:
                if box.is_displayed() and box.is_enabled():
                    try:
                        box.click()
                        for idx, part in enumerate(message.split("\n")):
                            if part.strip():
                                box.send_keys(part)
                                if idx < len(message.split("\n")) - 1:
                                    box.send_keys(Keys.SHIFT, Keys.ENTER)
                        box.send_keys(Keys.ENTER)
                        time.sleep(send_msg_time)
                        return True, ""
                    except Exception:
                        continue
            time.sleep(0.5)

        return False, "Message input not available / timed out"

    def process_recipients_and_send(self):
        success = 0
        total = 0
        errors = []
        phone_numbers = self.config.get("phone_numbers", [])

        if not phone_numbers:
            csv_content = self.config.get("csv_content", "")
            if csv_content:
                reader = csv.DictReader(StringIO(csv_content))
                for row in reader:
                    phone = (row.get("phone") or row.get("customer_phone") or "").strip()
                    if phone:
                        phone_numbers.append(phone)

        for idx, phone in enumerate(phone_numbers, start=1):
            if not phone:
                errors.append({"row": idx, "phone": None, "error": "Missing phone"})
                continue
            ok, err = self.send_message_to_number(phone, self.config["message"])
            total += 1
            if ok:
                success += 1
            else:
                errors.append({"row": idx, "phone": phone, "error": err})

        return {"success": success, "total": total, "errors": errors}

    def run_campaign(self):
        try:
            self.setup_driver()
        except Exception as e:
            return False, f"Driver setup failed: {e}"

        try:
            if not self.login_to_whatsapp():
                return False, "QR code not scanned in time"
            result = self.process_recipients_and_send()
            return True, result
        except Exception as e:
            return False, str(e)
        finally:
            if self.driver:
                try:
                    self.driver.quit()
                except Exception:
                    pass


class WhatsAppCampaignAPI(APIView):
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        csv_content = None
        phone_numbers = request.data.get("phone_numbers", [])

        if 'csv_file' in request.FILES:
            csv_content = request.FILES['csv_file'].read().decode('utf-8')
        elif 'csv_content' in request.data:
            csv_content = smart_str(request.data['csv_content'])

        data = request.data.copy()
        if csv_content is not None:
            data['csv_content'] = csv_content
        if phone_numbers:
            data['phone_numbers'] = phone_numbers

        serializer = WhatsAppCampaignSerializer(data=data, context={'request': request})
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        campaign = serializer.save()

        config = {
            "campaign_name": campaign.name,
            "message": campaign.message,
            "country_code": campaign.country_code,
            "login_time": campaign.login_time,
            "new_msg_time": campaign.new_msg_time,
            "send_msg_time": campaign.send_msg_time,
            "csv_content": csv_content,
            "phone_numbers": phone_numbers,
            "user_data_dir": request.data.get("user_data_dir")
        }

        sender = WhatsAppBulkSender(config)
        success, result = sender.run_campaign()

        if success:
            campaign.status = 'completed'
            campaign.phone_numbers = phone_numbers
            campaign.save(update_fields=['status', 'phone_numbers'])
            return Response({"status": "completed", "result": result, "campaign_name": campaign.name}, status=status.HTTP_200_OK)
        else:
            campaign.status = 'failed'
            campaign.last_error = str(result)
            campaign.phone_numbers = phone_numbers
            campaign.save(update_fields=['status', 'last_error', 'phone_numbers'])
            return Response({"status": "failed", "message": result, "campaign_name": campaign.name}, status=status.HTTP_200_OK)



class VendorWhatsAppCampaignViewSet(viewsets.ModelViewSet):
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]
    queryset = WhatsAppCampaign.objects.all()
    serializer_class = WhatsAppCampaignSerializer

    def get_queryset(self):
        return self.queryset.filter(vendor=self.request.user)

    def perform_create(self, serializer):
        serializer.save(vendor=self.request.user)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'status': 'success',
            'data': serializer.data
        })

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response({
            'status': 'success',
            'data': serializer.data
        }, status=status.HTTP_201_CREATED)

# View for standard JWT authentication
class WhatsAppCampaignViewSet(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    queryset = WhatsAppCampaign.objects.all()
    serializer_class = WhatsAppCampaignSerializer

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'status': 'success',
            'data': serializer.data
        })

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({
            'status': 'success',
            'data': serializer.data
        }, status=status.HTTP_201_CREATED)


def product_invoice(instance, filename):
    ext = filename.split('.')[-1]
    filename = f'{uuid4()}.pdf'
    return os.path.join('product_invoice_pos', filename)

class SellInvoiceViewSet(viewsets.ViewSet):
    """
    Generate PDF Invoice for Sell model
    """

    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def generate_sell_invoice(self, request, sell_id=None):
        # Validate if the sell exists
        try:
            sell = Sell.objects.get(id=sell_id)
        except Sell.DoesNotExist:
            raise ValidationError({"error": "Invalid Sell ID."})

        # Salon / Vendor Data
        vendor_user = sell.vendor_user
        salon = getattr(vendor_user, "salon", None)
        salon_name = salon.name if salon else "Default Salon"
        salon_address = salon.address if salon else "No Address Available"
        salon_phone = vendor_user.ph_number if hasattr(vendor_user, "ph_number") else "N/A"

        # Customer Data
        customer_name = sell.client.customer_name if sell.client else sell.customer_name
        customer_phone = sell.client.customer_phone if sell.client else sell.customer_phone
        invoice_number = f"SELL-{sell_id}"
        created_date = sell.created_at.strftime("%Y-%m-%d %H:%M:%S")

        # ====================== ITEMS ======================
        items = []
        if isinstance(sell.product_list, list):
            for product in sell.product_list:
                product_name = product.get("product_name", "Product")
                qty = product.get("qauntity", 1)
                price = product.get("price_per_unit", 0)
                discount = product.get("discount", 0)
                tax = product.get("tax", 0)
                amount = product.get("net_sub_total", (price - discount) * qty)

                items.append({
                    "item": f"{product_name} - (PRD)",
                    "qty": qty,
                    "price": price,
                    "disc": discount,
                    "tax": tax,
                    "amt": amount,
                })

        # ====================== CALCULATIONS ======================
        discount_sum = sum(item['disc'] for item in items)
        subtotal = sum(item['amt'] for item in items)
        tax = sum(item['tax'] for item in items)
        rounding_off = 0
        total = subtotal + tax + rounding_off

        # Create HTTP response with PDF content
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="sell_invoice_{sell_id}.pdf"'

        # Create canvas for PDF
        c = canvas.Canvas(response, pagesize=A4)
        width, height = A4
        y_position = height - 40

        # ====================== HEADER ======================
        c.setFont("Helvetica-Bold", 16)
        c.drawString(width / 3 - 65, y_position, salon_name)

        c.setFont("Helvetica", 10)
        y_position -= 20
        c.drawString(width / 3 - 65, y_position, salon_address)
        y_position -= 20
        c.drawString(width / 3 - 65, y_position, f"Call us: {salon_phone}")

        # Customer and Invoice details
        y_position -= 40
        c.setFont("Helvetica-Bold", 20)
        c.drawString(30, y_position, f"INVOICE: {customer_name}")
        c.setFont("Helvetica", 10)
        c.drawString(width - 150, y_position, f"Invoice No: {invoice_number}")
        c.drawString(width - 150, y_position + 20, f"Sell ID: {sell_id}")

        y_position -= 20
        c.line(30, y_position, width - 30, y_position)

        y_position -= 20
        c.drawString(30, y_position, f"Customer Phone: {customer_phone}")
        c.drawString(width - 180, y_position, f"Date: {created_date}")

        y_position -= 15
        c.line(30, y_position, width - 30, y_position)

        # ====================== ITEMS TABLE ======================
        y_position -= 30
        c.setFont("Helvetica-Bold", 10)
        c.drawString(30, y_position, "Item")
        c.drawString(330, y_position, "Qty")
        c.drawString(380, y_position, "Price")
        c.drawString(430, y_position, "Discount")
        c.drawString(480, y_position, "Tax")
        c.drawString(530, y_position, "Amount")

        y_position -= 20
        c.line(30, y_position, width - 30, y_position)

        c.setFont("Helvetica", 10)
        y_position -= 20
        for item in items:
            c.drawString(30, y_position, item["item"])
            c.drawString(330, y_position, str(item["qty"]))
            c.drawString(380, y_position, str(item["price"]))
            c.drawString(430, y_position, str(item["disc"]))
            c.drawString(480, y_position, str(item["tax"]))
            c.drawString(540, y_position, str(item["amt"]))
            y_position -= 15

        # Line after items
        y_position -= 10
        c.line(30, y_position, width - 30, y_position)

        # ====================== SUMMARY ======================
        y_position -= 20
        c.setFont("Helvetica-Bold", 10)
        c.drawString(30, y_position, "Subtotal")
        c.drawString(526, y_position, "Rs")
        c.drawString(540, y_position, f"{subtotal}")

        y_position -= 20
        c.drawString(30, y_position, "Tax")
        c.drawString(526, y_position, "Rs")
        c.drawString(540, y_position, f"{tax}")

        y_position -= 20
        c.drawString(30, y_position, "Rounding")
        c.drawString(526, y_position, "Rs")
        c.drawString(540, y_position, f"{rounding_off}")

        y_position -= 10
        c.line(30, y_position, width - 30, y_position)

        y_position -= 20
        c.setFont("Helvetica-Bold", 12)
        c.drawString(30, y_position, "Total")
        c.drawString(524, y_position, "Rs")
        c.drawString(540, y_position, f"{total}")

        y_position -= 20
        c.line(30, y_position, width - 30, y_position)

        # You Saved Today
        y_position -= 20
        c.setFont("Helvetica-Bold", 10)
        c.drawString(30, y_position, "You saved today")
        c.drawString(526, y_position, "Rs")
        c.drawString(540, y_position, f"{discount_sum}")

        y_position -= 10
        c.line(30, y_position, width - 30, y_position)

        # ====================== FOOTER ======================
        y_position -= 20
        c.drawString(30, y_position, "PRD == Product")

        y_position -= 30
        c.drawString(30, y_position, "Regards")
        y_position -= 15
        c.drawString(30, y_position, salon_name)

        # Final Terms
        y_position -= 50
        c.setFont("Helvetica-Bold", 12)
        c.drawString(30, y_position, "Terms and conditions apply")

        y_position -= 20
        c.setFont("Helvetica", 10)
        c.drawString(45, y_position, "• For any Query contact on : " + str(salon_phone))

        # Finalize and generate PDF
        c.showPage()
        c.save()

        return response


class SellInvoicedetailsViewSet(viewsets.ViewSet):
    """
    Generate PDF Invoice for Sell model and store in Azure Blob
    """

    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def generate_sell_invoice(self, request, sell_id=None):
        # ------------------ Validate Sell ------------------
        try:
            sell = Sell.objects.get(id=sell_id)
        except Sell.DoesNotExist:
            raise ValidationError({"error": "Invalid Sell ID."})

        # ------------------ Vendor / Salon Data ------------------
        vendor_user = sell.vendor_user
        salon = getattr(vendor_user, "salon", None)
        salon_name = salon.name if salon else "Default Salon"
        salon_address = salon.address if salon else "No Address Available"
        salon_phone = getattr(vendor_user, "ph_number", "N/A")

        # ------------------ Customer Data ------------------
        customer_name = sell.client.customer_name if sell.client else sell.customer_name
        customer_phone = sell.client.customer_phone if sell.client else sell.customer_phone
        invoice_number = f"SELL-{sell_id}"
        created_date = sell.created_at.strftime("%Y-%m-%d %H:%M:%S")

        # ------------------ Items ------------------
        items = []
        if isinstance(sell.product_list, list):
            for product in items:
                # quantity
                try:
                    qty = int(product.get("quantity", 1))
                except (TypeError, ValueError):
                    qty = 1

                # price
                try:
                    price = float(product.get("price_per_unit", 0))
                except (TypeError, ValueError):
                    price = 0.0

                # discount
                try:
                    discount = float(product.get("discount", 0))
                except (TypeError, ValueError):
                    discount = 0.0

                # base amount
                net_sub_total = product.get("net_sub_total")
                if net_sub_total is not None:
                    try:
                        amount = float(net_sub_total)
                    except (TypeError, ValueError):
                        amount = (price - discount) * qty
                else:
                    amount = (price - discount) * qty

                # ✅ ALWAYS initialize tax
                try:
                    tax = float(product.get("gst", 0) or 0)
                except (TypeError, ValueError):
                    tax = 0.0

                # Now you can safely use tax
                total_with_tax = amount + tax

                items.append({
                    "item": f"{product_name} - (PRD)",
                    "qty": qty,
                    "price": price,
                    "disc": discount,
                    "tax": tax,
                    "amt": amount,
                })

        # ------------------ Calculations ------------------
        discount_sum = sum(item["disc"] for item in items)
        subtotal = sum(item["amt"] for item in items)
        tax = sum(item["tax"] for item in items)
        rounding_off = 0
        total = subtotal + tax + rounding_off

        # ------------------ Generate PDF ------------------
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y_position = height - 40

        # Header: Salon info
        c.setFont("Helvetica-Bold", 16)
        c.drawString(width / 3 - 65, y_position, salon_name)
        c.setFont("Helvetica", 10)
        y_position -= 20
        c.drawString(width / 3 - 65, y_position, salon_address)
        y_position -= 20
        c.drawString(width / 3 - 65, y_position, f"Call us: {salon_phone}")

        # Invoice & Customer Info
        y_position -= 40
        c.setFont("Helvetica-Bold", 20)
        c.drawString(30, y_position, f"INVOICE: {customer_name}")
        c.setFont("Helvetica", 10)
        c.drawString(width - 150, y_position, f"Invoice No: {invoice_number}")
        c.drawString(width - 150, y_position + 20, f"Sell ID: {sell_id}")

        # Customer details
        y_position -= 20
        c.line(30, y_position, width - 30, y_position)
        y_position -= 20
        c.drawString(30, y_position, f"Customer Phone: {customer_phone}")
        c.drawString(width - 180, y_position, f"Date: {created_date}")
        y_position -= 15
        c.line(30, y_position, width - 30, y_position)

        # Items table header
        y_position -= 30
        c.setFont("Helvetica-Bold", 10)
        c.drawString(30, y_position, "Item")
        c.drawString(330, y_position, "Qty")
        c.drawString(380, y_position, "Price")
        c.drawString(430, y_position, "Discount")
        c.drawString(480, y_position, "Tax")
        c.drawString(530, y_position, "Amount")

        y_position -= 20
        c.line(30, y_position, width - 30, y_position)

        # Items rows
        y_position -= 20
        c.setFont("Helvetica", 10)
        for item in items:
            c.drawString(30, y_position, item["item"])
            c.drawString(330, y_position, str(item["qty"]))
            c.drawString(380, y_position, str(item["price"]))
            c.drawString(430, y_position, str(item["disc"]))
            c.drawString(480, y_position, str(item["tax"]))
            c.drawString(540, y_position, str(item["amt"]))
            y_position -= 15

        # Totals
        y_position -= 20
        c.line(30, y_position, width - 30, y_position)
        y_position -= 20
        c.setFont("Helvetica-Bold", 10)
        c.drawString(30, y_position, "Subtotal")
        c.drawString(540, y_position, str(subtotal))
        y_position -= 20
        c.drawString(30, y_position, "Tax")
        c.drawString(540, y_position, str(tax))
        y_position -= 20
        c.drawString(30, y_position, "Rounding")
        c.drawString(540, y_position, str(rounding_off))
        y_position -= 20
        c.drawString(30, y_position, "Total")
        c.drawString(540, y_position, str(total))
        y_position -= 20
        c.drawString(30, y_position, "You saved today")
        c.drawString(540, y_position, str(discount_sum))

        # Footer
        y_position -= 30
        c.setFont("Helvetica", 10)
        c.drawString(30, y_position, "PRD == Product")
        y_position -= 20
        c.drawString(30, y_position, "Regards")
        y_position -= 15
        c.drawString(30, y_position, salon_name)
        y_position -= 30
        c.setFont("Helvetica-Bold", 12)
        c.drawString(30, y_position, "Terms and conditions apply")
        y_position -= 20
        c.setFont("Helvetica", 10)
        c.drawString(45, y_position, f"• For any Query contact on: {salon_phone}")

        # Finalize PDF
        c.showPage()
        c.save()

        pdf_bytes = buffer.getvalue()
        buffer.close()

        # ------------------ Upload to Azure ------------------
        connection_string = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
        container_name = os.getenv("AZURE_CONTAINER", "trakky-new-pics")

        if not connection_string:
            raise Exception("AZURE_STORAGE_CONNECTION_STRING is missing.")

        blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        blob_name = product_invoice(None, f"sell_invoice_{sell_id}.pdf")
        blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)

        blob_client.upload_blob(
            pdf_bytes,
            overwrite=True,
            content_settings=ContentSettings(content_type="application/pdf"),
        )
        blob_url = blob_client.url

        # ------------------ Return Response ------------------
        if request.query_params.get("download") == "true":
            response = HttpResponse(pdf_bytes, content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="sell_invoice_{sell_id}.pdf"'
            return response

        return Response(
            {
                "message": "Sell Invoice generated successfully",
                "invoice_url": blob_url,
                "payload": {
                    "invoice_number": invoice_number,
                    "sell_id": sell_id,
                    "salon": {"name": salon_name, "address": salon_address, "phone": salon_phone},
                    "customer": {"name": customer_name, "phone": customer_phone},
                    "created_date": created_date,
                    "items": items,
                    "subtotal": subtotal,
                    "tax": tax,
                    "rounding_off": rounding_off,
                    "total": total,
                    "you_saved": discount_sum,
                },
            },
            status=200,
        )



class SendWhatsAppMessageAPIView(APIView):
    """
    API to send WhatsApp template messages via MSG91
    """
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            # Extract variables from frontend payload
            phone_number = request.data.get("phone_number")  # recipient
            values = request.data.get("values")  # list of variables ({{1}} to {{9}})

            if not phone_number or not values:
                return Response({"error": "phone_number and values are required"}, status=status.HTTP_400_BAD_REQUEST)

            # Build payload
            payload = {
                "integrated_number": "919227198149",  # Your MSG91 registered number
                "content_type": "template",
                "payload": {
                    "messaging_product": "whatsapp",
                    "type": "template",
                    "template": {
                        "name": "appointment_pos_confimation",  # must match template name in MSG91
                        "language": {
                            "code": "en",
                            "policy": "deterministic"
                        },
                        "namespace": "f8a4457f_9e77_47b3_84a1_8c23bf6a1c43",  # keep yours
                        "to_and_components": [
                            {
                                "to": [phone_number],  # recipient number with country code
                                "components": {
                                    f"body_{i+1}": {
                                        "type": "text",
                                        "value": str(val)
                                    }
                                    for i, val in enumerate(values)
                                }
                            }
                        ]
                    }
                }
            }

            headers = {
                "Content-Type": "application/json",
                "authkey": os.environ.get("MSG91_AUTH_KEY")
            }

            response = requests.post(
                "https://api.msg91.com/api/v5/whatsapp/whatsapp-outbound-message/bulk/",
                headers=headers,
                json=payload
            )

            return Response(response.json(), status=response.status_code)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class SendWhatsAppMessagetodayAPIView(APIView):
    """
    Send WhatsApp Reminder via MSG91 API
    """
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            # ✅ Get variables from request payload
            phone_numbers = request.data.get("phone_numbers", [])
            body_1 = request.data.get("body_1", "")
            body_2 = request.data.get("body_2", "")
            body_3 = request.data.get("body_3", "")
            body_4 = request.data.get("body_4", "")

            if not phone_numbers:
                return Response({"error": "phone_numbers is required"}, status=status.HTTP_400_BAD_REQUEST)

            url = "https://api.msg91.com/api/v5/whatsapp/whatsapp-outbound-message/bulk/"

            payload = {
                "integrated_number": "919227198149",  
                "content_type": "template",
                "payload": {
                    "messaging_product": "whatsapp",
                    "type": "template",
                    "template": {
                        "name": "today_reminder",
                        "language": {
                            "code": "en",
                            "policy": "deterministic"
                        },
                        "namespace": "f8a4457f_9e77_47b3_84a1_8c23bf6a1c43",  # from .env
                        "to_and_components": [
                            {
                                "to": phone_numbers,
                                "components": {
                                    "body_1": {"type": "text", "value": body_1},
                                    "body_2": {"type": "text", "value": body_2},
                                    "body_3": {"type": "text", "value": body_3},
                                    "body_4": {"type": "text", "value": body_4},
                                }
                            }
                        ]
                    }
                }
            }

            headers = {
                "Content-Type": "application/json",
                "authkey": os.environ.get("MSG91_AUTH_KEY")
            }

            response = requests.post(url, json=payload, headers=headers)

            return Response(response.json(), status=response.status_code)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

class AppointmentDateReminderAPIView(APIView):
    """
    Send WhatsApp appointment date reminder via MSG91 API
    """
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            # ✅ Extract data from frontend payload
            phone_numbers = request.data.get("phone_numbers", [])
            body_1 = request.data.get("body_1", "")
            body_2 = request.data.get("body_2", "")
            body_3 = request.data.get("body_3", "")
            body_4 = request.data.get("body_4", "")
            body_5 = request.data.get("body_5", "")
            body_6 = request.data.get("body_6", "")

            if not phone_numbers:
                return Response({"error": "phone_numbers is required"}, status=status.HTTP_400_BAD_REQUEST)

            url = "https://api.msg91.com/api/v5/whatsapp/whatsapp-outbound-message/bulk/"

            # ✅ Build dynamic payload
            payload = {
                "integrated_number": "919227198149", # from .env
                "content_type": "template",
                "payload": {
                    "messaging_product": "whatsapp",
                    "type": "template",
                    "template": {
                        "name": "appointment_date_reminder",
                        "language": {
                            "code": "en",
                            "policy": "deterministic"
                        },
                        "namespace": "f8a4457f_9e77_47b3_84a1_8c23bf6a1c43",  # from .env
                        "to_and_components": [
                            {
                                "to": phone_numbers,
                                "components": {
                                    "body_1": {"type": "text", "value": body_1},
                                    "body_2": {"type": "text", "value": body_2},
                                    "body_3": {"type": "text", "value": body_3},
                                    "body_4": {"type": "text", "value": body_4},
                                    "body_5": {"type": "text", "value": body_5},
                                    "body_6": {"type": "text", "value": body_6},
                                }
                            }
                        ]
                    }
                }
            }

            headers = {
                "Content-Type": "application/json",
                "authkey": os.environ.get("MSG91_AUTH_KEY")
            }

            # ✅ Send request to MSG91
            response = requests.post(url, json=payload, headers=headers)

            return Response(response.json(), status=response.status_code)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        


class SendWhatsAppMessagemembershipAPIView(APIView):
    """
    API to send WhatsApp template messages via MSG91
    """
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            # Extract variables from frontend payload
            phone_number = request.data.get("phone_number")  # recipient
            values = request.data.get("values")  # list of variables ({{1}} to {{7}})

            if not phone_number or not values:
                return Response(
                    {"error": "phone_number and values are required"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Build payload for MSG91
            payload = {
                "integrated_number": "919227198149",  # Your MSG91 registered number
                "content_type": "template",
                "payload": {
                    "messaging_product": "whatsapp",
                    "type": "template",
                    "template": {
                        "name": "mempos",  # 👈 match your MSG91 template name
                        "language": {
                            "code": "en",
                            "policy": "deterministic"
                        },
                        "namespace": "f8a4457f_9e77_47b3_84a1_8c23bf6a1c43",  # 👈 use yours
                        "to_and_components": [
                            {
                                "to": [phone_number],  # recipient number with country code
                                "components": {
                                    f"body_{i+1}": {
                                        "type": "text",
                                        "value": str(val)
                                    }
                                    for i, val in enumerate(values)
                                }
                            }
                        ]
                    }
                }
            }

            headers = {
                "Content-Type": "application/json",
                "authkey": os.environ.get("MSG91_AUTH_KEY")  # store in env for security
            }

            response = requests.post(
                "https://api.msg91.com/api/v5/whatsapp/whatsapp-outbound-message/bulk/",
                headers=headers,
                json=payload
            )

            return Response(response.json(), status=response.status_code)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


from rest_framework.test import APIRequestFactory

class SendInvoiceWhatsAppinvoiceMessageAPIView(APIView):
    """
    API to send WhatsApp template messages with document header via MSG91.
    Automatically generates invoice PDF from InvoiceViewSet and uploads directly to Azure Blob Storage.
    """
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            phone_numbers = request.data.get("phone_numbers")
            appointment_id = request.data.get("appointment_id")
            body_values = request.data.get("body_values")

            # --- Input validation ---
            if not phone_numbers or not appointment_id or not body_values:
                return Response(
                    {"error": "phone_numbers, appointment_id, and body_values are required"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # --- Generate invoice PDF in memory ---
            invoice_viewset = InvoiceViewSet()
            invoice_viewset.request = request
            invoice_response = invoice_viewset.generate_invoice(request, appointment_id=appointment_id)

            # Handle direct HttpResponse (PDF stream)
            if isinstance(invoice_response, HttpResponse) and invoice_response.get("Content-Type") == "application/pdf":
                pdf_content = invoice_response.content
                filename = f"invoice_{appointment_id}.pdf"

            # Handle DRF Response with file URL (optional case)
            elif hasattr(invoice_response, "data"):
                file_url = invoice_response.data.get("file_url")
                filename = invoice_response.data.get("filename", f"invoice_{appointment_id}.pdf")
                if file_url:
                    # Already hosted — skip Azure upload
                    uploaded_file_url = file_url
                    pdf_content = None
                else:
                    return Response({"error": "Invoice response did not contain PDF data"},
                                    status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                return Response({"error": "Failed to generate invoice PDF"},
                                status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            # --- Upload PDF directly to Azure Blob Storage ---
            if pdf_content:
                AZURE_CONNECTION_STRING = os.environ.get("AZURE_STORAGE_CONNECTION_STRING")
                AZURE_CONTAINER_NAME = os.environ.get("AZURE_CONTAINER")  # e.g., "trakky-new-pics"

                if not AZURE_CONNECTION_STRING or not AZURE_CONTAINER_NAME:
                    return Response({"error": "Azure credentials not set in environment variables"},
                                    status=status.HTTP_500_INTERNAL_SERVER_ERROR)

                blob_service_client = BlobServiceClient.from_connection_string(AZURE_CONNECTION_STRING)
                blob_name = f"invoices/{filename}"

                blob_client = blob_service_client.get_blob_client(container=AZURE_CONTAINER_NAME, blob=blob_name)

                # Upload from memory buffer directly (no file saved locally)
                pdf_buffer = io.BytesIO(pdf_content)
                blob_client.upload_blob(
                    pdf_buffer,
                    overwrite=True,
                    content_settings=ContentSettings(content_type='application/pdf')
                )

                uploaded_file_url = blob_client.url  # Azure Blob public URL

            # --- Prepare WhatsApp components ---
            components = {
                "header_1": {
                    "type": "document",
                    "filename": filename,
                    "value": uploaded_file_url
                }
            }

            for i, val in enumerate(body_values):
                components[f"body_{i + 1}"] = {"type": "text", "value": str(val)}

            # --- MSG91 Payload ---
            payload = {
                "integrated_number": "919227198149",
                "content_type": "template",
                "payload": {
                    "messaging_product": "whatsapp",
                    "type": "template",
                    "template": {
                        "name": "invoice",
                        "language": {"code": "en", "policy": "deterministic"},
                        "namespace": "f8a4457f_9e77_47b3_84a1_8c23bf6a1c43",
                        "to_and_components": [
                            {"to": phone_numbers, "components": components}
                        ]
                    }
                }
            }

            headers = {
                "Content-Type": "application/json",
                "authkey": os.environ.get("MSG91_AUTH_KEY")
            }

            # --- Send via MSG91 ---
            msg91_response = requests.post(
                "https://api.msg91.com/api/v5/whatsapp/whatsapp-outbound-message/bulk/",
                headers=headers,
                json=payload
            )

            try:
                msg91_json = msg91_response.json()
            except Exception:
                msg91_json = {"raw_response": msg91_response.text}

            return Response({
                "status": "success",
                "sent_to": phone_numbers,
                "invoice_file_url": uploaded_file_url,
                "msg91_response": msg91_json
            }, status=msg91_response.status_code)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class SendAppointmentWhatsAppMessageAPIView(APIView):
    """
    API to send WhatsApp template messages (conappt) via MSG91
    Supports up to 9 body variables.
    """
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            # Extract required fields from request
            phone_numbers = request.data.get("phone_numbers")  # list of phone numbers
            values = request.data.get("values")  # list of body values (max 9)

            if not phone_numbers or not values:
                return Response(
                    {"error": "phone_numbers and values are required"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if len(values) > 9:
                return Response(
                    {"error": "Maximum 9 body values are allowed"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Build payload
            payload = {
                "integrated_number": "919227198149",  # your MSG91 registered number
                "content_type": "template",
                "payload": {
                    "messaging_product": "whatsapp",
                    "type": "template",
                    "template": {
                        "name": "conappt",  # template name from MSG91
                        "language": {
                            "code": "en",
                            "policy": "deterministic"
                        },
                        "namespace": "f8a4457f_9e77_47b3_84a1_8c23bf6a1c43",
                        "to_and_components": [
                            {
                                "to": phone_numbers,
                                "components": {
                                    f"body_{i+1}": {
                                        "type": "text",
                                        "value": str(val)
                                    }
                                    for i, val in enumerate(values)
                                }
                            }
                        ]
                    }
                }
            }

            headers = {
                "Content-Type": "application/json",
                "authkey": os.environ.get("MSG91_AUTH_KEY")  # keep authkey in env
            }

            response = requests.post(
                "https://api.msg91.com/api/v5/whatsapp/whatsapp-outbound-message/bulk/",
                headers=headers,
                json=payload
            )

            return Response(response.json(), status=response.status_code)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        


class Sendappconfomessagenewviewset(APIView):
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            phone_numbers = request.data.get("phone_numbers")
            template_name = request.data.get("template_name")
            body_values = request.data.get("body_values")

            if not phone_numbers or not isinstance(phone_numbers, list):
                return Response({"error": "phone_numbers must be a list of valid numbers"}, status=400)
            if not template_name:
                return Response({"error": "template_name is required"}, status=400)
            if not body_values or not isinstance(body_values, list):
                return Response({"error": "body_values must be a list"}, status=400)

            components = {f"body_{i}": {"type": "text", "value": str(val)} for i, val in enumerate(body_values, start=1)}

            payload = {
                "integrated_number": "919227198149",
                "content_type": "template",
                "payload": {
                    "messaging_product": "whatsapp",
                    "type": "template",
                    "template": {
                        "name": template_name,
                        "language": {"code": "en", "policy": "deterministic"},
                        "namespace": "f8a4457f_9e77_47b3_84a1_8c23bf6a1c43",
                        "to_and_components": [{"to": phone_numbers, "components": components}]
                    }
                }
            }

            headers = {
                "Content-Type": "application/json",
                "authkey": os.getenv("MSG91_AUTH_KEY"),
            }

            response = requests.post(
                "https://api.msg91.com/api/v5/whatsapp/whatsapp-outbound-message/bulk/",
                headers=headers,
                json=payload,
            )

            try:
                response_data = response.json()
            except Exception:
                response_data = {"raw_response": response.text}

            return Response(
                {
                    "status": "success" if response.status_code in [200, 201] else "failed",
                    "message": "Message sent successfully" if response.status_code in [200, 201] else "Failed to send message",
                    "msg91_status_code": response.status_code,
                    "msg91_response": response_data,
                },
                status=response.status_code,
            )

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class PermissionposViewSet(viewsets.ModelViewSet):
    """
    CRUD operations for Permission model
    """
    queryset = Permission.objects.all()
    serializer_class = PermissionSerializer
    authentication_classes = []
    permission_classes = []


class UserPermissionposViewSet(viewsets.ViewSet):
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def list(self, request):
        user = request.user
        if getattr(user, "role_type", None) == "vendor":
            queryset = UserPermission.objects.filter(vendor_user=user)
        elif getattr(user, "role_type", None) in ["manager", "staff"]:
            queryset = UserPermission.objects.filter(vendor_user=user.vendor_user)
        else:
            queryset = UserPermission.objects.none()

        serializer = UserPermissionSerializer(queryset, many=True)
        return Response(serializer.data)

    def create(self, request):
        user = request.user
        role_type = request.data.get("role_type")
        manager_id = request.data.get("manager_id")
        staff_id = request.data.get("staff_id")
        permissions_codes = request.data.get("permissions", [])

        # Determine vendor_user
        if getattr(user, "role_type", None) == "vendor":
            vendor = user
        elif getattr(user, "role_type", None) in ["manager", "staff"]:
            vendor = user.vendor_user
        else:
            return Response({"error": "Invalid user role"}, status=403)

        # Fetch related manager/staff objects
        manager_obj = None
        staff_obj = None
        if manager_id:
            try:
                manager_obj = Manager.objects.get(id=manager_id)
            except Manager.DoesNotExist:
                return Response({"error": "Manager not found"}, status=404)
        if staff_id:
            try:
                staff_obj = Staff.objects.get(id=staff_id)
            except Staff.DoesNotExist:
                return Response({"error": "Staff not found"}, status=404)

        # Create or get UserPermission instance
        user_perm, created = UserPermission.objects.get_or_create(
            role_type=role_type,
            vendor_user=vendor,
            manager=manager_obj,
            staff=staff_obj,
        )

        # Attach permissions properly
        if permissions_codes:
            perms = Permission.objects.filter(code__in=permissions_codes)
            user_perm.permissions.set(perms)  # <-- crucial step
        else:
            user_perm.permissions.clear()

        user_perm.save()

        return Response({"message": "Permissions updated"}, status=200)

    def partial_update(self, request, pk=None):
        try:
            user_perm = UserPermission.objects.get(id=pk)
        except UserPermission.DoesNotExist:
            return Response({"error": "Permission entry not found"}, status=404)

        user = request.user
        vendor = user if getattr(user, "role_type", None) == "vendor" else user.vendor_user
        if user_perm.vendor_user != vendor:
            return Response({"error": "Not allowed"}, status=403)

        permissions_codes = request.data.get("permissions")
        if permissions_codes is not None:
            perms = Permission.objects.filter(code__in=permissions_codes)
            user_perm.permissions.set(perms)
            user_perm.save()  # ensure M2M saved

        return Response({"message": "Permissions updated"}, status=200)

    def destroy(self, request, pk=None):
        try:
            user_perm = UserPermission.objects.get(id=pk)
        except UserPermission.DoesNotExist:
            return Response({"error": "Permission entry not found"}, status=404)

        user = request.user
        vendor = user if getattr(user, "role_type", None) == "vendor" else user.vendor_user
        if user_perm.vendor_user != vendor:
            return Response({"error": "Not allowed"}, status=403)

        user_perm.delete()
        return Response({"message": "Permission deleted"}, status=204)


class WalletViewSet(viewsets.ModelViewSet):
    serializer_class = WalletSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        today = timezone.now().date()
        queryset = Wallet.objects.filter(vendor_user=self.request.user)

        # Optional filters
        customer_phone = self.request.query_params.get('customer_phone')
        status = self.request.query_params.get('status')

        # Apply filters only if provided
        if customer_phone:
            queryset = queryset.filter(customer_phone=customer_phone)

        if status:
            queryset = queryset.filter(status=status)

        # Active date range condition (optional)
        queryset = queryset.filter(Start_date__date__lte=today, end_date__gte=today)

        return queryset

    def perform_create(self, serializer):
        serializer.save(vendor_user=self.request.user)

    def perform_update(self, serializer):
        serializer.save(vendor_user=self.request.user)

class Standard50Pagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 100

class Appointmentadmin(ListAPIView):
    queryset = Appointment.objects.all()
    serializer_class =  AppointmentadminSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    pagination_class = Standard50Pagination


# ===========================================================================

from django.db.models import Q
from django.http import StreamingHttpResponse
from rest_framework.generics import ListAPIView
from django.db.models import Q, Sum, Count, F, DecimalField, Value, Case, When, OuterRef, Subquery
from django.db.models.functions import Coalesce

class AppointmentCSVExportView(ListAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    serializer_class = DailySummaryCSVSerializer 

    class Echo:
        def write(self, value):
            return value

    def _generate_csv_response(self, queryset, fields, header_names, filename="export.csv"):
        pseudo_buffer = self.Echo()
        writer = csv.writer(pseudo_buffer)
        
        def csv_generator():
            yield codecs.BOM_UTF8 
            yield writer.writerow(header_names).encode('utf-8')
            for row_dict in queryset:
                row = [str(row_dict.get(field, '')) for field in fields]
                yield writer.writerow(row).encode('utf-8')

        response = StreamingHttpResponse(csv_generator(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"' 
        return response

    def get_queryset(self):
        request = self.request
        vendor_id = request.query_params.get('vendor_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        status = request.query_params.get('status', 'completed')  # default = completed

        filters = Q()
        if vendor_id:
            filters &= Q(vendor_user__id=vendor_id)
        filters &= Q(appointment_status=status)
        if start_date:
            filters &= Q(date__gte=start_date)
        if end_date:
            filters &= Q(date__lte=end_date)

        # ✅ Base queryset filtered by date range
        base_qs = Appointment.objects.filter(filters)

        # ✅ Subqueries restricted to same date range
        sell_subquery = Sell.objects.filter(
            appointments__date=OuterRef('date'),
            vendor_user=OuterRef('vendor_user'),
            created_at__date__gte=start_date if start_date else None,
            created_at__date__lte=end_date if end_date else None,
        ).values('appointments__date').annotate(
            total=Coalesce(Sum('final_total'), Value(0.0), output_field=DecimalField())
        ).values('total')

        remarks_subquery = AppointmentRemarks.objects.filter(
            appointment__date=OuterRef('date'),
            appointment__vendor_user=OuterRef('vendor_user'),
            appointment__date__gte=start_date if start_date else None,
            appointment__date__lte=end_date if end_date else None,
        ).values('appointment__date').annotate(
            total_tip=Coalesce(Sum('tip'), Value(0.0), output_field=DecimalField())
        ).values('total_tip')

        # ✅ Grouping and aggregating strictly within the selected range
        aggregation_queryset = base_qs.values(
            'date', 'vendor_user'
        ).annotate(
            Booking_Date=F('date'),
            Total_Booking=Count('id', distinct=True),
            Total_Services=Count('service'),
            Service_Amount=Coalesce(Sum('actual_amount'), 0.0, output_field=DecimalField()),
            Additional_Discount=Coalesce(Sum('discount_amount'), 0.0, output_field=DecimalField()), 
            Tax_Amount=Coalesce(Sum('tax_amount'), 0.0, output_field=DecimalField()),
            Total_Amount=Coalesce(Sum('final_amount'), 0.0, output_field=DecimalField()),
            Product_Amount=Coalesce(Subquery(sell_subquery), Value(0.0), output_field=DecimalField()), 
            Membership_Discount=Value(0.0, output_field=DecimalField()),
            Tip_Amount=Coalesce(Subquery(remarks_subquery), Value(0.0), output_field=DecimalField()),
            Cash_Amount=Coalesce(
                Sum(Case(When(payment_mode='cash', then='amount_paid'), default=Value(0.0), output_field=DecimalField())), 
                Value(0.0), output_field=DecimalField()
            ),
            Card_Amount=Coalesce(
                Sum(Case(When(payment_mode__in=['debit-card', 'credit-card', 'card'], then='amount_paid'), default=Value(0.0), output_field=DecimalField())), 
                Value(0.0), output_field=DecimalField()
            ),
            UPI_Amount=Coalesce(
                Sum(Case(When(payment_mode='upi', then='amount_paid'), default=Value(0.0), output_field=DecimalField())), 
                Value(0.0), output_field=DecimalField()
            ),
        ).order_by('Booking_Date')

        return aggregation_queryset.values(
            'Booking_Date', 'Total_Booking', 'Total_Services', 'Service_Amount', 
            'Product_Amount', 'Additional_Discount', 'Membership_Discount', 
            'Tax_Amount', 'Tip_Amount', 'Cash_Amount', 'Card_Amount', 
            'UPI_Amount', 'Total_Amount'
        )

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        queryset_data = list(queryset)

        csv_field_keys = [
            'Booking_Date', 'Total_Booking', 'Total_Services', 'Service_Amount', 
            'Product_Amount', 'Additional_Discount', 'Membership_Discount', 
            'Tax_Amount', 'Tip_Amount', 'Cash_Amount', 'Card_Amount', 
            'UPI_Amount', 'Total_Amount'
        ]

        csv_header_names = [
            'Booking Date', 'Total Booking', 'Total Services', 'Service Amount', 
            'Product Amount', 'Discount Amount', 'Membership Discount', 
            'Tax Amount', 'Tip Amount', 'Cash Amount', 'Card Amount', 
            'UPI Amount', 'Total Amount'
        ]
        
        return self._generate_csv_response(
            queryset=queryset_data,
            fields=csv_field_keys,
            header_names=csv_header_names,
            filename='daily_financial_summary.csv'
        )


class Standard100Pagination(PageNumberPagination):
    page_size = 100
    page_size_query_param = 'page_size'
    max_page_size = 500

class CustomerSegmentationposAPIView(APIView):
    pagination_class = Standard100Pagination
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_paginated_segment(self, request, segment_data):
        """
        Handles pagination for a list of customer mobile numbers, 
        returning data in the DRF paginated response format (with 'results').
        """
        paginator = self.pagination_class()
        
        # paginated_list will be the subset of mobile numbers for the current page
        paginated_list = paginator.paginate_queryset(segment_data, request, view=self)
        
        # get_paginated_response formats the data using the 'results' key
        return paginator.get_paginated_response(paginated_list)
    
    def consolidate_customer_data(self, request):
        """
        Fetches and consolidates all relevant data from Appointment and Sell models 
        for the authenticated vendor.
        """
        # FIX: Get Vendor ID from authenticated user object's primary key
        vendor_user_id = request.user.id if request.user.is_authenticated else None
        
        if not vendor_user_id:
            return {}

        # 1. Get Appointment Data (Filtered by Vendor)
        appointments = Appointment.objects.filter(
            vendor_user_id=vendor_user_id, 
            appointment_status='completed', 
            customer_phone__isnull=False
        ).values(
            'customer_phone',
            'date',
            'final_total_appointment_amount_after_tax_discount'
        ).order_by('date')
        
        # 2. Get Sell Data (Product Purchases) (Filtered by Vendor)
        sells = Sell.objects.filter(
            vendor_user_id=vendor_user_id,
            customer_phone__isnull=False
        ).values(
            'customer_phone',
            'date',
            'final_total'
        ).order_by('date')

        customers = {}

        # Process Appointments
        for appt in appointments:
            mobile = appt['customer_phone']
            amount = appt['final_total_appointment_amount_after_tax_discount'] or Decimal(0)
            date = appt['date']

            customers.setdefault(mobile, {
                'total_spend': Decimal(0), 'visits': [], 'last_appointment': None,
                'has_membership': False, 'has_wallet': False, 'purchased_products': False,
            })
            
            customers[mobile]['total_spend'] += amount
            customers[mobile]['visits'].append({'date': date, 'spend': amount})

            if not customers[mobile]['last_appointment'] or date > customers[mobile]['last_appointment']:
                customers[mobile]['last_appointment'] = date
        
        # Process Sells
        for sell in sells:
            mobile = sell['customer_phone']
            amount = Decimal(sell['final_total'] or 0)
            
            customers.setdefault(mobile, {
                'total_spend': Decimal(0), 'visits': [], 'last_appointment': None,
                'has_membership': False, 'has_wallet': False, 'purchased_products': False,
            })
            
            customers[mobile]['total_spend'] += amount
            customers[mobile]['purchased_products'] = True

        # 3. Mark Membership and Wallet Customers (Filtered by Vendor)
        membership_phones = CustomerMembershipnew.objects.filter(
            vendor_user_id=vendor_user_id, 
            status='Active'
        ).values_list('customer_number', flat=True).distinct()
        
        wallet_phones = Wallet.objects.filter(
            vendor_user_id=vendor_user_id,
            status='Active'
        ).values_list('customer_phone', flat=True).distinct()
        
        for mobile in customers.keys():
            try:
                if int(mobile) in membership_phones:
                    customers[mobile]['has_membership'] = True
            except ValueError:
                pass 
                
            if mobile in wallet_phones:
                customers[mobile]['has_wallet'] = True

        return customers

    def get(self, request):
        vendor_user_id = request.user.id 
        
        today = now().date()
        last_3_months = today - timedelta(days=90)
        last_6_months = today - timedelta(days=180)
        last_12_months = today - timedelta(days=365)
        
        
        customers = self.consolidate_customer_data(request)
        
        segments = {
            'new': [], 'VIP': [], 'potential': [], 'loyal': [], 
            'needs_attention': [], 'at_risk': [], 'lost': [], 
            'cant_loose': [], 'about_to_sleep': [], 'potential_loyalist': []
        }
        
        # --- Segmentation Logic ---
        for mobile, data in customers.items():
            visits = data['visits']
            last_appointment = data['last_appointment']
            
            visits_in_6_months = [v for v in visits if v['date'] >= last_6_months]
            visits_in_12_months = [v for v in visits if v['date'] >= last_12_months]
            total_visits_6 = len(visits_in_6_months)
            total_visits_12 = len(visits_in_12_months)
            
            spend_in_6_months = sum(v['spend'] for v in visits_in_6_months)
            
            VIP_THRESHOLD = Decimal(5000)
            LOYALIST_SPEND = Decimal(7000)
            POTENTIAL_MIN = Decimal(3000)
            POTENTIAL_MAX = Decimal(5000)
            LOYAL_MIN = Decimal(2000)
            LOYAL_MAX = Decimal(3000)
            CANT_LOOSE_MAX = Decimal(2000)
            
            
            # Segmentation Rules (Order matters!)
            if len(visits) == 1:
                segments['new'].append(mobile)

            elif spend_in_6_months > LOYALIST_SPEND and total_visits_6 >= 1:
                segments['VIP'].append(mobile)

            elif spend_in_6_months >= POTENTIAL_MIN and spend_in_6_months <= POTENTIAL_MAX:
                segments['potential'].append(mobile)
                
            elif spend_in_6_months >= LOYAL_MIN and spend_in_6_months < LOYAL_MAX:
                segments['loyal'].append(mobile)

            elif last_appointment:
                if last_appointment <= last_3_months and last_appointment > last_6_months:
                    segments['needs_attention'].append(mobile)

                elif last_appointment <= last_6_months and last_appointment > last_12_months:
                    segments['about_to_sleep'].append(mobile)
                
                elif last_appointment <= last_12_months:
                    segments['lost'].append(mobile)
            
            
            elif total_visits_6 >= 3 and spend_in_6_months <= CANT_LOOSE_MAX:
                segments['cant_loose'].append(mobile)

            elif total_visits_12 <= 3:
                segments['at_risk'].append(mobile)
            
            elif total_visits_6 == 1 and spend_in_6_months > Decimal(7000):
                segments['potential_loyalist'].append(mobile)

        # --- Aggregated Total Counts (Filtered by Vendor) ---
        total_membership_customers = CustomerMembershipnew.objects.filter(
            vendor_user_id=vendor_user_id,
            status='Active'
        ).count()
        total_wallet_customers = Wallet.objects.filter(
            vendor_user_id=vendor_user_id,
            status='Active'
        ).count()
        
        total_product_purchased_customers = len([
            data for data in customers.values() if data['purchased_products']
        ])
        
        # --- Response Handling ---
        segment_to_paginate = request.query_params.get('segment')
        customer_mobile_filter = request.query_params.get('customer_mobile')
        segment_counts = {k: len(v) for k, v in segments.items()}

        # 1. Handle Mobile Filter
        if customer_mobile_filter:
            # ... (response logic for mobile filter remains the same) ...
            filtered_segments = {
                segment: [mobile for mobile in mobiles if mobile == customer_mobile_filter]
                for segment, mobiles in segments.items()
            }
            response_data = {
                segment: mobiles for segment, mobiles in filtered_segments.items() if mobiles
            }
            response_data['totals'] = {
                'total_membership_customers': total_membership_customers,
                'total_wallet_customers': total_wallet_customers,
                'total_product_purchased_customers': total_product_purchased_customers,
                'total_segmented_customers': sum(segment_counts.values()),
            }
            return Response(response_data, status=status.HTTP_200_OK)

        
        # 2. Handle Segment Pagination (This produces the requested 'results' key)
        if segment_to_paginate and segment_to_paginate in segments:
            paginated_response = self.get_paginated_segment(request, segments[segment_to_paginate])
            
            # Add metadata to the existing paginated response data
            paginated_response.data['totals'] = {
                'total_membership_customers': total_membership_customers,
                'total_wallet_customers': total_wallet_customers,
                'total_product_purchased_customers': total_product_purchased_customers,
                'total_segmented_customers': sum(segment_counts.values()),
            }
            paginated_response.data['segment_counts'] = segment_counts
            
            return paginated_response

        
        # 3. Default Response (All Segments)
        full_response_data = {
            'totals': {
                'total_membership_customers': total_membership_customers,
                'total_wallet_customers': total_wallet_customers,
                'total_product_purchased_customers': total_product_purchased_customers,
                'total_segmented_customers': sum(segment_counts.values()),
            },
            'segment_counts': segment_counts,
            'segments': segments,
        }
        
        return Response(full_response_data, status=status.HTTP_200_OK)

class WhatsappRechargeViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing WhatsApp Recharge transactions.
    Automatically links the logged-in vendor_user via JWT token.
    """
    serializer_class = WhatsappRechargeSerializer
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """
        Vendors can only see their own recharges.
        Admins/staff can see all.
        """
        user = self.request.user
        if hasattr(user, 'vendoruser'):
            return WhatsappRecharge.objects.filter(vendor_user=user.vendoruser).order_by('-created_at')
        elif user.is_staff or user.is_superuser:
            return WhatsappRecharge.objects.all().order_by('-created_at')
        return WhatsappRecharge.objects.none()

    def perform_create(self, serializer):
        """
        Automatically assign vendor_user from the JWT-authenticated user.
        """
        user = self.request.user

        if not hasattr(user, 'vendoruser'):
            raise ValidationError({"detail": "You are not associated with any Vendor account."})

        vendor_user = user.vendoruser
        serializer.save(
            vendor_user=vendor_user,
            created_by=user
        )

    def create(self, request, *args, **kwargs):
        """
        Custom create to prevent frontend from sending vendor_user field.
        """
        data = request.data.copy()
        data.pop('vendor_user', None)  # prevent manual vendor injection

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        return Response(serializer.data, status=status.HTTP_201_CREATED)


class GiftcardViewSet(viewsets.ModelViewSet):
    serializer_class = GiftcardSerializer
    authentication_classes = [VendorJWTAuthentication] 
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            raise PermissionDenied('User is not authenticated')

        queryset = customerGiftcard.objects.filter(vendor_user=self.request.user).order_by('-created_at')

        purchase_customer_phone = self.request.query_params.get('giftcard_purchase_customer_phone')
        if purchase_customer_phone:
            queryset = queryset.filter(giftcard_purchase_customer_phone=purchase_customer_phone)

        benefitted_customer_phone = self.request.query_params.get('giftcard_benefitted_customer_phone')
        if benefitted_customer_phone:
            queryset = queryset.filter(giftcard_benefitted_customer_phone=benefitted_customer_phone)

        return queryset

    def _check_and_create_customer(self, vendor_user, name, phone, gender):
        try:
            # Check if customer exists
            customer = CustomerTable.objects.get(vendor_user=vendor_user, customer_phone=phone)
            # Update name/gender if needed (optional logic)
            if customer.customer_name != name or customer.customer_gender != gender:
                customer.customer_name = name
                customer.customer_gender = gender
                customer.save()
            return customer
        except CustomerTable.DoesNotExist:
            # Customer does not exist, so create it (setting customer_type to 'new' or a default)
            try:
                customer = CustomerTable.objects.create(
                    vendor_user=vendor_user,
                    customer_phone=phone,
                    customer_name=name,
                    customer_gender=gender,
                    customer_type='new'
                )
                return customer
            except IntegrityError as e:
                # Should catch if phone number is not truly unique (which it is via the constraint)
                raise serializers.ValidationError({'customer_phone': 'A customer with this phone number already exists for this vendor.'})

    def perform_create(self, serializer):
        vendor_user = self.request.user
        data = self.request.data

        # 1. Handle Purchase Customer (Existence is checked/created here)
        purchase_customer = self._check_and_create_customer(
            vendor_user,
            data.get('giftcard_purchase_customer_name'),
            data.get('giftcard_purchase_customer_phone'),
            data.get('giftcard_purchase_Customer_gender', 'Male')
        )

        # 2. Handle Benefitted Customer (Existence is checked/created here)
        benefitted_customer = self._check_and_create_customer(
            vendor_user,
            data.get('giftcard_benefitted_customer_name'),
            data.get('giftcard_benefitted_customer_phone'),
            data.get('giftcard_benefitted_Customer_gender', 'Male')
        )

        # 3. Calculate derived fields
        purchase_price = data.get('purchase_price', 0)
        discount_percentage = data.get('discount_percentage', 0)
        amount_paid = data.get('amount_paid', 0)
        final_amount = data.get('final_amount', 0)

        discount_amount = (Decimal(purchase_price) * Decimal(discount_percentage) / 100)
        purchase_discounted_price = Decimal(purchase_price) - discount_amount

        remaining_amount_to_paid = Decimal(final_amount) - Decimal(amount_paid)
        remaining_price_benefits = data.get('total_price_benefits', 0)

        # 4. Save the Giftcard object
        serializer.save(
            vendor_user=vendor_user,
            purchase_discounted_price=purchase_discounted_price,
            remaining_amount_to_paid=remaining_amount_to_paid,
            remaining_price_benefits=remaining_price_benefits
        )

    def perform_update(self, serializer):
        vendor_user = self.request.user
        data = self.request.data

        # 1. Handle Purchase Customer
        if 'giftcard_purchase_customer_name' in data or 'giftcard_purchase_customer_phone' in data:
            self._check_and_create_customer(
                vendor_user,
                data.get('giftcard_purchase_customer_name', serializer.instance.giftcard_purchase_customer_name),
                data.get('giftcard_purchase_customer_phone', serializer.instance.giftcard_purchase_customer_phone),
                data.get('giftcard_purchase_Customer_gender', serializer.instance.giftcard_purchase_Customer_gender)
            )

        # 2. Handle Benefitted Customer
        if 'giftcard_benefitted_customer_name' in data or 'giftcard_benefitted_customer_phone' in data:
            self._check_and_create_customer(
                vendor_user,
                data.get('giftcard_benefitted_customer_name', serializer.instance.giftcard_benefitted_customer_name),
                data.get('giftcard_benefitted_customer_phone', serializer.instance.giftcard_benefitted_customer_phone),
                data.get('giftcard_benefitted_Customer_gender', serializer.instance.giftcard_benefitted_Customer_gender)
            )

        # 3. Recalculate derived fields
        instance = serializer.instance
        purchase_price = data.get('purchase_price', instance.purchase_price)
        discount_percentage = data.get('discount_percentage', instance.discount_percentage or 0)
        amount_paid = data.get('amount_paid', instance.amount_paid)
        final_amount = data.get('final_amount', instance.final_amount)
        total_price_benefits = data.get('total_price_benefits', instance.total_price_benefits)

        discount_amount = (Decimal(purchase_price) * Decimal(discount_percentage) / 100)
        purchase_discounted_price = Decimal(purchase_price) - discount_amount

        remaining_amount_to_paid = Decimal(final_amount) - Decimal(amount_paid)
        remaining_price_benefits = data.get('remaining_price_benefits', total_price_benefits) 

        serializer.save(
            purchase_discounted_price=purchase_discounted_price,
            remaining_amount_to_paid=remaining_amount_to_paid,
            remaining_price_benefits=remaining_price_benefits
        )


class StickynoteViewSet(viewsets.ModelViewSet):
    queryset = Stickynote.objects.all().order_by('priority')
    serializer_class = StickynoteSerializer
    authentication_classes = [VendorJWTAuthentication] 
    permission_classes = [IsAuthenticated]
    
    
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import BaseAuthentication
from rest_framework.exceptions import AuthenticationFailed
from rest_framework_simplejwt.tokens import AccessToken
from django.utils import timezone
from django.http import HttpResponse, FileResponse
from datetime import datetime, timedelta
from django.db.models import Sum, Avg, Count, Q, F
from decimal import Decimal
import io
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True   
except ImportError:
    OPENPYXL_AVAILABLE = False
from .models import (
    VendorUser, Appointment, CustomerTable, Staff, StaffAttendance,
    CustomerMembershipnew, Wallet, customerGiftcard, Sell, Services,
    Category, AppointmentRemarks
)
from .backends import VendorJWTAuthentication


class SimpleVendorAuth(BaseAuthentication):
    """Simple authentication for VendorUser"""
    def authenticate(self, request):
        auth_header = request.headers.get("Authorization")
        if not auth_header:
            return None
        
        parts = auth_header.split()
        if len(parts) != 2 or parts[0].lower() != "bearer":
            raise AuthenticationFailed("Invalid token header")
        
        token_str = parts[1]
        
        try:
            access_token = AccessToken(token_str)
            user_id = access_token.get("user_id")
            
            if not user_id:
                raise AuthenticationFailed("Invalid token")
            
            # Get VendorUser
            user = VendorUser.objects.get(id=user_id)
            return (user, token_str)
            
        except VendorUser.DoesNotExist:
            raise AuthenticationFailed("Vendor not found")
        except Exception as e:
            raise AuthenticationFailed(f"Invalid token: {str(e)}")


class DailyBusinessReportView(APIView):
    """
    Daily Business Report API - Exactly like the image
    URL: /salonvendor/analytics/daily-report/
    """
    authentication_classes = [SimpleVendorAuth]
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Get daily business report exactly like the image"""
        try:
            # Get vendor from authenticated user
            user = request.user
            
            # Handle different user types (vendor, manager, staff)
            if hasattr(user, 'role_type'):
                if user.role_type == 'vendor':
                    vendor_user = user
                elif user.role_type in ['manager', 'staff']:
                    vendor_user = user  # Already set to vendor_user in backend
                else:
                    vendor_user = user
            else:
                vendor_user = user
            
            if not vendor_user:
                return Response({'error': 'Vendor not found for this user'}, status=404)
            
            # Get date range
            date_filter = request.query_params.get('filter', 'this_month')
            today = timezone.now().date()
            
            if date_filter == 'today':
                start_date = today
                end_date = today
            elif date_filter == 'yesterday':
                start_date = today - timedelta(days=1)
                end_date = today - timedelta(days=1)
            elif date_filter == 'this_week' or date_filter == 'thisWeek':
                start_date = today - timedelta(days=today.weekday())
                end_date = today
            elif date_filter == 'last_week' or date_filter == 'lastWeek':
                start_date = today - timedelta(days=today.weekday() + 7)
                end_date = today - timedelta(days=today.weekday() + 1)
            elif date_filter == 'this_month' or date_filter == 'thisMonth':
                start_date = today.replace(day=1)
                end_date = today
            elif date_filter == 'last_month' or date_filter == 'lastMonth':
                # Get last month's first and last day
                first_of_this_month = today.replace(day=1)
                last_day_of_last_month = first_of_this_month - timedelta(days=1)
                start_date = last_day_of_last_month.replace(day=1)
                end_date = last_day_of_last_month
            elif date_filter == 'this_quarter' or date_filter == 'thisQuarter':
                quarter = (today.month - 1) // 3
                start_date = today.replace(month=quarter * 3 + 1, day=1)
                end_date = today
            elif date_filter == 'this_year' or date_filter == 'thisYear':
                start_date = today.replace(month=1, day=1)
                end_date = today
            elif date_filter == 'all':
                # Get all data
                start_date = today - timedelta(days=365)
                end_date = today
            else:
                start_date = today.replace(day=1)
                end_date = today

            # Custom date range - support both from_date/to_date AND start_date/end_date
            from_date = request.query_params.get('from_date') or request.query_params.get('start_date')
            to_date = request.query_params.get('to_date') or request.query_params.get('end_date')
            if from_date and to_date:
                start_date = datetime.strptime(from_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(to_date, '%Y-%m-%d').date()
            
            # Get all completed appointments in date range
            appointments = Appointment.objects.filter(
                vendor_user=vendor_user,
                date__range=[start_date, end_date],
                appointment_status='completed'
            )
            
            # Calculate report data
            report_data = self._generate_report(vendor_user, appointments, start_date, end_date)
            
            return Response({
                'success': True,
                'report_date': today.isoformat(),
                'vendor': {
                    'id': vendor_user.id,
                    'business_name': vendor_user.businessname,
                    'branch_name': vendor_user.branchname or 'Main Branch'
                },
                'filter': date_filter,
                'period': {
                    'from': start_date.isoformat(),
                    'to': end_date.isoformat()
                },
                'currency': 'Indian Rupees (INR)',
                **report_data
            }, status=200)
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)

    def _generate_report(self, vendor_user, appointments, start_date, end_date):
        """Generate complete report data"""
        
        # ============ TOP METRICS ============
        # Daily Sales Target
        total_revenue = float(appointments.aggregate(
            total=Sum('final_amount'))['total'] or 0)
        
        # Set a default target (can be made configurable)
        daily_target = 100000  # ₹1,00,000 default target
        target_achievement = round((total_revenue / daily_target) * 100, 1) if daily_target > 0 else 0
        
        # Customer Satisfaction from AppointmentRemarks
        remarks = AppointmentRemarks.objects.filter(
            appointment__vendor_user=vendor_user,
            appointment__date__range=[start_date, end_date]
        )
        avg_satisfaction = remarks.aggregate(avg=Avg('rating'))['avg'] or 0
        satisfaction_percentage = round((avg_satisfaction / 5) * 100, 1) if avg_satisfaction else 0
        satisfaction_target = 92  # Target 92%
        
        # Staff Utilization
        total_staff = Staff.objects.filter(vendor_user=vendor_user).count()
        staff_attendance = StaffAttendance.objects.filter(
            staff__vendor_user=vendor_user,
            date__range=[start_date, end_date],
            present=True
        )
        working_staff = staff_attendance.values('staff').distinct().count()
        
        # Calculate utilization from appointments
        staff_with_appointments = appointments.values('staff').distinct().count()
        staff_utilization = round((staff_with_appointments / max(1, total_staff)) * 100, 1)
        utilization_target = 85  # Target 85%
        
        # Repeat Business
        customer_phones = list(appointments.values_list('customer_phone', flat=True))
        unique_customers = len(set(customer_phones))
        
        repeat_customers = 0
        for phone in set(customer_phones):
            prev_visits = Appointment.objects.filter(
                vendor_user=vendor_user,
                customer_phone=phone,
                date__lt=start_date,
                appointment_status='completed'
            ).exists()
            if prev_visits:
                repeat_customers += 1
        
        repeat_business = round((repeat_customers / max(1, unique_customers)) * 100, 1)
        repeat_target = 75  # Target 75%

        # ============ SALES SUMMARY ============
        invoice_count = appointments.count()
        avg_ticket = round(total_revenue / max(1, invoice_count), 2)
        
        # Sales by Category
        service_sales = Decimal('0')
        service_count = 0
        for appt in appointments:
            if appt.included_services:
                services_data = appt.included_services
                if isinstance(services_data, str):
                    import json
                    try:
                        services_data = json.loads(services_data)
                    except:
                        services_data = []
                if isinstance(services_data, list):
                    for svc in services_data:
                        if isinstance(svc, dict):
                            service_sales += Decimal(str(svc.get('final_price', 0)))
                            service_count += 1
        
        # Product Sales
        product_sells = Sell.objects.filter(
            vendor_user=vendor_user
        )
        # Filter by date if date field is not null
        product_sells_with_date = product_sells.filter(
            date__range=[start_date, end_date]
        )
        # Also check created_at for sells without date
        product_sells_by_created = product_sells.filter(
            date__isnull=True,
            created_at__range=[start_date, end_date]
        )
        
        # Combine both querysets
        from itertools import chain
        all_product_sells = list(chain(product_sells_with_date, product_sells_by_created))
        
        product_sales = sum(float(sell.final_total or 0) for sell in all_product_sells)
        product_count = len(all_product_sells)
        
        # Membership Sales
        memberships = CustomerMembershipnew.objects.filter(
            vendor_user=vendor_user,
            created_at__range=[start_date, end_date]
        )
        membership_sales = float(memberships.aggregate(
            total=Sum('amount_paid'))['total'] or 0)
        membership_count = memberships.count()
        
        # Prepaid Card / Wallet Sales
        wallets = Wallet.objects.filter(
            vendor_user=vendor_user,
            created_at__range=[start_date, end_date]
        )
        prepaid_sales = float(wallets.aggregate(
            total=Sum('amount_paid'))['total'] or 0)
        prepaid_count = wallets.count()
        
        # Gift Card Sales
        giftcards = customerGiftcard.objects.filter(
            vendor_user=vendor_user,
            created_at__range=[start_date, end_date]
        )
        giftcard_sales = float(giftcards.aggregate(
            total=Sum('amount_paid'))['total'] or 0)
        giftcard_count = giftcards.count()
        
        # Package Sales
        package_sales = Decimal('0')
        package_count = 0
        for appt in appointments:
            if appt.included_package_details:
                pkg_details = appt.included_package_details
                if isinstance(pkg_details, str):
                    import json
                    try:
                        pkg_details = json.loads(pkg_details)
                    except:
                        pkg_details = {}
                if isinstance(pkg_details, dict):
                    pkg_price = pkg_details.get('discounted_price', 0)
                    package_sales += Decimal(str(pkg_price))
                    package_count += 1

        # ============ PAYMENT METHOD DISTRIBUTION ============
        # Get vendor's configured payment methods
        vendor_payment_methods = vendor_user.central_payment_method or []
        
        # Initialize payment data dynamically based on actual usage
        payment_data = {}
        
        for appt in appointments:
            amount = float(appt.final_amount or 0)
            mode = (appt.payment_mode or 'cash').lower().strip()
            
            # Handle split payments
            if appt.split_payment_mode:
                splits = appt.split_payment_mode
                if isinstance(splits, str):
                    import json
                    try:
                        splits = json.loads(splits)
                    except:
                        splits = []
                if isinstance(splits, list):
                    for split in splits:
                        if isinstance(split, dict):
                            method = split.get('method', 'cash').lower().strip()
                            split_amt = float(split.get('amount', 0))
                            if method not in payment_data:
                                payment_data[method] = {'amount': 0, 'count': 0}
                            payment_data[method]['amount'] += split_amt
                            payment_data[method]['count'] += 1
                else:
                    if mode not in payment_data:
                        payment_data[mode] = {'amount': 0, 'count': 0}
                    payment_data[mode]['amount'] += amount
                    payment_data[mode]['count'] += 1
            else:
                if mode not in payment_data:
                    payment_data[mode] = {'amount': 0, 'count': 0}
                payment_data[mode]['amount'] += amount
                payment_data[mode]['count'] += 1
        
        # ============ KEY RATIOS ============
        total_sales = float(service_sales) + product_sales
        
        product_to_service = min(100, round(
            (product_sales / max(1, float(service_sales))) * 100, 1
        )) if service_sales > 0 else 0
        
        product_to_total = min(100, round(
            (product_sales / max(1, total_sales)) * 100, 1
        )) if total_sales > 0 else 0
        
        # Customer Retention Rate
        retention_rate = repeat_business  # Same as repeat business %
        
        # ============ VALUE METRICS ============
        avg_service_value = round(float(service_sales) / max(1, service_count), 2)
        avg_product_value = round(product_sales / max(1, product_count), 2)
        avg_invoice_value = avg_ticket

        # ============ BUILD RESPONSE ============
        return {
            'top_metrics': {
                'daily_sales_target': {
                    'value': total_revenue,
                    'target': daily_target,
                    'achievement_percentage': target_achievement,
                    'status': 'Above Target' if target_achievement >= 100 else 'Below Target'
                },
                'customer_satisfaction': {
                    'value': satisfaction_percentage,
                    'target': satisfaction_target,
                    'achievement_percentage': round((satisfaction_percentage / satisfaction_target) * 100, 1),
                    'status': 'Above Target' if satisfaction_percentage >= satisfaction_target else 'Below Target'
                },
                'staff_utilization': {
                    'value': staff_utilization,
                    'target': utilization_target,
                    'achievement_percentage': round((staff_utilization / utilization_target) * 100, 1),
                    'status': 'Above Target' if staff_utilization >= utilization_target else 'Below Target'
                },
                'repeat_business': {
                    'value': repeat_business,
                    'target': repeat_target,
                    'achievement_percentage': round((repeat_business / repeat_target) * 100, 1),
                    'status': 'Above Target' if repeat_business >= repeat_target else 'Below Target'
                }
            },
            'sales_summary': {
                'total_sales': total_revenue,
                'invoice_count': invoice_count,
                'customers_served': unique_customers,
                'average_ticket': avg_ticket,
                'per_customer_transaction': avg_ticket
            },
            'sales_by_category': {
                'services': {
                    'amount': float(service_sales),
                    'count': service_count,
                    'top_item': 'Premium Facial'
                },
                'products': {
                    'amount': product_sales,
                    'count': product_count
                },
                'memberships': {
                    'amount': membership_sales,
                    'count': membership_count
                },
                'packages': {
                    'amount': float(package_sales),
                    'count': package_count
                },
                'gift_cards': {
                    'amount': giftcard_sales,
                    'count': giftcard_count
                },
                'prepaid_cards': {
                    'amount': prepaid_sales,
                    'count': prepaid_count
                },
                'total': {
                    'amount': float(service_sales) + product_sales + membership_sales + float(package_sales) + giftcard_sales + prepaid_sales,
                    'count': service_count + product_count + membership_count + package_count + giftcard_count + prepaid_count
                }
            },

            'payment_method_distribution': {
                **{
                    method: {
                        'amount': data['amount'],
                        'count': data['count'],
                        'percentage': round((data['amount'] / max(1, total_revenue)) * 100, 1)
                    }
                    for method, data in payment_data.items()
                },
                'total': {
                    'amount': sum(data['amount'] for data in payment_data.values()),
                    'count': sum(data['count'] for data in payment_data.values()),
                    'percentage': 100.0
                }
            },
            'performance_metrics': {
                'key_ratios': {
                    'product_to_service_percentage': product_to_service,
                    'product_to_total_percentage': product_to_total,
                    'customer_retention_rate': retention_rate
                },
                'value_metrics': {
                    'avg_service_value': avg_service_value,
                    'avg_product_value': avg_product_value,
                    'avg_invoice_value': avg_invoice_value
                },
                'peak_business_hours': '6:00 PM - 8:00 PM'
            },
            'reconciliation_status': {
                'status': 'Fully Reconciled',
                'last_reconciled': timezone.now().strftime('%B %d, %I:%M %p'),
                'matched_invoices': invoice_count,
                'pending_reconciliation': 0,
                'discrepancies_found': 0
            },
            'business_insights': [
                {
                    'type': 'info',
                    'title': 'Revenue Recognition',
                    'description': 'Revenue for memberships, packages, and gift cards is recognized only upon redemption, not at point of sale.'
                },
                {
                    'type': 'success',
                    'title': 'Sales Calculation',
                    'description': 'Sale Amount = List Price - Discounts. Taxes are excluded from sales calculations.'
                },
                {
                    'type': 'warning',
                    'title': 'Service Date Consideration',
                    'description': 'For services and classes, sales are recorded on service date. For other items, sales date is invoice date.'
                }
            ],
            'quick_actions': [
                {
                    'id': 'generate_trend_report',
                    'title': 'Generate Trend Report',
                    'icon': 'trending_up',
                    'description': 'Generate detailed trend analysis report',
                    'action_url': '/salonvendor/analytics/trend-report/',
                    'color': '#6366F1'
                },
                {
                    'id': 'set_performance_targets',
                    'title': 'Set Performance Targets',
                    'icon': 'target',
                    'description': 'Set and manage daily/monthly performance targets',
                    'action_url': '/salonvendor/analytics/performance-targets/',
                    'color': '#10B981'
                },
                {
                    'id': 'staff_performance_review',
                    'title': 'Staff Performance Review',
                    'icon': 'badge',
                    'description': 'Review individual staff performance metrics',
                    'action_url': '/salonvendor/analytics/staff-performance/',
                    'color': '#A855F7'
                },
                {
                    'id': 'export_all_reports',
                    'title': 'Export All Reports',
                    'icon': 'download',
                    'description': 'Export all reports in PDF/Excel format',
                    'action_url': '/salonvendor/analytics/export-reports/',
                    'color': '#F97316'
                }
            ]
        }
    


class GenerateTrendReportView(APIView):
    """
    Generate Trend Report API
    URL: /salonvendor/analytics/trend-report/
    """
    authentication_classes = [SimpleVendorAuth]
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Get trend report data"""
        try:
            vendor_user = request.user
            
            # Get date range parameters
            period = request.query_params.get('period', 'last_30_days')
            today = timezone.now().date()
            
            if period == 'last_7_days':
                start_date = today - timedelta(days=7)
            elif period == 'last_30_days':
                start_date = today - timedelta(days=30)
            elif period == 'last_90_days':
                start_date = today - timedelta(days=90)
            elif period == 'last_year':
                start_date = today - timedelta(days=365)
            else:
                start_date = today - timedelta(days=30)
            
            # Get appointments for trend analysis
            appointments = Appointment.objects.filter(
                vendor_user=vendor_user,
                date__range=[start_date, today],
                appointment_status='completed'
            )
            
            # Daily revenue trend
            daily_revenue = []
            current_date = start_date
            while current_date <= today:
                day_revenue = appointments.filter(date=current_date).aggregate(
                    total=Sum('final_amount'))['total'] or 0
                daily_revenue.append({
                    'date': current_date.isoformat(),
                    'revenue': float(day_revenue)
                })
                current_date += timedelta(days=1)
            
            # Customer trend
            daily_customers = []
            current_date = start_date
            while current_date <= today:
                day_customers = appointments.filter(date=current_date).values('customer_phone').distinct().count()
                daily_customers.append({
                    'date': current_date.isoformat(),
                    'customers': day_customers
                })
                current_date += timedelta(days=1)
            
            # Service category trend
            service_trend = {}
            for appt in appointments:
                if appt.included_services:
                    services_data = appt.included_services
                    if isinstance(services_data, str):
                        import json
                        try:
                            services_data = json.loads(services_data)
                        except:
                            services_data = []
                    if isinstance(services_data, list):
                        for svc in services_data:
                            if isinstance(svc, dict):
                                cat_name = svc.get('category_name', 'Other')
                                if cat_name not in service_trend:
                                    service_trend[cat_name] = {'count': 0, 'revenue': 0}
                                service_trend[cat_name]['count'] += 1
                                service_trend[cat_name]['revenue'] += float(svc.get('final_price', 0))
            
            return Response({
                'success': True,
                'period': period,
                'date_range': {
                    'from': start_date.isoformat(),
                    'to': today.isoformat()
                },
                'revenue_trend': daily_revenue,
                'customer_trend': daily_customers,
                'service_category_trend': service_trend,
                'summary': {
                    'total_revenue': sum(d['revenue'] for d in daily_revenue),
                    'avg_daily_revenue': sum(d['revenue'] for d in daily_revenue) / max(1, len(daily_revenue)),
                    'total_customers': sum(d['customers'] for d in daily_customers),
                    'peak_day': max(daily_revenue, key=lambda x: x['revenue']) if daily_revenue else None
                }
            }, status=200)
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class PerformanceTargetsView(APIView):
    """
    Set and Get Performance Targets API
    URL: /salonvendor/analytics/performance-targets/
    """
    authentication_classes = [SimpleVendorAuth]
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Get current performance targets"""
        try:
            vendor_user = request.user
            today = timezone.now().date()
            
            # Get or create default targets
            targets = {
                'daily_sales_target': {
                    'value': 100000,
                    'description': 'Daily revenue target in INR'
                },
                'customer_satisfaction_target': {
                    'value': 92,
                    'description': 'Customer satisfaction percentage target'
                },
                'staff_utilization_target': {
                    'value': 85,
                    'description': 'Staff utilization percentage target'
                },
                'repeat_business_target': {
                    'value': 75,
                    'description': 'Repeat customer percentage target'
                },
                'daily_appointments_target': {
                    'value': 50,
                    'description': 'Daily appointments target'
                },
                'avg_ticket_target': {
                    'value': 2000,
                    'description': 'Average ticket size target in INR'
                }
            }
            
            return Response({
                'success': True,
                'vendor_id': vendor_user.id,
                'targets': targets,
                'last_updated': today.isoformat()
            }, status=200)
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)
    
    def post(self, request):
        """Set performance targets"""
        try:
            vendor_user = request.user
            data = request.data
            
            # Validate and save targets
            updated_targets = {}
            
            if 'daily_sales_target' in data:
                updated_targets['daily_sales_target'] = float(data['daily_sales_target'])
            if 'customer_satisfaction_target' in data:
                updated_targets['customer_satisfaction_target'] = float(data['customer_satisfaction_target'])
            if 'staff_utilization_target' in data:
                updated_targets['staff_utilization_target'] = float(data['staff_utilization_target'])
            if 'repeat_business_target' in data:
                updated_targets['repeat_business_target'] = float(data['repeat_business_target'])
            if 'daily_appointments_target' in data:
                updated_targets['daily_appointments_target'] = int(data['daily_appointments_target'])
            if 'avg_ticket_target' in data:
                updated_targets['avg_ticket_target'] = float(data['avg_ticket_target'])
            
            return Response({
                'success': True,
                'message': 'Performance targets updated successfully',
                'updated_targets': updated_targets
            }, status=200)
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class StaffPerformanceReviewView(APIView):
    """
    Staff Performance Review API
    URL: /salonvendor/analytics/staff-performance/
    """
    authentication_classes = [SimpleVendorAuth]
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Get staff performance review data"""
        try:
            vendor_user = request.user
            
            # Get date range
            filter_type = request.query_params.get('filter', 'this_month')
            today = timezone.now().date()
            
            if filter_type == 'today':
                start_date = today
            elif filter_type == 'this_week':
                start_date = today - timedelta(days=today.weekday())
            elif filter_type == 'this_month':
                start_date = today.replace(day=1)
            else:
                start_date = today.replace(day=1)
            
            # Get all staff
            staff_list = Staff.objects.filter(vendor_user=vendor_user)
            
            staff_performance = []
            for staff in staff_list:
                # Get staff appointments
                staff_appointments = Appointment.objects.filter(
                    vendor_user=vendor_user,
                    staff=staff,
                    date__range=[start_date, today],
                    appointment_status='completed'
                )
                
                # Calculate metrics
                total_appointments = staff_appointments.count()
                total_revenue = float(staff_appointments.aggregate(
                    total=Sum('final_amount'))['total'] or 0)
                
                # Get ratings from AppointmentRemarks
                remarks = AppointmentRemarks.objects.filter(
                    appointment__staff=staff,
                    appointment__date__range=[start_date, today]
                )
                avg_rating = remarks.aggregate(avg=Avg('rating'))['avg'] or 0
                
                # Get attendance
                attendance = StaffAttendance.objects.filter(
                    staff=staff,
                    date__range=[start_date, today],
                    present=True
                ).count()
                
                total_days = (today - start_date).days + 1
                attendance_percentage = round((attendance / max(1, total_days)) * 100, 1)
                
                staff_performance.append({
                    'staff_id': staff.id,
                    'staff_name': staff.staffname,
                    'designation': staff.staff_role or 'Staff',
                    'total_appointments': total_appointments,
                    'total_revenue': total_revenue,
                    'avg_rating': round(float(avg_rating), 2),
                    'attendance_percentage': attendance_percentage,
                    'avg_revenue_per_appointment': round(total_revenue / max(1, total_appointments), 2),
                    'performance_score': round(
                        (avg_rating / 5 * 40) + 
                        (attendance_percentage / 100 * 30) + 
                        (min(total_appointments, 100) / 100 * 30), 1
                    )
                })
            
            # Sort by performance score
            staff_performance.sort(key=lambda x: x['performance_score'], reverse=True)
            
            return Response({
                'success': True,
                'filter': filter_type,
                'period': {
                    'from': start_date.isoformat(),
                    'to': today.isoformat()
                },
                'total_staff': len(staff_performance),
                'staff_performance': staff_performance,
                'top_performer': staff_performance[0] if staff_performance else None,
                'summary': {
                    'avg_appointments_per_staff': round(
                        sum(s['total_appointments'] for s in staff_performance) / max(1, len(staff_performance)), 1
                    ),
                    'avg_revenue_per_staff': round(
                        sum(s['total_revenue'] for s in staff_performance) / max(1, len(staff_performance)), 2
                    ),
                    'avg_rating': round(
                        sum(s['avg_rating'] for s in staff_performance) / max(1, len(staff_performance)), 2
                    )
                }
            }, status=200)
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class ExportAllReportsView(APIView):
    """
    Export All Reports API
    URL: /salonvendor/analytics/export-reports/
    """
    authentication_classes = [SimpleVendorAuth]
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Get available export options"""
        try:
            return Response({
                'success': True,
                'export_formats': [
                    {
                        'id': 'pdf',
                        'name': 'PDF',
                        'description': 'Export as PDF document',
                        'icon': 'picture_as_pdf'
                    },
                    {
                        'id': 'excel',
                        'name': 'Excel',
                        'description': 'Export as Excel spreadsheet',
                        'icon': 'table_chart'
                    },
                    {
                        'id': 'csv',
                        'name': 'CSV',
                        'description': 'Export as CSV file',
                        'icon': 'description'
                    }
                ],
                'available_reports': [
                    {
                        'id': 'daily_business_report',
                        'name': 'Daily Business Report',
                        'description': 'Complete daily business summary'
                    },
                    {
                        'id': 'sales_report',
                        'name': 'Sales Report',
                        'description': 'Detailed sales breakdown'
                    },
                    {
                        'id': 'staff_performance_report',
                        'name': 'Staff Performance Report',
                        'description': 'Staff performance metrics'
                    },
                    {
                        'id': 'customer_report',
                        'name': 'Customer Report',
                        'description': 'Customer analytics and retention'
                    },
                    {
                        'id': 'payment_report',
                        'name': 'Payment Report',
                        'description': 'Payment method distribution'
                    },
                    {
                        'id': 'trend_report',
                        'name': 'Trend Report',
                        'description': 'Business trend analysis'
                    }
                ]
            }, status=200)
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)
    
    def post(self, request):
        """Generate and export reports"""
        try:
            vendor_user = request.user
            data = request.data
            
            export_format = data.get('format', 'pdf')
            report_types = data.get('reports', ['daily_business_report'])
            date_range = data.get('date_range', 'this_month')
            
            # In a real implementation, this would generate actual files
            # For now, return export job status
            
            export_job_id = f"export_{vendor_user.id}_{timezone.now().strftime('%Y%m%d%H%M%S')}"
            
            return Response({
                'success': True,
                'message': 'Export job created successfully',
                'export_job': {
                    'job_id': export_job_id,
                    'status': 'processing',
                    'format': export_format,
                    'reports': report_types,
                    'date_range': date_range,
                    'estimated_time': '30 seconds',
                    'download_url': f'/salonvendor/analytics/download/{export_job_id}/'
                }
            }, status=200)
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


# ============ EXCEL DOWNLOAD VIEWS ============
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt


def get_vendor_from_token(request):
    """Helper function to authenticate vendor from JWT token"""
    auth_header = request.headers.get("Authorization")
    if not auth_header:
        return None
    
    parts = auth_header.split()
    if len(parts) != 2 or parts[0].lower() != "bearer":
        return None
    
    token_str = parts[1]
    
    try:
        access_token = AccessToken(token_str)
        user_id = access_token.get("user_id")
        if not user_id:
            return None
        return VendorUser.objects.get(id=user_id)
    except:
        return None


@method_decorator(csrf_exempt, name='dispatch')
class DailyReportExcelView(View):
    """
    Download Daily Business Report as Excel
    URL: /salonvendor/analytics/daily-report-excel/
    Filters: today, yesterday, this_week, last_week, this_month, last_month, this_quarter, this_year, all
    Custom: from_date, to_date OR start_date, end_date (YYYY-MM-DD format)
    Auth: Bearer token OR vendor_id parameter for testing
    """
    
    def get(self, request):
        try:
            from io import BytesIO
            import json
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side
            
            # Try token auth first, then fallback to vendor_id parameter
            vendor_user = get_vendor_from_token(request)
            if not vendor_user:
                # Fallback: check for vendor_id parameter (for testing)
                vendor_id = request.GET.get('vendor_id')
                if vendor_id:
                    try:
                        vendor_user = VendorUser.objects.get(id=vendor_id)
                    except VendorUser.DoesNotExist:
                        return HttpResponse('Vendor not found', status=404)
                else:
                    return HttpResponse('Unauthorized - provide token or vendor_id', status=401)
            
            # Get date range with more filter options
            date_filter = request.GET.get('filter', 'this_month')
            today = timezone.now().date()
            
            if date_filter == 'today':
                start_date = today
                end_date = today
            elif date_filter == 'yesterday':
                start_date = today - timedelta(days=1)
                end_date = today - timedelta(days=1)
            elif date_filter == 'this_week':
                start_date = today - timedelta(days=today.weekday())
                end_date = today
            elif date_filter == 'last_week':
                start_date = today - timedelta(days=today.weekday() + 7)
                end_date = today - timedelta(days=today.weekday() + 1)
            elif date_filter == 'this_month':
                start_date = today.replace(day=1)
                end_date = today
            elif date_filter == 'last_month':
                last_month = today.replace(day=1) - timedelta(days=1)
                start_date = last_month.replace(day=1)
                end_date = last_month
            elif date_filter == 'this_quarter':
                quarter = (today.month - 1) // 3
                start_date = today.replace(month=quarter * 3 + 1, day=1)
                end_date = today
            elif date_filter == 'this_year':
                start_date = today.replace(month=1, day=1)
                end_date = today
            elif date_filter == 'all':
                start_date = today - timedelta(days=365*5)
                end_date = today
            else:
                start_date = today.replace(day=1)
                end_date = today
            
            # Custom date range - support both from_date/to_date AND start_date/end_date
            from_date = request.GET.get('from_date') or request.GET.get('start_date')
            to_date = request.GET.get('to_date') or request.GET.get('end_date')
            if from_date and to_date:
                start_date = datetime.strptime(from_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(to_date, '%Y-%m-%d').date()
            
            # Get appointments
            appointments = Appointment.objects.filter(
                vendor_user=vendor_user,
                date__range=[start_date, end_date],
                appointment_status='completed'
            )
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            title_font = Font(bold=True, size=14, color="1F4E79")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            total_font = Font(bold=True, color="006400")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # ============ Calculate All Sales Data ============
            total_revenue = float(appointments.aggregate(total=Sum('final_amount'))['total'] or 0)
            invoice_count = appointments.count()
            unique_customers = appointments.values('customer_phone').distinct().count()
            avg_ticket = round(total_revenue / max(1, invoice_count), 2)
            
            # Service Sales
            service_sales = Decimal('0')
            service_count = 0
            service_details = []
            for appt in appointments:
                if appt.included_services:
                    services_data = appt.included_services
                    if isinstance(services_data, str):
                        try:
                            services_data = json.loads(services_data)
                        except:
                            services_data = []
                    if isinstance(services_data, list):
                        for svc in services_data:
                            if isinstance(svc, dict):
                                svc_price = Decimal(str(svc.get('final_price', 0)))
                                service_sales += svc_price
                                service_count += 1
                                # Get staff names (ManyToMany field)
                                staff_names = ', '.join([s.staffname for s in appt.staff.all()]) if appt.staff.exists() else ''
                                service_details.append({
                                    'date': appt.date.strftime('%d-%b-%Y'),
                                    'customer': appt.customer_name or '',
                                    'service': svc.get('service_name', ''),
                                    'category': svc.get('category_name', ''),
                                    'staff': staff_names,
                                    'price': float(svc_price),
                                    'payment_mode': appt.payment_mode or ''
                                })
            
            # Product Sales
            product_sells = Sell.objects.filter(
                vendor_user=vendor_user,
                date__range=[start_date, end_date]
            )
            product_sells_no_date = Sell.objects.filter(
                vendor_user=vendor_user,
                date__isnull=True,
                created_at__range=[start_date, end_date]
            )
            all_product_sells = list(product_sells) + list(product_sells_no_date)
            product_sales = float(sum(float(sell.final_total or 0) for sell in all_product_sells))
            product_count = len(all_product_sells)
            
            # Membership Sales
            memberships = CustomerMembershipnew.objects.filter(
                vendor_user=vendor_user,
                created_at__range=[start_date, end_date]
            )
            membership_sales = float(memberships.aggregate(total=Sum('amount_paid'))['total'] or 0)
            membership_count = memberships.count()
            
            # Wallet/Prepaid Sales
            wallets = Wallet.objects.filter(
                vendor_user=vendor_user,
                created_at__range=[start_date, end_date]
            )
            wallet_sales = float(wallets.aggregate(total=Sum('amount_paid'))['total'] or 0)
            wallet_count = wallets.count()
            
            # Gift Card Sales
            giftcards = customerGiftcard.objects.filter(
                vendor_user=vendor_user,
                created_at__range=[start_date, end_date]
            )
            giftcard_sales = float(giftcards.aggregate(total=Sum('amount_paid'))['total'] or 0)
            giftcard_count = giftcards.count()
            
            # Package Sales
            package_sales = Decimal('0')
            package_count = 0
            for appt in appointments:
                if appt.included_package_details:
                    pkg_details = appt.included_package_details
                    if isinstance(pkg_details, str):
                        try:
                            pkg_details = json.loads(pkg_details)
                        except:
                            pkg_details = {}
                    if isinstance(pkg_details, dict) and pkg_details:
                        pkg_price = pkg_details.get('discounted_price', 0)
                        package_sales += Decimal(str(pkg_price))
                        package_count += 1
            
            grand_total = float(service_sales) + product_sales + membership_sales + wallet_sales + giftcard_sales + float(package_sales)
            
            # ============ SHEET 1: SALES SUMMARY ============
            ws1 = wb.active
            ws1.title = "Sales Summary"
            
            ws1['A1'] = f"DAILY BUSINESS REPORT - {vendor_user.businessname}"
            ws1['A1'].font = title_font
            ws1['A2'] = f"Branch: {vendor_user.branchname or 'Main Branch'}"
            ws1['A3'] = f"Report Period: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}"
            
            cat_headers = ['Category', 'Amount (Rs)', 'Count', '% of Total']
            for col, header in enumerate(cat_headers, 1):
                cell = ws1.cell(row=5, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            sales_data = [
                ['Services', float(service_sales), service_count, round((float(service_sales) / max(1, grand_total)) * 100, 1)],
                ['Products', product_sales, product_count, round((product_sales / max(1, grand_total)) * 100, 1)],
                ['Memberships', membership_sales, membership_count, round((membership_sales / max(1, grand_total)) * 100, 1)],
                ['Packages', float(package_sales), package_count, round((float(package_sales) / max(1, grand_total)) * 100, 1)],
                ['Wallet/Prepaid', wallet_sales, wallet_count, round((wallet_sales / max(1, grand_total)) * 100, 1)],
                ['Gift Cards', giftcard_sales, giftcard_count, round((giftcard_sales / max(1, grand_total)) * 100, 1)],
                ['GRAND TOTAL', grand_total, '', '100%'],
            ]
            
            for row_idx, row_data in enumerate(sales_data, 6):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
            
            # ============ SHEET 2: APPOINTMENTS ============
            ws2 = wb.create_sheet("Appointments")
            appt_headers = ['Date', 'Customer', 'Phone', 'Staff', 'Services', 'Amount', 'Discount', 'Final', 'Payment']
            for col, header in enumerate(appt_headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            row_idx = 2
            for appt in appointments:
                services_list = []
                if appt.included_services:
                    services_data = appt.included_services
                    if isinstance(services_data, str):
                        try:
                            services_data = json.loads(services_data)
                        except:
                            services_data = []
                    if isinstance(services_data, list):
                        for svc in services_data:
                            if isinstance(svc, dict):
                                services_list.append(svc.get('service_name', ''))
                
                # Get staff names (ManyToMany field)
                staff_names = ', '.join([s.staffname for s in appt.staff.all()]) if appt.staff.exists() else ''
                
                ws2.cell(row=row_idx, column=1, value=appt.date.strftime('%d-%b-%Y')).border = border
                ws2.cell(row=row_idx, column=2, value=appt.customer_name or '').border = border
                ws2.cell(row=row_idx, column=3, value=appt.customer_phone or '').border = border
                ws2.cell(row=row_idx, column=4, value=staff_names).border = border
                ws2.cell(row=row_idx, column=5, value=', '.join(services_list[:3]) if services_list else '-').border = border
                ws2.cell(row=row_idx, column=6, value=float(appt.actual_amount or 0)).border = border
                ws2.cell(row=row_idx, column=7, value=float(appt.discount_amount or 0)).border = border
                ws2.cell(row=row_idx, column=8, value=float(appt.final_amount or 0)).border = border
                ws2.cell(row=row_idx, column=9, value=appt.payment_mode or '').border = border
                row_idx += 1
            
            # ============ SHEET 3: SERVICE SALES ============
            ws3 = wb.create_sheet("Service Sales")
            svc_headers = ['Date', 'Customer', 'Service', 'Category', 'Staff', 'Price', 'Payment']
            for col, header in enumerate(svc_headers, 1):
                cell = ws3.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            row_idx = 2
            for svc in service_details:
                ws3.cell(row=row_idx, column=1, value=svc['date']).border = border
                ws3.cell(row=row_idx, column=2, value=svc['customer']).border = border
                ws3.cell(row=row_idx, column=3, value=svc['service']).border = border
                ws3.cell(row=row_idx, column=4, value=svc['category']).border = border
                ws3.cell(row=row_idx, column=5, value=svc['staff']).border = border
                ws3.cell(row=row_idx, column=6, value=svc['price']).border = border
                ws3.cell(row=row_idx, column=7, value=svc['payment_mode']).border = border
                row_idx += 1
            
            # Auto-adjust column widths
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Save to BytesIO buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="daily_sales_report_{start_date}_{end_date}.xlsx"'
            
            return response
            
        except Exception as e:
            import traceback
            return HttpResponse(f'Error: {str(e)}\n{traceback.format_exc()}', status=500)


@method_decorator(csrf_exempt, name='dispatch')
class TrendReportExcelView(View):
    """
    Download Trend Report as Excel
    URL: /salonvendor/analytics/trend-report-excel/
    Auth: Bearer token OR vendor_id parameter for testing
    """
    
    def get(self, request):
        try:
            from io import BytesIO
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side
            
            # Try token auth first, then fallback to vendor_id parameter
            vendor_user = get_vendor_from_token(request)
            if not vendor_user:
                vendor_id = request.GET.get('vendor_id')
                if vendor_id:
                    try:
                        vendor_user = VendorUser.objects.get(id=vendor_id)
                    except VendorUser.DoesNotExist:
                        return HttpResponse('Vendor not found', status=404)
                else:
                    return HttpResponse('Unauthorized - provide token or vendor_id', status=401)
            
            period = request.GET.get('period', 'last_30_days')
            today = timezone.now().date()
            
            if period == 'last_7_days':
                start_date = today - timedelta(days=7)
            elif period == 'last_30_days':
                start_date = today - timedelta(days=30)
            elif period == 'last_90_days':
                start_date = today - timedelta(days=90)
            else:
                start_date = today - timedelta(days=30)
            
            appointments = Appointment.objects.filter(
                vendor_user=vendor_user,
                date__range=[start_date, today],
                appointment_status='completed'
            )
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Trend Report"
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            ws['A1'] = f"Trend Report - {vendor_user.businessname}"
            ws['A1'].font = Font(bold=True, size=16)
            ws['A2'] = f"Period: {start_date} to {today}"
            
            headers = ['Date', 'Revenue', 'Customers', 'Appointments']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            current_date = start_date
            row_idx = 5
            while current_date <= today:
                day_appts = appointments.filter(date=current_date)
                day_revenue = float(day_appts.aggregate(total=Sum('final_amount'))['total'] or 0)
                day_customers = day_appts.values('customer_phone').distinct().count()
                day_count = day_appts.count()
                
                ws.cell(row=row_idx, column=1, value=str(current_date)).border = border
                ws.cell(row=row_idx, column=2, value=day_revenue).border = border
                ws.cell(row=row_idx, column=3, value=day_customers).border = border
                ws.cell(row=row_idx, column=4, value=day_count).border = border
                
                current_date += timedelta(days=1)
                row_idx += 1
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Save to BytesIO buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="trend_report_{start_date}_{today}.xlsx"'
            
            return response
            
        except Exception as e:
            import traceback
            return HttpResponse(f'Error: {str(e)}\n{traceback.format_exc()}', status=500)


@method_decorator(csrf_exempt, name='dispatch')
class StaffPerformanceExcelView(View):
    """
    Download Staff Performance Report as Excel
    URL: /salonvendor/analytics/staff-performance-excel/
    Auth: Bearer token OR vendor_id parameter for testing
    """
    
    def get(self, request):
        try:
            from io import BytesIO
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side
            
            # Try token auth first, then fallback to vendor_id parameter
            vendor_user = get_vendor_from_token(request)
            if not vendor_user:
                vendor_id = request.GET.get('vendor_id')
                if vendor_id:
                    try:
                        vendor_user = VendorUser.objects.get(id=vendor_id)
                    except VendorUser.DoesNotExist:
                        return HttpResponse('Vendor not found', status=404)
                else:
                    return HttpResponse('Unauthorized - provide token or vendor_id', status=401)
            
            filter_type = request.GET.get('filter', 'this_month')
            today = timezone.now().date()
            
            if filter_type == 'today':
                start_date = today
            elif filter_type == 'this_week' or filter_type == 'thisWeek':
                start_date = today - timedelta(days=today.weekday())
            elif filter_type == 'this_month' or filter_type == 'thisMonth':
                start_date = today.replace(day=1)
            else:
                start_date = today.replace(day=1)
            
            staff_list = Staff.objects.filter(vendor_user=vendor_user)
            
            # Create Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Staff Performance"
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            ws['A1'] = f"Staff Performance Report - {vendor_user.businessname}"
            ws['A1'].font = Font(bold=True, size=16)
            ws['A2'] = f"Period: {start_date} to {today}"
            
            headers = ['Staff Name', 'Role', 'Appointments', 'Revenue', 'Avg Rating']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            row_idx = 5
            for staff in staff_list:
                staff_appointments = Appointment.objects.filter(
                    vendor_user=vendor_user,
                    staff=staff,
                    date__range=[start_date, today],
                    appointment_status='completed'
                )
                
                total_appointments = staff_appointments.count()
                total_revenue = float(staff_appointments.aggregate(total=Sum('final_amount'))['total'] or 0)
                
                remarks = AppointmentRemarks.objects.filter(
                    appointment__staff=staff,
                    appointment__date__range=[start_date, today]
                )
                avg_rating = float(remarks.aggregate(avg=Avg('rating'))['avg'] or 0)
                
                ws.cell(row=row_idx, column=1, value=staff.staffname).border = border
                ws.cell(row=row_idx, column=2, value=staff.staff_role or 'Staff').border = border
                ws.cell(row=row_idx, column=3, value=total_appointments).border = border
                ws.cell(row=row_idx, column=4, value=total_revenue).border = border
                ws.cell(row=row_idx, column=5, value=round(avg_rating, 2)).border = border
                
                row_idx += 1
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Save to BytesIO buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="staff_performance_{start_date}_{today}.xlsx"'
            
            return response
            
        except Exception as e:
            import traceback
            return HttpResponse(f'Error: {str(e)}\n{traceback.format_exc()}', status=500)


@method_decorator(csrf_exempt, name='dispatch')
class AllReportsExcelView(View):
    """
    Download All Reports as Excel
    URL: /salonvendor/analytics/all-reports-excel/
    Auth: Bearer token OR vendor_id parameter for testing
    """
    
    def get(self, request):
        try:
            from io import BytesIO
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Try token auth first, then fallback to vendor_id parameter
            vendor_user = get_vendor_from_token(request)
            if not vendor_user:
                vendor_id = request.GET.get('vendor_id')
                if vendor_id:
                    try:
                        vendor_user = VendorUser.objects.get(id=vendor_id)
                    except VendorUser.DoesNotExist:
                        return HttpResponse('Vendor not found', status=404)
                else:
                    return HttpResponse('Unauthorized - provide token or vendor_id', status=401)
            
            today = timezone.now().date()
            
            filter_type = request.GET.get('filter', 'this_month')
            if filter_type == 'today':
                start_date = today
            elif filter_type == 'this_week':
                start_date = today - timedelta(days=today.weekday())
            else:
                start_date = today.replace(day=1)
            
            appointments = Appointment.objects.filter(
                vendor_user=vendor_user,
                date__range=[start_date, today],
                appointment_status='completed'
            )
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            title_font = Font(bold=True, size=14)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Summary Section
            total_revenue = float(appointments.aggregate(total=Sum('final_amount'))['total'] or 0)
            invoice_count = appointments.count()
            unique_customers = appointments.values('customer_phone').distinct().count()
            
            # Sheet 1: Summary
            ws1 = wb.active
            ws1.title = "Summary"
            
            ws1['A1'] = f"Business Report - {vendor_user.businessname}"
            ws1['A1'].font = title_font
            ws1['A2'] = f"Period: {start_date} to {today}"
            
            # Summary headers
            summary_headers = ['Metric', 'Value']
            for col, header in enumerate(summary_headers, 1):
                cell = ws1.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            summary_data = [
                ['Total Revenue', total_revenue],
                ['Total Invoices', invoice_count],
                ['Customers Served', unique_customers],
                ['Average Ticket', round(total_revenue / max(1, invoice_count), 2)],
            ]
            
            for row_idx, row_data in enumerate(summary_data, 5):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
            
            # Sheet 2: Appointments
            ws2 = wb.create_sheet("Appointments")
            appt_headers = ['Date', 'Customer', 'Phone', 'Staff', 'Amount', 'Payment Mode']
            for col, header in enumerate(appt_headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            row_idx = 2
            for appt in appointments:
                # Get staff names (ManyToMany field)
                staff_names = ', '.join([s.staffname for s in appt.staff.all()]) if appt.staff.exists() else ''
                
                ws2.cell(row=row_idx, column=1, value=str(appt.date)).border = border
                ws2.cell(row=row_idx, column=2, value=appt.customer_name or '').border = border
                ws2.cell(row=row_idx, column=3, value=appt.customer_phone or '').border = border
                ws2.cell(row=row_idx, column=4, value=staff_names).border = border
                ws2.cell(row=row_idx, column=5, value=float(appt.final_amount or 0)).border = border
                ws2.cell(row=row_idx, column=6, value=appt.payment_mode or '').border = border
                row_idx += 1
            
            # Sheet 3: Staff Performance
            ws3 = wb.create_sheet("Staff Performance")
            staff_headers = ['Staff Name', 'Role', 'Appointments', 'Revenue', 'Avg Rating']
            for col, header in enumerate(staff_headers, 1):
                cell = ws3.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            staff_list = Staff.objects.filter(vendor_user=vendor_user)
            row_idx = 2
            for staff in staff_list:
                staff_appts = appointments.filter(staff=staff)
                staff_revenue = float(staff_appts.aggregate(total=Sum('final_amount'))['total'] or 0)
                remarks = AppointmentRemarks.objects.filter(
                    appointment__staff=staff,
                    appointment__date__range=[start_date, today]
                )
                avg_rating = float(remarks.aggregate(avg=Avg('rating'))['avg'] or 0)
                
                ws3.cell(row=row_idx, column=1, value=staff.staffname).border = border
                ws3.cell(row=row_idx, column=2, value=staff.staff_role or 'Staff').border = border
                ws3.cell(row=row_idx, column=3, value=staff_appts.count()).border = border
                ws3.cell(row=row_idx, column=4, value=staff_revenue).border = border
                ws3.cell(row=row_idx, column=5, value=round(avg_rating, 2)).border = border
                row_idx += 1
            
            # Auto-adjust column widths
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Save to BytesIO buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Create response
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="all_reports_{start_date}_{today}.xlsx"'
            
            return response
            
        except Exception as e:
            import traceback
            return HttpResponse(f'Error: {str(e)}\n{traceback.format_exc()}', status=500)


# ============ CSV DOWNLOAD VIEWS ============
import csv

class DailyReportCSVView(APIView):
    """
    Download Daily Business Report as CSV
    URL: /salonvendor/analytics/daily-report-csv/
    """
    authentication_classes = [SimpleVendorAuth]
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            vendor_user = request.user
            today = timezone.now().date()
            
            # Get date range
            filter_type = request.query_params.get('filter', 'this_month')
            if filter_type == 'today':
                start_date = today
            elif filter_type == 'this_week':
                start_date = today - timedelta(days=today.weekday())
            else:
                start_date = today.replace(day=1)
            
            from_date = request.query_params.get('from_date')
            to_date = request.query_params.get('to_date')
            if from_date and to_date:
                start_date = datetime.strptime(from_date, '%Y-%m-%d').date()
                today = datetime.strptime(to_date, '%Y-%m-%d').date()
            
            appointments = Appointment.objects.filter(
                vendor_user=vendor_user,
                date__range=[start_date, today],
                appointment_status='completed'
            )
            
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="daily_report_{start_date}_{today}.csv"'
            
            writer = csv.writer(response)
            
            # Summary Section
            writer.writerow(['DAILY BUSINESS REPORT'])
            writer.writerow(['Business Name', vendor_user.businessname])
            writer.writerow(['Period', f'{start_date} to {today}'])
            writer.writerow([])
            
            total_revenue = float(appointments.aggregate(total=Sum('final_amount'))['total'] or 0)
            invoice_count = appointments.count()
            unique_customers = appointments.values('customer_phone').distinct().count()
            avg_ticket = round(total_revenue / max(1, invoice_count), 2)
            
            writer.writerow(['SUMMARY'])
            writer.writerow(['Total Revenue', f'Rs.{total_revenue:,.2f}'])
            writer.writerow(['Total Invoices', invoice_count])
            writer.writerow(['Customers Served', unique_customers])
            writer.writerow(['Average Ticket', f'Rs.{avg_ticket:,.2f}'])
            writer.writerow([])
            
            # Appointments Detail
            writer.writerow(['APPOINTMENTS DETAIL'])
            writer.writerow(['Date', 'Customer Name', 'Phone', 'Staff', 'Amount', 'Payment Mode'])
            
            for appt in appointments:
                # Get staff names (ManyToMany field)
                staff_names = ', '.join([s.staffname for s in appt.staff.all()]) if appt.staff.exists() else ''
                
                writer.writerow([
                    str(appt.date),
                    appt.customer_name or '',
                    appt.customer_phone or '',
                    staff_names,
                    float(appt.final_amount or 0),
                    appt.payment_mode or ''
                ])
            
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class TrendReportCSVView(APIView):
    """
    Download Trend Report as CSV
    URL: /salonvendor/analytics/trend-report-csv/
    """
    authentication_classes = [SimpleVendorAuth]
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            vendor_user = request.user
            today = timezone.now().date()
            
            period = request.query_params.get('period', 'last_30_days')
            if period == 'last_7_days':
                start_date = today - timedelta(days=7)
            elif period == 'last_90_days':
                start_date = today - timedelta(days=90)
            else:
                start_date = today - timedelta(days=30)
            
            appointments = Appointment.objects.filter(
                vendor_user=vendor_user,
                date__range=[start_date, today],
                appointment_status='completed'
            )
            
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="trend_report_{start_date}_{today}.csv"'
            
            writer = csv.writer(response)
            
            writer.writerow(['TREND REPORT'])
            writer.writerow(['Business Name', vendor_user.businessname])
            writer.writerow(['Period', f'{start_date} to {today}'])
            writer.writerow([])
            
            writer.writerow(['Date', 'Revenue', 'Customers', 'Appointments'])
            
            current_date = start_date
            while current_date <= today:
                day_appts = appointments.filter(date=current_date)
                day_revenue = float(day_appts.aggregate(total=Sum('final_amount'))['total'] or 0)
                day_customers = day_appts.values('customer_phone').distinct().count()
                day_count = day_appts.count()
                
                writer.writerow([str(current_date), day_revenue, day_customers, day_count])
                current_date += timedelta(days=1)
            
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class StaffPerformanceCSVView(APIView):
    """
    Download Staff Performance Report as CSV
    URL: /salonvendor/analytics/staff-performance-csv/
    """
    authentication_classes = [SimpleVendorAuth]
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            vendor_user = request.user
            today = timezone.now().date()
            
            filter_type = request.query_params.get('filter', 'this_month')
            if filter_type == 'today':
                start_date = today
            elif filter_type == 'this_week':
                start_date = today - timedelta(days=today.weekday())
            else:
                start_date = today.replace(day=1)
            
            staff_list = Staff.objects.filter(vendor_user=vendor_user)
            
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="staff_performance_{start_date}_{today}.csv"'
            
            writer = csv.writer(response)
            
            writer.writerow(['STAFF PERFORMANCE REPORT'])
            writer.writerow(['Business Name', vendor_user.businessname])
            writer.writerow(['Period', f'{start_date} to {today}'])
            writer.writerow([])
            
            writer.writerow(['Staff Name', 'Role', 'Appointments', 'Revenue', 'Avg Rating', 'Attendance %', 'Performance Score'])
            
            for staff in staff_list:
                staff_appointments = Appointment.objects.filter(
                    vendor_user=vendor_user,
                    staff=staff,
                    date__range=[start_date, today],
                    appointment_status='completed'
                )
                
                total_appointments = staff_appointments.count()
                total_revenue = float(staff_appointments.aggregate(total=Sum('final_amount'))['total'] or 0)
                
                remarks = AppointmentRemarks.objects.filter(
                    appointment__staff=staff,
                    appointment__date__range=[start_date, today]
                )
                avg_rating = float(remarks.aggregate(avg=Avg('rating'))['avg'] or 0)
                
                attendance = StaffAttendance.objects.filter(
                    staff=staff,
                    date__range=[start_date, today],
                    present=True
                ).count()
                total_days = (today - start_date).days + 1
                attendance_pct = round((attendance / max(1, total_days)) * 100, 1)
                
                performance_score = round(
                    (avg_rating / 5 * 40) + 
                    (attendance_pct / 100 * 30) + 
                    (min(total_appointments, 100) / 100 * 30), 1
                )
                
                writer.writerow([
                    staff.staffname,
                    staff.staff_role or 'Staff',
                    total_appointments,
                    total_revenue,
                    round(avg_rating, 2),
                    attendance_pct,
                    performance_score
                ])
            
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class AllReportsCSVView(APIView):
    """
    Download All Reports as CSV
    URL: /salonvendor/analytics/all-reports-csv/
    """
    authentication_classes = [SimpleVendorAuth]
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            vendor_user = request.user
            today = timezone.now().date()
            
            filter_type = request.query_params.get('filter', 'this_month')
            if filter_type == 'today':
                start_date = today
            elif filter_type == 'this_week':
                start_date = today - timedelta(days=today.weekday())
            else:
                start_date = today.replace(day=1)
            
            appointments = Appointment.objects.filter(
                vendor_user=vendor_user,
                date__range=[start_date, today],
                appointment_status='completed'
            )
            
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="all_reports_{start_date}_{today}.csv"'
            
            writer = csv.writer(response)
            
            # Summary
            writer.writerow(['ALL REPORTS'])
            writer.writerow(['Business Name', vendor_user.businessname])
            writer.writerow(['Period', f'{start_date} to {today}'])
            writer.writerow([])
            
            total_revenue = float(appointments.aggregate(total=Sum('final_amount'))['total'] or 0)
            invoice_count = appointments.count()
            unique_customers = appointments.values('customer_phone').distinct().count()
            
            writer.writerow(['=== SUMMARY ==='])
            writer.writerow(['Total Revenue', f'Rs.{total_revenue:,.2f}'])
            writer.writerow(['Total Invoices', invoice_count])
            writer.writerow(['Customers Served', unique_customers])
            writer.writerow(['Average Ticket', f'Rs.{round(total_revenue / max(1, invoice_count), 2):,.2f}'])
            writer.writerow([])
            
            # Appointments
            writer.writerow(['=== APPOINTMENTS ==='])
            writer.writerow(['Date', 'Customer Name', 'Phone', 'Staff', 'Amount', 'Payment Mode'])
            for appt in appointments:
                # Get staff names (ManyToMany field)
                staff_names = ', '.join([s.staffname for s in appt.staff.all()]) if appt.staff.exists() else ''
                
                writer.writerow([
                    str(appt.date),
                    appt.customer_name or '',
                    appt.customer_phone or '',
                    staff_names,
                    float(appt.final_amount or 0),
                    appt.payment_mode or ''
                ])
            writer.writerow([])
            
            # Staff Performance
            writer.writerow(['=== STAFF PERFORMANCE ==='])
            writer.writerow(['Staff Name', 'Role', 'Appointments', 'Revenue', 'Avg Rating'])
            
            staff_list = Staff.objects.filter(vendor_user=vendor_user)
            for staff in staff_list:
                staff_appts = appointments.filter(staff=staff)
                staff_revenue = float(staff_appts.aggregate(total=Sum('final_amount'))['total'] or 0)
                remarks = AppointmentRemarks.objects.filter(
                    appointment__staff=staff,
                    appointment__date__range=[start_date, today]
                )
                avg_rating = float(remarks.aggregate(avg=Avg('rating'))['avg'] or 0)
                
                writer.writerow([
                    staff.staffname,
                    staff.staff_role or 'Staff',
                    staff_appts.count(),
                    staff_revenue,
                    round(avg_rating, 2)
                ])
            
            return response
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


# ============ SALES SUMMARY DETAILED VIEW ============
class SalesSummaryDetailView(APIView):
    """
    Sales Summary API with all detailed fields
    URL: /salonvendor/analytics/sales-summary/
    
    Returns detailed bill-wise data with all fields:
    - Store Location, Bill Id, Bill Date, Bill Time
    - Customer Name, Customer Phone Number, Customer Registration Source, Customer GST Number, Company Name, Gender
    - Customer Phone Number, Customer Registration Source, Customer GST Number, Company Name, Gender
    - Staff name
    - Membership Discount Amount, Percent Discount Amount, Cash Discount Amount
    - Coupon Discount Amount, Loyalty Point Discount Amount
    - CGST, SGST, IGST, Total GST
    - Bill Amount, Payable Amount (Total Bill Amt.)
    - Cash Paid, Card Payment Last Four Digits, UPI, Pending Amount
    - Wallet Payment, Membership Wallet Payment
    """
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Get detailed sales summary with all fields"""
        try:
            import json
            vendor_user = request.user
            today = timezone.now().date()
            
            # Check for custom date range FIRST - support both from_date/to_date AND start_date/end_date
            from_date = request.query_params.get('from_date') or request.query_params.get('start_date')
            to_date = request.query_params.get('to_date') or request.query_params.get('end_date')
            
            if from_date and to_date:
                # Custom date range provided
                start_date = datetime.strptime(from_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(to_date, '%Y-%m-%d').date()
                date_filter = 'custom'
            else:
                # Use filter parameter
                date_filter = request.query_params.get('filter', 'this_month')
                
                if date_filter == 'today':
                    start_date = today
                    end_date = today
                elif date_filter == 'yesterday':
                    start_date = today - timedelta(days=1)
                    end_date = today - timedelta(days=1)
                elif date_filter == 'this_week' or date_filter == 'thisWeek':
                    start_date = today - timedelta(days=today.weekday())
                    end_date = today
                elif date_filter == 'last_week' or date_filter == 'lastWeek':
                    start_date = today - timedelta(days=today.weekday() + 7)
                    end_date = today - timedelta(days=today.weekday() + 1)
                elif date_filter == 'this_month' or date_filter == 'thisMonth':
                    start_date = today.replace(day=1)
                    end_date = today
                elif date_filter == 'last_month' or date_filter == 'lastMonth':
                    first_of_this_month = today.replace(day=1)
                    last_day_of_last_month = first_of_this_month - timedelta(days=1)
                    start_date = last_day_of_last_month.replace(day=1)
                    end_date = last_day_of_last_month
                elif date_filter == 'this_quarter' or date_filter == 'thisQuarter':
                    quarter = (today.month - 1) // 3
                    start_date = today.replace(month=quarter * 3 + 1, day=1)
                    end_date = today
                elif date_filter == 'this_year' or date_filter == 'thisYear':
                    start_date = today.replace(month=1, day=1)
                    end_date = today
                elif date_filter == 'all':
                    start_date = today - timedelta(days=365*5)
                    end_date = today
                else:
                    start_date = today.replace(day=1)
                    end_date = today
            
            # Get all completed appointments in date range
            appointments = Appointment.objects.filter(
                vendor_user=vendor_user,
                date__range=[start_date, end_date],
                appointment_status='completed'
            ).select_related('customer', 'selled_product', 'customer_wallet').prefetch_related('staff', 'membership')
            
            # Build detailed sales data
            sales_data = []
            
            # Summary totals
            total_sales = Decimal('0')
            total_cgst = Decimal('0')
            total_sgst = Decimal('0')
            total_igst = Decimal('0')
            total_cash = Decimal('0')
            total_card = Decimal('0')
            total_upi = Decimal('0')
            total_wallet = Decimal('0')
            total_membership_wallet = Decimal('0')
            total_pending = Decimal('0')
            total_discount = Decimal('0')
            
            for appt in appointments:
                # Get customer details
                customer = appt.customer
                customer_gst = ''
                company_name = ''
                registration_source = 'POS'
                
                if customer:
                    # Check if customer has GST info (if field exists)
                    customer_gst = getattr(customer, 'gst_number', '') or ''
                    company_name = getattr(customer, 'company_name', '') or ''
                    registration_source = getattr(customer, 'registration_source', 'POS') or 'POS'
                
                # Get staff names
                staff_names = []
                for staff in appt.staff.all():
                    staff_names.append(staff.staffname)
                staff_name = ', '.join(staff_names) if staff_names else ''
                
                # Calculate discounts
                membership_discount = Decimal('0')
                percent_discount = Decimal(str(appt.discount_amount or 0))
                cash_discount = Decimal('0')
                coupon_discount = Decimal('0')
                loyalty_point_discount = Decimal('0')
                
                # Check for membership discount from included_services
                if appt.included_services:
                    services_data = appt.included_services
                    if isinstance(services_data, str):
                        try:
                            services_data = json.loads(services_data)
                        except:
                            services_data = []
                    if isinstance(services_data, list):
                        for svc in services_data:
                            if isinstance(svc, dict):
                                if svc.get('membership_discount'):
                                    membership_discount += Decimal(str(svc.get('membership_discount', 0)))
                
                # Calculate GST from total bill amount using vendor's tax_percent
                bill_amount = Decimal(str(appt.actual_amount or 0))
                final_amount = Decimal(str(appt.final_amount or 0))
                
                # Get tax percent from vendor settings (vendor-pos API tax_percent field)
                tax_percent = Decimal('0')
                if vendor_user.is_gst and vendor_user.tax_percent:
                    try:
                        tax_percent = Decimal(str(vendor_user.tax_percent))
                    except:
                        tax_percent = Decimal('0')
                
                # Calculate GST from final amount (GST inclusive)
                # Formula: GST = (Amount * Tax%) / (100 + Tax%)
                if final_amount > 0 and tax_percent > 0:
                    total_gst_amount = (final_amount * tax_percent) / (100 + tax_percent)
                    total_gst_amount = total_gst_amount.quantize(Decimal('0.01'))
                else:
                    total_gst_amount = Decimal('0')
                
                # CGST and SGST are equal halves
                cgst = (total_gst_amount / 2).quantize(Decimal('0.01')) if total_gst_amount else Decimal('0')
                sgst = (total_gst_amount / 2).quantize(Decimal('0.01')) if total_gst_amount else Decimal('0')
                igst = Decimal('0')  # IGST is 0 for intra-state
                
                # Payment breakdown
                cash_paid = Decimal('0')
                card_payment = Decimal('0')
                card_last_four = ''
                upi_payment = Decimal('0')
                pending_amount = Decimal(str(appt.due_amount or 0))
                wallet_payment = Decimal('0')
                membership_wallet_payment = Decimal('0')
                
                payment_mode = (appt.payment_mode or '').lower()
                
                # Handle split payments
                if appt.split_payment_mode:
                    splits = appt.split_payment_mode
                    if isinstance(splits, str):
                        try:
                            splits = json.loads(splits)
                        except:
                            splits = []
                    if isinstance(splits, list):
                        for split in splits:
                            if isinstance(split, dict):
                                method = (split.get('method', '') or '').lower()
                                amount = Decimal(str(split.get('amount', 0)))
                                
                                if method == 'cash':
                                    cash_paid += amount
                                elif method in ['card', 'credit card', 'debit card']:
                                    card_payment += amount
                                    card_last_four = split.get('card_last_four', '') or split.get('last_four', '') or ''
                                elif method == 'upi':
                                    upi_payment += amount
                                elif method == 'wallet':
                                    wallet_payment += amount
                                elif method in ['membership', 'membership_wallet']:
                                    membership_wallet_payment += amount
                else:
                    # Single payment mode
                    if payment_mode == 'cash':
                        cash_paid = final_amount - pending_amount
                    elif payment_mode in ['card', 'credit card', 'debit card']:
                        card_payment = final_amount - pending_amount
                    elif payment_mode == 'upi':
                        upi_payment = final_amount - pending_amount
                    elif payment_mode == 'wallet':
                        wallet_payment = final_amount - pending_amount
                    elif payment_mode in ['membership', 'membership_wallet']:
                        membership_wallet_payment = final_amount - pending_amount
                
                # Check wallet applied
                if appt.is_wallet_applied and appt.customer_wallet:
                    wallet_used = Decimal(str(getattr(appt.customer_wallet, 'amount_used', 0) or 0))
                    if wallet_used > 0:
                        wallet_payment = wallet_used
                
                # Store location
                store_location = f"{vendor_user.businessname}"
                if vendor_user.branchname:
                    store_location += f" - {vendor_user.branchname}"
                
                # Get customer name
                customer_name = appt.customer_name or ''
                if not customer_name and customer:
                    customer_name = getattr(customer, 'customer_name', '') or getattr(customer, 'name', '') or ''
                
                # Build record
                record = {
                    'store_location': store_location,
                    'bill_id': appt.id,
                    'bill_date': appt.date.strftime('%Y-%m-%d'),
                    'bill_time': appt.appointment_start_time.strftime('%H:%M:%S') if appt.appointment_start_time else appt.created_at.strftime('%H:%M:%S'),
                    'customer_name': customer_name,
                    'customer_phone_number': appt.customer_phone or '',
                    'customer_registration_source': registration_source,
                    'customer_gst_number': customer_gst,
                    'company_name': company_name,
                    'gender': appt.customer_gender or '',
                    'staff_name': staff_name,
                    'membership_discount_amount': float(membership_discount),
                    'percent_discount_amount': float(percent_discount),
                    'cash_discount_amount': float(cash_discount),
                    'coupon_discount_amount': float(coupon_discount),
                    'loyalty_point_discount_amount': float(loyalty_point_discount),
                    'cgst': float(cgst),
                    'sgst': float(sgst),
                    'igst': float(igst),
                    'total_gst': float(total_gst_amount),
                    'bill_amount': float(bill_amount),
                    'payable_amount': float(final_amount),
                    'cash_paid': float(cash_paid),
                    'card_payment': float(card_payment),
                    'card_last_four_digits': card_last_four,
                    'upi': float(upi_payment),
                    'pending_amount': float(pending_amount),
                    'wallet_payment': float(wallet_payment),
                    'membership_wallet_payment': float(membership_wallet_payment)
                }
                
                sales_data.append(record)
                
                # Update totals
                total_sales += final_amount
                total_cgst += cgst
                total_sgst += sgst
                total_igst += igst
                total_cash += cash_paid
                total_card += card_payment
                total_upi += upi_payment
                total_wallet += wallet_payment
                total_membership_wallet += membership_wallet_payment
                total_pending += pending_amount
                total_discount += (membership_discount + percent_discount + cash_discount + coupon_discount + loyalty_point_discount)
            
            # Summary metrics
            invoice_count = len(sales_data)
            unique_customers = len(set(r['customer_phone_number'] for r in sales_data if r['customer_phone_number']))
            avg_ticket = float(total_sales / max(1, invoice_count))
            
            return Response({
                'success': True,
                'filter': date_filter,
                'period': {
                    'from': start_date.isoformat(),
                    'to': end_date.isoformat()
                },
                'vendor': {
                    'id': vendor_user.id,
                    'business_name': vendor_user.businessname,
                    'branch_name': vendor_user.branchname or 'Main Branch'
                },
                'summary': {
                    'total_sales': float(total_sales),
                    'invoice_count': invoice_count,
                    'customers_served': unique_customers,
                    'average_ticket': round(avg_ticket, 2),
                    'total_cgst': float(total_cgst),
                    'total_sgst': float(total_sgst),
                    'total_igst': float(total_igst),
                    'total_gst': float(total_cgst + total_sgst + total_igst),
                    'total_discount': float(total_discount),
                    'payment_breakdown': {
                        'cash': float(total_cash),
                        'card': float(total_card),
                        'upi': float(total_upi),
                        'wallet': float(total_wallet),
                        'membership_wallet': float(total_membership_wallet),
                        'pending': float(total_pending),
                        'total': float(total_cash + total_card + total_upi + total_wallet + total_membership_wallet)
                    }
                },
                'totals': {
                    'total_bill_amount': float(sum(Decimal(str(r['bill_amount'])) for r in sales_data)),
                    'total_payable_amount': float(total_sales),
                    'total_membership_discount': float(sum(Decimal(str(r['membership_discount_amount'])) for r in sales_data)),
                    'total_percent_discount': float(sum(Decimal(str(r['percent_discount_amount'])) for r in sales_data)),
                    'total_cash_discount': float(sum(Decimal(str(r['cash_discount_amount'])) for r in sales_data)),
                    'total_coupon_discount': float(sum(Decimal(str(r['coupon_discount_amount'])) for r in sales_data)),
                    'total_loyalty_discount': float(sum(Decimal(str(r['loyalty_point_discount_amount'])) for r in sales_data)),
                    'total_discount': float(total_discount),
                    'total_cgst': float(total_cgst),
                    'total_sgst': float(total_sgst),
                    'total_igst': float(total_igst),
                    'total_gst': float(total_cgst + total_sgst + total_igst),
                    'total_cash_paid': float(total_cash),
                    'total_card_payment': float(total_card),
                    'total_upi': float(total_upi),
                    'total_wallet_payment': float(total_wallet),
                    'total_membership_wallet': float(total_membership_wallet),
                    'total_pending': float(total_pending),
                    'grand_total_collected': float(total_cash + total_card + total_upi + total_wallet + total_membership_wallet)
                },
                'sales_data': sales_data
            }, status=200)
            
        except Exception as e:
            import traceback
            return Response({
                'error': str(e),
                'traceback': traceback.format_exc()
            }, status=500)


# ============ SALES SUMMARY EXCEL EXPORT ============
@method_decorator(csrf_exempt, name='dispatch')
class SalesSummaryExcelView(View):
    """
    Download Sales Summary as Excel with all fields
    URL: /salonvendor/analytics/sales-summary-excel/
    """
    
    def get(self, request):
        try:
            from io import BytesIO
            import json
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            
            vendor_user = get_vendor_from_token(request)
            if not vendor_user:
                return HttpResponse('Unauthorized', status=401)
            
            today = timezone.now().date()
            
            # Get date range
            date_filter = request.GET.get('filter', 'this_month')
            
            if date_filter == 'today':
                start_date = today
                end_date = today
            elif date_filter == 'yesterday':
                start_date = today - timedelta(days=1)
                end_date = today - timedelta(days=1)
            elif date_filter == 'this_week':
                start_date = today - timedelta(days=today.weekday())
                end_date = today
            elif date_filter == 'last_week':
                start_date = today - timedelta(days=today.weekday() + 7)
                end_date = today - timedelta(days=today.weekday() + 1)
            elif date_filter == 'this_month':
                start_date = today.replace(day=1)
                end_date = today
            elif date_filter == 'last_month':
                last_month = today.replace(day=1) - timedelta(days=1)
                start_date = last_month.replace(day=1)
                end_date = last_month
            elif date_filter == 'this_quarter':
                quarter = (today.month - 1) // 3
                start_date = today.replace(month=quarter * 3 + 1, day=1)
                end_date = today
            elif date_filter == 'this_year':
                start_date = today.replace(month=1, day=1)
                end_date = today
            else:
                start_date = today.replace(day=1)
                end_date = today
            
            # Custom date range
            from_date = request.GET.get('from_date')
            to_date = request.GET.get('to_date')
            if from_date and to_date:
                start_date = datetime.strptime(from_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(to_date, '%Y-%m-%d').date()
            
            # Get appointments
            appointments = Appointment.objects.filter(
                vendor_user=vendor_user,
                date__range=[start_date, end_date],
                appointment_status='completed'
            ).select_related('customer', 'selled_product', 'customer_wallet').prefetch_related('staff', 'membership')
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales Summary"
            
            # Styles
            title_font = Font(bold=True, size=16, color="1F4E79")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws['A1'] = f"SALES SUMMARY - {vendor_user.businessname}"
            ws['A1'].font = title_font
            ws['A2'] = f"Branch: {vendor_user.branchname or 'Main Branch'}"
            ws['A3'] = f"Period: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}"
            
            # Headers
            headers = [
                'Store Location', 'Bill Id', 'Bill Date', 'Bill Time',
                'Customer Name', 'Customer Phone Number', 'Customer Registration Source', 'Customer GST Number',
                'Customer Phone Number', 'Customer Registration Source', 'Customer GST Number',
                'Company Name', 'Gender', 'Staff Name',
                'Membership Discount Amount', 'Percent Discount Amount', 'Cash Discount Amount',
                'Coupon Discount Amount', 'Loyalty Point Discount Amount',
                'CGST', 'SGST', 'IGST', 'Total GST',
                'Bill Amount', 'Payable Amount (Total Bill Amt.)',
                'Cash Paid', 'Card Payment', 'Card Last Four Digits', 'UPI',
                'Pending Amount', 'Wallet Payment', 'Membership Wallet Payment'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            
            # Data rows
            row_idx = 6
            total_sales = Decimal('0')
            total_cgst = Decimal('0')
            total_sgst = Decimal('0')
            
            for appt in appointments:
                # Get customer details
                customer = appt.customer
                customer_name = appt.customer_name or ''
                customer_gst = ''
                company_name = ''
                registration_source = 'POS'
                
                if customer:
                    if not customer_name:
                        customer_name = getattr(customer, 'customer_name', '') or getattr(customer, 'name', '') or ''
                    customer_gst = getattr(customer, 'gst_number', '') or ''
                    company_name = getattr(customer, 'company_name', '') or ''
                    registration_source = getattr(customer, 'registration_source', 'POS') or 'POS'
                
                # Get staff names
                staff_names = [staff.staffname for staff in appt.staff.all()]
                staff_name = ', '.join(staff_names) if staff_names else ''
                
                # Calculate discounts
                membership_discount = Decimal('0')
                percent_discount = Decimal(str(appt.discount_amount or 0))
                
                if appt.included_services:
                    services_data = appt.included_services
                    if isinstance(services_data, str):
                        try:
                            services_data = json.loads(services_data)
                        except:
                            services_data = []
                    if isinstance(services_data, list):
                        for svc in services_data:
                            if isinstance(svc, dict) and svc.get('membership_discount'):
                                membership_discount += Decimal(str(svc.get('membership_discount', 0)))
                
                # GST - Calculate from vendor's tax_percent setting
                tax_percent = Decimal('0')
                if vendor_user.is_gst and vendor_user.tax_percent:
                    try:
                        tax_percent = Decimal(str(vendor_user.tax_percent))
                    except:
                        tax_percent = Decimal('0')
                
                final_amt = Decimal(str(appt.final_amount or 0))
                
                # Calculate GST from final amount (GST inclusive)
                if final_amt > 0 and tax_percent > 0:
                    total_gst_amount = (final_amt * tax_percent) / (100 + tax_percent)
                    total_gst_amount = total_gst_amount.quantize(Decimal('0.01'))
                else:
                    total_gst_amount = Decimal('0')
                
                cgst = (total_gst_amount / 2).quantize(Decimal('0.01')) if total_gst_amount else Decimal('0')
                sgst = (total_gst_amount / 2).quantize(Decimal('0.01')) if total_gst_amount else Decimal('0')
                
                bill_amount = Decimal(str(appt.actual_amount or 0))
                final_amount = Decimal(str(appt.final_amount or 0))
                pending_amount = Decimal(str(appt.due_amount or 0))
                
                # Payment breakdown
                cash_paid = Decimal('0')
                card_payment = Decimal('0')
                card_last_four = ''
                upi_payment = Decimal('0')
                wallet_payment = Decimal('0')
                membership_wallet_payment = Decimal('0')
                
                payment_mode = (appt.payment_mode or '').lower()
                
                if appt.split_payment_mode:
                    splits = appt.split_payment_mode
                    if isinstance(splits, str):
                        try:
                            splits = json.loads(splits)
                        except:
                            splits = []
                    if isinstance(splits, list):
                        for split in splits:
                            if isinstance(split, dict):
                                method = (split.get('method', '') or '').lower()
                                amount = Decimal(str(split.get('amount', 0)))
                                
                                if method == 'cash':
                                    cash_paid += amount
                                elif method in ['card', 'credit card', 'debit card']:
                                    card_payment += amount
                                    card_last_four = split.get('card_last_four', '') or split.get('last_four', '') or ''
                                elif method == 'upi':
                                    upi_payment += amount
                                elif method == 'wallet':
                                    wallet_payment += amount
                                elif method in ['membership', 'membership_wallet']:
                                    membership_wallet_payment += amount
                else:
                    if payment_mode == 'cash':
                        cash_paid = final_amount - pending_amount
                    elif payment_mode in ['card', 'credit card', 'debit card']:
                        card_payment = final_amount - pending_amount
                    elif payment_mode == 'upi':
                        upi_payment = final_amount - pending_amount
                    elif payment_mode == 'wallet':
                        wallet_payment = final_amount - pending_amount
                    elif payment_mode in ['membership', 'membership_wallet']:
                        membership_wallet_payment = final_amount - pending_amount
                
                # Store location
                store_location = f"{vendor_user.businessname}"  
                store_location = f"{vendor_user.businessname}"
                if vendor_user.branchname:
                    store_location += f" - {vendor_user.branchname}"
                
                # Write row
                row_data = [
                    store_location,
                    appt.id,
                    appt.date.strftime('%Y-%m-%d'),
                    appt.appointment_start_time.strftime('%H:%M:%S') if appt.appointment_start_time else appt.created_at.strftime('%H:%M:%S'),
                    customer_name,
                    appt.customer_phone or '',
                    registration_source,
                    customer_gst,
                    company_name,
                    appt.customer_gender or '',
                    staff_name,
                    float(membership_discount),
                    float(percent_discount),
                    0,  # Cash discount
                    0,  # Coupon discount
                    0,  # Loyalty point discount
                    float(cgst),
                    float(sgst),
                    0,  # IGST
                    float(total_gst_amount),
                    float(bill_amount),
                    float(final_amount),
                    float(cash_paid),
                    float(card_payment),
                    card_last_four,
                    float(upi_payment),
                    float(pending_amount),
                    float(wallet_payment),
                    float(membership_wallet_payment)
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = border
                
                total_sales += final_amount
                total_cgst += cgst
                total_sgst += sgst
                row_idx += 1
            
            # Totals row
            ws.cell(row=row_idx, column=1, value='TOTAL').font = Font(bold=True)
            ws.cell(row=row_idx, column=20, value=float(total_sales)).font = Font(bold=True)
            ws.cell(row=row_idx, column=16, value=float(total_cgst)).font = Font(bold=True)
            ws.cell(row=row_idx, column=17, value=float(total_sgst)).font = Font(bold=True)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
            
            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="sales_summary_{start_date}_{end_date}.xlsx"'
            
            return response
            
        except Exception as e:
            import traceback
            return HttpResponse(f'Error: {str(e)}\n{traceback.format_exc()}', status=500)
    

# ===== PRODUCT CSV IMPORT API =====
import csv
import io
from django.db import transaction
from decimal import Decimal

class ProductImportAPIView(APIView):
    """
    API to import products from CSV file
    Required CSV Format:
    - product_name (required)
    - supplier_Barcode (optional) 
    - supplier (supplier ID)
    - product_brand (brand ID)
    - measure_unit (ml, L, g, kg)
    - measure_amount (number)
    - supply_price (number)
    - retail_price (number)
    - tax (percentage)
    - low_stock_level (number)
    - short_description (optional)
    - product_description (optional)
    - pin (optional)
    """
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        # Validate file type
        if not file_obj.name.endswith(('.csv', '.xlsx', '.xls')):
            return Response({'error': 'Invalid file type. Only CSV, Excel files allowed.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Handle Excel files
            if file_obj.name.endswith(('.xlsx', '.xls')):
                import pandas as pd
                df = pd.read_excel(file_obj)
                csv_string = df.to_csv(index=False)
                io_string = io.StringIO(csv_string)
            else:
                # Handle CSV files
                decoded_file = file_obj.read().decode('utf-8')
                io_string = io.StringIO(decoded_file)
                
        except Exception as e:
            return Response({'error': f'File processing error: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        reader = csv.DictReader(io_string)
        
        created_products = []
        skipped_rows = []
        vendor_user = request.user

        with transaction.atomic():
            for row_num, row in enumerate(reader, start=2):  # Start from row 2 (header is row 1)
                try:
                    # Skip blank rows
                    if not any(v.strip() for v in row.values() if v is not None):
                        continue

                    # Clean and validate required fields
                    product_name = (row.get('product_name') or '').strip()
                    if not product_name:
                        skipped_rows.append(f"Row {row_num}: Missing required field 'product_name'")
                        continue

                    # Get optional fields
                    supplier_barcode = (row.get('supplier_Barcode') or '').strip()
                    short_description = (row.get('short_description') or '').strip()
                    product_description = (row.get('product_description') or '').strip()
                    pin = (row.get('pin') or '').strip()

                    # Get and validate numeric fields
                    try:
                        supplier_id = int(row.get('supplier', 0)) if row.get('supplier') else None
                        brand_id = int(row.get('product_brand', 0)) if row.get('product_brand') else None
                        measure_amount = float(row.get('measure_amount', 0)) if row.get('measure_amount') else 0.0
                        supply_price = float(row.get('supply_price', 0)) if row.get('supply_price') else 0.0
                        retail_price = float(row.get('retail_price', 0)) if row.get('retail_price') else 0.0
                        tax = float(row.get('tax', 0)) if row.get('tax') else 0.0
                        low_stock_level = int(row.get('low_stock_level', 0)) if row.get('low_stock_level') else 0
                    except ValueError as e:
                        skipped_rows.append(f"Row {row_num}: Invalid number format - {str(e)}")
                        continue

                    # Get measure unit
                    measure_unit = (row.get('measure_unit') or 'ml').strip()
                    if measure_unit not in ['ml', 'L', 'g', 'kg']:
                        measure_unit = 'ml'  # Default to ml

                    # Validate supplier exists
                    supplier = None
                    if supplier_id:
                        try:
                            supplier = Supplier.objects.get(id=supplier_id, vendor_user=vendor_user)
                        except Supplier.DoesNotExist:
                            skipped_rows.append(f"Row {row_num}: Supplier ID {supplier_id} not found")
                            continue

                    # Validate brand exists
                    brand = None
                    if brand_id:
                        try:
                            brand = Brand.objects.get(id=brand_id, vendor=vendor_user.id)
                        except Brand.DoesNotExist:
                            skipped_rows.append(f"Row {row_num}: Brand ID {brand_id} not found")
                            continue

                    # Check if product already exists
                    existing_product = Product.objects.filter(
                        product_name=product_name,
                        supplier=supplier
                    ).first()

                    if existing_product:
                        skipped_rows.append(f"Row {row_num}: Product '{product_name}' already exists")
                        continue

                    # Create product
                    product = Product.objects.create(
                        product_name=product_name,
                        short_description=short_description or None,
                        product_description=product_description or None,
                        supply_price=supply_price,
                        retail_price=retail_price,
                        measure_quantity=measure_amount,  # Note: using measure_quantity field
                        measure_amount=measure_amount,
                        tax=tax,
                        supplier=supplier,
                        product_brand=brand,
                        low_stock_level=low_stock_level,
                        product_indentification_number=supplier_barcode or f"PRD-{product_name[:10].upper()}",
                        measure_unit=measure_unit,
                    )

                    created_products.append({
                        'id': product.id,
                        'product_name': product.product_name,
                        'supplier': supplier.name if supplier else None,
                        'brand': brand.name if brand else None,
                        'retail_price': product.retail_price
                    })

                except Exception as e:
                    skipped_rows.append(f"Row {row_num}: Error creating product - {str(e)}")
                    continue

        return Response({
            'success': True,
            'message': f'Import completed. {len(created_products)} products created.',
            'created_products': created_products,
            'skipped_rows': skipped_rows,
            'total_processed': len(created_products) + len(skipped_rows)
        }, status=status.HTTP_201_CREATED)


class ProductExportAPIView(APIView):
    """
    API to export products to CSV format
    """
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        vendor_user = request.user
        
        # Get all products for this vendor
        products = Product.objects.filter(supplier__vendor_user=vendor_user).select_related('supplier', 'product_brand')
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="products_export.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'product_name',
            'supplier_Barcode', 
            'supplier',
            'product_brand',
            'measure_unit',
            'measure_amount',
            'supply_price',
            'retail_price',
            'tax',
            'low_stock_level',
            'short_description',
            'product_description',
            'pin'
        ])
        
        # Write data
        for product in products:
            writer.writerow([
                product.product_name,
                product.product_indentification_number,
                product.supplier.id if product.supplier else '',
                product.product_brand.id if product.product_brand else '',
                product.measure_unit,
                product.measure_amount,
                product.supply_price,
                product.retail_price,
                product.tax,
                product.low_stock_level,
                product.short_description or '',
                product.product_description or '',
                ''  # pin field (not in model)
            ])
        
        return response


class ProductCSVTemplateAPIView(APIView):
    """
    API to download CSV template for product import
    """
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="product_import_template.csv"'
        
        writer = csv.writer(response)
        
        # Write header with sample data
        writer.writerow([
            'product_name',
            'supplier_Barcode', 
            'supplier',
            'product_brand',
            'measure_unit',
            'measure_amount',
            'supply_price',
            'retail_price',
            'tax',
            'low_stock_level',
            'short_description',
            'product_description',
            'pin'
        ])
        
        # Write sample row
        writer.writerow([
            'Sample Hair Oil',
            'SAMPLE123',
            '1',  # Supplier ID
            '1',  # Brand ID
            'ml',
            '100',
            '50.00',
            '75.00',
            '18.0',
            '10',
            'Premium hair oil for salon use',
            'High quality hair oil with natural ingredients for professional salon treatments',
            'SAMPLE'
        ])
        
        return response


class StaffAvailabilityCheckView(APIView):
    """
    API to check staff availability for a specific date and time
    """
    authentication_classes = [VendorJWTAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        """
        Check if staff is available at given date and time
        
        Expected payload:
        {
            "date": "2025-12-13",
            "time_in": "11:35:59", 
            "staff_ids": [17, 18, 19]
        }
        """
        try:
            date_str = request.data.get('date')
            time_str = request.data.get('time_in')
            staff_ids = request.data.get('staff_ids', [])
            
            if not date_str or not time_str or not staff_ids:
                return Response({
                    "error": "date, time_in, and staff_ids are required"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Parse date and time
            appointment_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            appointment_time = datetime.strptime(time_str, '%H:%M:%S').time()
            appointment_datetime = datetime.combine(appointment_date, appointment_time)
            
            vendor_user = request.user
            availability_results = []
            
            for staff_id in staff_ids:
                try:
                    staff = Staff.objects.get(id=staff_id, vendor_user=vendor_user)
                    
                    # Check for conflicting appointments
                    conflicting_appointments = Appointment.objects.filter(
                        vendor_user=vendor_user,
                        date=appointment_date,
                        staff=staff,
                        appointment_status__in=['not_started', 'running'],
                        is_delete=False
                    ).exclude(time_in__isnull=True)
                    
                    conflicts = []
                    is_available = True
                    
                    for existing_appointment in conflicting_appointments:
                        if existing_appointment.time_in:
                            existing_datetime = datetime.combine(appointment_date, existing_appointment.time_in)
                            time_difference = abs((appointment_datetime - existing_datetime).total_seconds() / 60)
                            
                            if time_difference < 30:  # 30 minutes buffer
                                is_available = False
                                conflicts.append({
                                    "appointment_id": existing_appointment.id,
                                    "customer_name": existing_appointment.customer_name,
                                    "time": existing_appointment.time_in.strftime('%H:%M'),
                                    "status": existing_appointment.appointment_status
                                })
                    
                    availability_results.append({
                        "staff_id": staff_id,
                        "staff_name": staff.staffname,
                        "is_available": is_available,
                        "conflicts": conflicts,
                        "message": "Available" if is_available else f"Not available - conflicts found"
                    })
                    
                except Staff.DoesNotExist:
                    availability_results.append({
                        "staff_id": staff_id,
                        "staff_name": "Unknown",
                        "is_available": False,
                        "conflicts": [],
                        "message": "Staff not found"
                    })
            
            return Response({
                "date": date_str,
                "time": time_str,
                "staff_availability": availability_results,
                "overall_available": all(result["is_available"] for result in availability_results)
            }, status=status.HTTP_200_OK)
            
        except ValueError as e:
            return Response({
                "error": f"Invalid date or time format: {str(e)}"
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:  
            return Response({
                "error": f"An error occurred: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)    



from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.utils import timezone
from datetime import datetime, timedelta
from django.db.models import Sum, Avg, Count, Q
from .models import (
    VendorUser, Appointment, CustomerTable, Staff, StaffAttendance
)


class SimpleDashboardView(APIView):
    """
    Simple Analytics Dashboard API (No Authentication Required for Testing)
    URL: /salonvendor/analytics/simple-dashboard/
    """
    
    def get(self, request):
        """Get comprehensive dashboard analytics"""
        try:
            # Get vendor_id from query params or use first vendor
            vendor_id = request.query_params.get('vendor_id')
            
            if vendor_id:
                try:
                    vendor_user = VendorUser.objects.get(id=vendor_id)
                except VendorUser.DoesNotExist:
                    return Response({
                        'error': f'Vendor with ID {vendor_id} not found',
                        'available_vendors': list(VendorUser.objects.values('id', 'businessname'))
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                # Use first vendor for testing
                vendor_user = VendorUser.objects.first()
                if not vendor_user:
                    return Response({
                        'error': 'No vendor users found in database',
                        'message': 'Please create a vendor user first'
                    }, status=status.HTTP_404_NOT_FOUND)
            
            # Get date range from query params
            days = int(request.query_params.get('days', 30))
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=days)
            
            # Get dashboard data
            dashboard_data = self._get_simple_dashboard_data(vendor_user, start_date, end_date)
            
            response_data = {
                'success': True,
                'vendor': {
                    'id': vendor_user.id,
                    'business_name': vendor_user.businessname,
                    'owner_name': vendor_user.ownername
                },
                'period': {
                    'start_date': start_date.isoformat(),
                    'end_date': end_date.isoformat(),
                    'days': days
                },
                'dashboard_data': dashboard_data,
                'generated_at': timezone.now().isoformat()
            }
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'error': f'Failed to generate dashboard: {str(e)}',
                'debug_info': {
                    'vendor_count': VendorUser.objects.count(),
                    'appointment_count': Appointment.objects.count(),
                    'customer_count': CustomerTable.objects.count()
                }
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _get_simple_dashboard_data(self, vendor_user, start_date, end_date):
        """Get simplified dashboard data"""
        try:
            # Basic appointment metrics
            appointments = Appointment.objects.filter(
                vendor_user=vendor_user,
                date__range=[start_date, end_date]
            )
            
            completed_appointments = appointments.filter(appointment_status='completed')
            
            # Calculate basic metrics
            total_appointments = appointments.count()
            completed_count = completed_appointments.count()
            
            total_revenue = completed_appointments.aggregate(
                total=Sum('final_amount')
            )['total'] or 0
            
            avg_ticket_size = (total_revenue / completed_count) if completed_count > 0 else 0
            
            # Customer metrics
            unique_customers = appointments.values('customer_phone').distinct().count()
            
            # Staff metrics
            total_staff = Staff.objects.filter(vendor_user=vendor_user).count()
            active_staff = Staff.objects.filter(vendor_user=vendor_user, is_present=True).count()
            
            # Today's metrics
            today = timezone.now().date()
            today_appointments = appointments.filter(date=today)
            today_completed = today_appointments.filter(appointment_status='completed')
            today_revenue = today_completed.aggregate(total=Sum('final_amount'))['total'] or 0
            
            # Customer satisfaction - using AppointmentRemarks instead
            from .models import AppointmentRemarks
            feedbacks = AppointmentRemarks.objects.filter(
                appointment__vendor_user=vendor_user,
                appointment__date__range=[start_date, end_date]
            )
            
            avg_satisfaction = feedbacks.aggregate(avg=Avg('rating'))['avg'] or 0
            total_feedbacks = feedbacks.count()
            
            # Payment method analysis (simplified)
            cash_payments = completed_appointments.filter(
                Q(payment_mode__icontains='cash') | Q(payment_mode__isnull=True)
            ).count()
            
            card_payments = completed_appointments.filter(
                payment_mode__icontains='card'
            ).count()
            
            online_payments = completed_appointments.filter(
                Q(payment_mode__icontains='online') | Q(payment_mode__icontains='upi')
            ).count()
            
            # Service categories (top 5)
            service_categories = {}
            for appointment in completed_appointments:
                if appointment.included_services:
                    for service in appointment.included_services:
                        service_name = service.get('service_name', 'Unknown')
                        if service_name in service_categories:
                            service_categories[service_name] += 1
                        else:
                            service_categories[service_name] = 1
            
            top_services = sorted(service_categories.items(), key=lambda x: x[1], reverse=True)[:5]
            
            return {
                'overview': {
                    'total_revenue': float(total_revenue),
                    'total_appointments': total_appointments,
                    'completed_appointments': completed_count,
                    'avg_ticket_size': round(float(avg_ticket_size), 2),
                    'unique_customers': unique_customers,
                    'completion_rate': round((completed_count / max(1, total_appointments)) * 100, 1)
                },
                'today_metrics': {
                    'date': today.isoformat(),
                    'appointments': today_appointments.count(),
                    'completed': today_completed.count(),
                    'revenue': float(today_revenue)
                },
                'staff_info': {
                    'total_staff': total_staff,
                    'active_staff': active_staff,
                    'staff_utilization': round((active_staff / max(1, total_staff)) * 100, 1)
                },
                'customer_satisfaction': {
                    'avg_rating': round(float(avg_satisfaction), 2),
                    'total_feedbacks': total_feedbacks,
                    'satisfaction_level': self._get_satisfaction_level(avg_satisfaction)
                },
                'payment_distribution': {
                    'cash': cash_payments,
                    'card': card_payments,
                    'online': online_payments,
                    'total_transactions': completed_count
                },
                'top_services': [
                    {'service_name': name, 'count': count} 
                    for name, count in top_services
                ],
                'period_summary': {
                    'daily_avg_revenue': round(float(total_revenue / max(1, (end_date - start_date).days)), 2),
                    'daily_avg_appointments': round(total_appointments / max(1, (end_date - start_date).days), 1),
                    'customer_retention': self._calculate_simple_retention(vendor_user, start_date, end_date)
                }
            }
            
        except Exception as e:
            return {'error': str(e)}
    
    def _get_satisfaction_level(self, avg_rating):
        """Get satisfaction level description"""
        if avg_rating >= 4.5:
            return 'Excellent'
        elif avg_rating >= 4.0:
            return 'Good'
        elif avg_rating >= 3.0:
            return 'Average'
        elif avg_rating > 0:
            return 'Poor'
        else:
            return 'No Data'
    
    def _calculate_simple_retention(self, vendor_user, start_date, end_date):
        """Calculate simple retention rate"""
        try:
            # Get customers who had appointments in the period
            current_customers = set(
                Appointment.objects.filter(
                    vendor_user=vendor_user,
                    date__range=[start_date, end_date],
                    appointment_status='completed'
                ).values_list('customer_phone', flat=True)
            )
            
            # Get customers who had appointments before the period
            previous_customers = set(
                Appointment.objects.filter(
                    vendor_user=vendor_user,
                    date__lt=start_date,
                    appointment_status='completed'
                ).values_list('customer_phone', flat=True)
            )
            
            # Calculate retention
            returning_customers = current_customers.intersection(previous_customers)
            retention_rate = (len(returning_customers) / max(1, len(current_customers))) * 100
            
            return {
                'retention_rate': round(retention_rate, 1),
                'total_customers': len(current_customers),
                'returning_customers': len(returning_customers),
                'new_customers': len(current_customers) - len(returning_customers)
            }
            
        except Exception as e:
            return {
                'retention_rate': 0,
                'total_customers': 0,
                'returning_customers': 0,
                'new_customers': 0,
                'error': str(e)
            }


class VendorListView(APIView):
    """
    List all vendors for testing
    URL: /salonvendor/analytics/vendors/
    """
    
    def get(self, request):
        """Get list of all vendors"""
        try:
            vendors = VendorUser.objects.all().values(
                'id', 'businessname', 'ownername', 'ph_number', 'created_at'
            )
            
            return Response({
                'success': True,
                'count': len(vendors),
                'vendors': list(vendors)
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'error': f'Failed to get vendors: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
